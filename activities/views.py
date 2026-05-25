from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse
import openpyxl
from .models import ActivityReport, ActivityIncomeExpense, ActivityEvidence, ActivityType
from accounts.models import AcademicYear


def _resolve_activity(slug_or_pk):
    """Lookup activity by slug (preferred) or pk (backward compat)."""
    if slug_or_pk.isdigit():
        return get_object_or_404(ActivityReport, pk=int(slug_or_pk))
    return get_object_or_404(ActivityReport, slug=slug_or_pk)


def _get_level_filter(user):
    """Return filter dict for level-based access."""
    if user.role == 'admin' or not user.assigned_level:
        return {}
    return {'pic__assigned_level': user.assigned_level}


def pic_required(view_func):
    def wrapper(request, *args, **kwargs):
        if request.user.role not in ['admin', 'pic_teacher']:
            messages.error(request, 'Akses ditolak.')
            return redirect('accounts:dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
def activity_list(request):
    level_filter = _get_level_filter(request.user)
    if request.user.role == 'pic_teacher':
        activities = ActivityReport.objects.filter(pic=request.user, **level_filter)
    elif request.user.role in ['admin', 'vp_activity', 'kepsek']:
        activities = ActivityReport.objects.filter(**level_filter)
    else:
        activities = ActivityReport.objects.none()
    return render(request, 'activities/activity_list.html', {'activities': activities})


@login_required
@pic_required
def create_activity(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date') or None
        google_drive_link1 = request.POST.get('google_drive_link1', '')
        google_drive_link2 = request.POST.get('google_drive_link2', '')
        invoice_number = request.POST.get('invoice_number', '')
        budget_real = request.POST.get('budget_real', 0)
        pic_notes = request.POST.get('pic_notes', '')
        activity_type_id = request.POST.get('activity_type')
        submit_action = request.POST.get('action', 'draft')

        academic_year = AcademicYear.objects.filter(is_active=True).first()
        activity = ActivityReport.objects.create(
            title=title,
            description=description,
            start_date=start_date,
            end_date=end_date,
            google_drive_link1=google_drive_link1,
            google_drive_link2=google_drive_link2,
            invoice_number=invoice_number,
            budget_real=budget_real,
            pic_notes=pic_notes,
            activity_type_id=activity_type_id or None,
            pic=request.user,
            academic_year=academic_year,
            status=submit_action,
            cover_image=request.FILES.get('cover_image'),
        )

        i = 0
        while request.POST.get(f'transaction_type_{i}'):
            ttype = request.POST[f'transaction_type_{i}']
            desc = request.POST.get(f'transaction_desc_{i}', '')
            amount = request.POST.get(f'transaction_amount_{i}', 0)
            if amount:
                ActivityIncomeExpense.objects.create(
                    activity=activity,
                    transaction_type=ttype,
                    description=desc,
                    amount=amount,
                    date=start_date,
                )
            i += 1

        for f in request.FILES.getlist('evidences'):
            ActivityEvidence.objects.create(
                activity=activity,
                file=f,
                description=request.POST.get('evidence_desc', ''),
            )

        msg = 'Laporan kegiatan berhasil disimpan sebagai draft.' if submit_action == 'draft' else 'Laporan berhasil diajukan ke PIC.'
        messages.success(request, msg)
        return redirect('activities:list')

    activity_types = ActivityType.objects.all()
    return render(request, 'activities/create_activity.html', {'activity_types': activity_types})


@login_required
@pic_required
def edit_activity(request, slug):
    activity = _resolve_activity(slug)
    if activity.pic != request.user:
        messages.error(request, 'Anda bukan PIC laporan ini.')
        return redirect('activities:list')
    if activity.status in ['approved', 'rejected']:
        messages.error(request, 'Laporan sudah final (disetujui/ditolak) tidak dapat diubah.')
        return redirect('activities:list')

    if request.method == 'POST':
        activity.title = request.POST.get('title')
        activity.description = request.POST.get('description')
        activity.start_date = request.POST.get('start_date')
        activity.end_date = request.POST.get('end_date') or None
        activity.google_drive_link1 = request.POST.get('google_drive_link1', '')
        activity.google_drive_link2 = request.POST.get('google_drive_link2', '')
        activity.invoice_number = request.POST.get('invoice_number', '')
        activity.budget_real = request.POST.get('budget_real', 0)
        activity.pic_notes = request.POST.get('pic_notes', '')
        activity.activity_type_id = request.POST.get('activity_type') or None
        if request.FILES.get('cover_image'):
            activity.cover_image = request.FILES['cover_image']
        submit_action = request.POST.get('action', activity.status)
        activity.status = submit_action
        activity.save()
        messages.success(request, 'Laporan berhasil diperbarui.')
        return redirect('activities:list')

    activity_types = ActivityType.objects.all()
    return render(request, 'activities/create_activity.html', {'activity': activity, 'activity_types': activity_types, 'editing': True})


@login_required
def activity_detail(request, slug):
    activity = _resolve_activity(slug)
    return render(request, 'activities/activity_detail.html', {'activity': activity})


@login_required
@pic_required
def publish_activity(request, slug):
    activity = _resolve_activity(slug)
    if activity.pic != request.user:
        messages.error(request, 'Akses ditolak.')
        return redirect('activities:list')
    activity.status = 'submitted'
    activity.save()
    messages.success(request, 'Laporan berhasil diajukan untuk review PIC.')
    return redirect('activities:list')


@login_required
def review_activity(request, slug):
    if request.user.role not in ['admin', 'kepsek', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')

    activity = _resolve_activity(slug)
    if activity.status not in ['pic_approved', 'submitted']:
        messages.error(request, 'Laporan belum siap direview.')
        return redirect('activities:list')

    if request.method == 'POST':
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')
        signature = request.POST.get('signature_data', '')
        if action == 'approved':
            activity.status = 'approved'
            activity.reviewed_by = request.user
            activity.reviewed_at = timezone.now()
            if notes:
                activity.feedback_notes = notes
                activity.feedback_by = request.user
                activity.feedback_at = timezone.now()
            messages.success(request, 'Laporan disetujui.')
        elif action == 'pic_approved':
            activity.status = 'pic_approved'
            activity.reviewed_by = request.user
            activity.reviewed_at = timezone.now()
            if signature:
                activity.pic_signature_data = signature
            if notes:
                activity.feedback_notes = notes
                activity.feedback_by = request.user
                activity.feedback_at = timezone.now()
            messages.success(request, 'Laporan disetujui dan siap review final.')
        elif action == 'revised':
            activity.status = 'revised'
            activity.rejection_notes = notes
            activity.feedback_notes = notes
            activity.feedback_by = request.user
            activity.feedback_at = timezone.now()
            messages.success(request, 'Laporan perlu direvisi.')
        elif action == 'rejected':
            activity.status = 'rejected'
            activity.rejection_notes = notes
            activity.feedback_notes = notes
            activity.feedback_by = request.user
            activity.feedback_at = timezone.now()
            activity.reviewed_by = request.user
            activity.reviewed_at = timezone.now()
            messages.error(request, 'Laporan ditolak.')
        activity.save()
        return redirect('activities:list')

    return render(request, 'activities/review_activity.html', {'activity': activity})


@login_required
def pic_review_activity(request, slug):
    if request.user.role not in ['admin', 'pic_teacher']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')

    activity = _resolve_activity(slug)
    if request.user.role == 'pic_teacher' and activity.pic != request.user:
        messages.error(request, 'Anda bukan PIC laporan ini.')
        return redirect('activities:list')
    if activity.status not in ['submitted', 'revised']:
        messages.error(request, 'Laporan tidak dalam status yang dapat direview.')
        return redirect('activities:list')

    if request.method == 'POST':
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')
        signature_data = request.POST.get('signature_data', '')

        if action == 'pic_approved':
            activity.status = 'pic_approved'
            activity.pic_reviewed_by = request.user
            activity.pic_reviewed_at = timezone.now()
            activity.pic_signature_data = signature_data
            if notes:
                activity.pic_notes = notes
            messages.success(request, 'Laporan disetujui dan diteruskan ke Kepala Sekolah.')
        elif action == 'revised':
            activity.status = 'revised'
            activity.rejection_notes = notes
            messages.success(request, 'Laporan perlu direvisi.')
        elif action == 'rejected':
            activity.status = 'rejected'
            activity.rejection_notes = notes
            activity.pic_reviewed_by = request.user
            activity.pic_reviewed_at = timezone.now()
            messages.success(request, 'Laporan ditolak.')
        activity.save()
        return redirect('activities:list')

    return render(request, 'activities/pic_review_activity.html', {'activity': activity})


@login_required
def activity_type_list(request):
    if request.user.role not in ['admin', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    types = ActivityType.objects.all()
    return render(request, 'activities/type_list.html', {'types': types})


@login_required
def create_activity_type(request):
    if request.user.role not in ['admin', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    if request.method == 'POST':
        ActivityType.objects.create(
            name=request.POST['name'],
            objectives=request.POST.get('objectives', ''),
            values_knowledge=request.POST.get('values_knowledge', ''),
            values_faith=request.POST.get('values_faith', ''),
            values_character=request.POST.get('values_character', ''),
            budget=request.POST.get('budget', 0),
            month=request.POST.get('month', ''),
            time_start=request.POST.get('time_start'),
            time_finish=request.POST.get('time_finish'),
            pic=request.POST.get('pic', ''),
            notes=request.POST.get('notes', ''),
            created_by=request.user,
        )
        messages.success(request, 'Jenis kegiatan berhasil ditambahkan.')
        return redirect('activities:type_list')
    return render(request, 'activities/type_form.html', {'action': 'Tambah'})


@login_required
def edit_activity_type(request, pk):
    if request.user.role not in ['admin', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    obj = get_object_or_404(ActivityType, id=pk)
    if request.method == 'POST':
        obj.name = request.POST['name']
        obj.objectives = request.POST.get('objectives', '')
        obj.values_knowledge = request.POST.get('values_knowledge', '')
        obj.values_faith = request.POST.get('values_faith', '')
        obj.values_character = request.POST.get('values_character', '')
        obj.budget = request.POST.get('budget', 0)
        obj.month = request.POST.get('month', '')
        obj.time_start = request.POST.get('time_start')
        obj.time_finish = request.POST.get('time_finish')
        obj.pic = request.POST.get('pic', '')
        obj.notes = request.POST.get('notes', '')
        obj.save()
        messages.success(request, 'Jenis kegiatan berhasil diperbarui.')
        return redirect('activities:type_list')
    return render(request, 'activities/type_form.html', {'obj': obj, 'action': 'Edit'})


@login_required
def delete_activity_type(request, pk):
    if request.user.role not in ['admin', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    obj = get_object_or_404(ActivityType, id=pk)
    obj.delete()
    messages.success(request, 'Jenis kegiatan berhasil dihapus.')
    return redirect('activities:type_list')


@login_required
def import_activity_types(request):
    if request.user.role not in ['admin', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        imported = 0
        errors = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            name = str(row[0]).strip() if row[0] else ''
            if not name:
                continue
            try:
                ActivityType.objects.create(
                    name=name, objectives=str(row[1] or ''), values_knowledge=str(row[2] or ''),
                    values_faith=str(row[3] or ''), values_character=str(row[4] or ''),
                    budget=float(row[5] or 0), month=str(row[6] or ''),
                    time_start=row[7], time_finish=row[8], pic=str(row[9] or ''),
                    notes=str(row[10] or ''), created_by=request.user,
                )
                imported += 1
            except Exception as e:
                errors.append(f'Baris {i}: {e}')
        messages.success(request, f'Berhasil import {imported} jenis kegiatan.')
        if errors:
            messages.error(request, '; '.join(errors[:5]))
        return redirect('activities:type_list')
    return render(request, 'activities/import_types.html')


@login_required
def download_type_template(request):
    if request.user.role not in ['admin', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Jenis Kegiatan"
    headers = ['Nama Kegiatan', 'Objectives/Goal', 'Value Knowledge', 'Value Faith', 'Value Character', 'Budget', 'Bulan', 'Time Start', 'Time Finish', 'PIC', 'Notes']
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="004085")
        cell.fill = header_fill
        ws.column_dimensions[chr(64+col)].width = 22
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=template_jenis_kegiatan.xlsx'
    wb.save(response)
    return response