from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font
import os
from datetime import datetime
from .models import (
    SchoolProfile, ContactInfo, SocialMedia, CurriculumInfo,
    SpecialProgram, TeacherProfile, GradingInfo,
    TeacherStructure, SchoolLevelInfo,
)


def public_required(view_func):
    def wrapper(request, *args, **kwargs):
        return view_func(request, *args, **kwargs)
    return wrapper


# ─── PUBLIC VIEWS (no login required) ───

def school_profile(request):
    profile = SchoolProfile.get_profile()
    contacts = ContactInfo.objects.filter(school=profile)
    social = SocialMedia.objects.filter(school=profile)
    return render(request, 'public_info/school_profile.html', {
        'profile': profile, 'contacts': contacts, 'social': social,
    })


def curriculum_info(request):
    curricula = CurriculumInfo.objects.all()
    programs = SpecialProgram.objects.filter(is_active=True)
    return render(request, 'public_info/curriculum.html', {
        'curricula': curricula, 'programs': programs,
    })


def teacher_list(request):
    teachers = TeacherProfile.objects.filter(is_active=True)
    return render(request, 'public_info/teacher_list.html', {'teachers': teachers})


def grading_info(request):
    gradings = GradingInfo.objects.all()
    return render(request, 'public_info/grading.html', {'gradings': gradings})


def public_calendar(request):
    from calendars.models import AcademicCalendar
    calendars = AcademicCalendar.objects.filter(is_active=True).order_by('-updated_at')
    return render(request, 'public_info/calendar.html', {'calendars': calendars})


def public_index(request):
    return render(request, 'public_info/index.html')


# ─── ADMIN CRUD VIEWS (admin, tu, vp_activity) ───

def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if request.user.role not in ['admin', 'tu', 'vp_activity']:
            messages.error(request, 'Akses ditolak.')
            return redirect('accounts:dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


def _apply_sorting(queryset, request, sort_map, default_field=None):
    sort = request.GET.get('sort')
    dir = request.GET.get('dir', 'asc')
    if sort and sort in sort_map:
        field = sort_map[sort]
        order_by = field if dir == 'asc' else f'-{field}'
        return queryset.order_by(order_by)
    if default_field:
        return queryset.order_by(default_field)
    return queryset


@login_required
@admin_required
def admin_school_profile(request):
    profile = SchoolProfile.get_profile()
    if request.method == 'POST':
        profile.name = request.POST.get('name')
        profile.address = request.POST.get('address')
        profile.phone = request.POST.get('phone')
        profile.email = request.POST.get('email')
        profile.website = request.POST.get('website', '')
        profile.vision = request.POST.get('vision')
        profile.mission = request.POST.get('mission')
        profile.core_values = request.POST.get('core_values', '')
        profile.history = request.POST.get('history', '')
        profile.accreditation = request.POST.get('accreditation', '')
        if request.FILES.get('logo'):
            profile.logo = request.FILES['logo']
        profile.save()
        for level in ['sd', 'smp', 'sma']:
            acc = request.POST.get(f'{level}_accreditation', '')
            acc_link = request.POST.get(f'{level}_accreditation_link', '')
            acc_valid_raw = request.POST.get(f'{level}_accreditation_valid_from')
            izin = request.POST.get(f'{level}_izin_operasional', '')
            izin_valid_raw = request.POST.get(f'{level}_izin_valid_from')
            acc_valid = datetime.strptime(acc_valid_raw, '%Y-%m-%d').date() if acc_valid_raw else None
            izin_valid = datetime.strptime(izin_valid_raw, '%Y-%m-%d').date() if izin_valid_raw else None
            SchoolLevelInfo.objects.update_or_create(
                level=level,
                defaults={
                    'accreditation': acc,
                    'accreditation_link': acc_link,
                    'accreditation_valid_from': acc_valid,
                    'izin_operasional': izin,
                    'izin_valid_from': izin_valid,
                }
            )
        messages.success(request, 'Profil sekolah berhasil disimpan.')
        return redirect('public_info:admin_school_profile')
    level_info = {li.level: li for li in SchoolLevelInfo.objects.all()}
    return render(request, 'public_info/admin_school_profile.html', {
        'profile': profile,
        'level_info': level_info,
    })


@login_required
@admin_required
def admin_contacts(request):
    contacts = _apply_sorting(
        ContactInfo.objects.all(), request,
        {'label': 'label', 'contact_type': 'contact_type', 'order': 'order'},
        default_field='order',
    )
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            ContactInfo.objects.create(
                label=request.POST.get('label'),
                contact_type=request.POST.get('contact_type'),
                value=request.POST.get('value'),
                is_emergency=request.POST.get('is_emergency') == 'on',
            )
            messages.success(request, 'Kontak berhasil ditambah.')
        elif action == 'delete':
            ContactInfo.objects.filter(id=request.POST.get('pk')).delete()
            messages.success(request, 'Kontak berhasil dihapus.')
        return redirect('public_info:admin_contacts')
    return render(request, 'public_info/admin_contacts.html', {'contacts': contacts})


@login_required
@admin_required
def admin_social(request):
    social = SocialMedia.objects.all()
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            SocialMedia.objects.create(
                platform=request.POST.get('platform'),
                url=request.POST.get('url'),
                icon=request.POST.get('icon', ''),
            )
            messages.success(request, 'Media sosial berhasil ditambah.')
        elif action == 'delete':
            SocialMedia.objects.filter(id=request.POST.get('pk')).delete()
            messages.success(request, 'Media sosial berhasil dihapus.')
        return redirect('public_info:admin_social')
    return render(request, 'public_info/admin_social.html', {'social': social})


@login_required
@admin_required
def admin_curriculum(request):
    curricula = _apply_sorting(
        CurriculumInfo.objects.all(), request,
        {'level': 'level', 'title': 'title', 'order': 'order'},
        default_field='order',
    )
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            CurriculumInfo.objects.create(
                level=request.POST.get('level'),
                title=request.POST.get('title'),
                description=request.POST.get('description'),
                subjects=request.POST.get('subjects', ''),
                pedagogical=request.POST.get('pedagogical', ''),
            )
            messages.success(request, 'Kurikulum berhasil ditambah.')
        elif action == 'delete':
            CurriculumInfo.objects.filter(id=request.POST.get('pk')).delete()
            messages.success(request, 'Kurikulum berhasil dihapus.')
        return redirect('public_info:admin_curriculum')
    return render(request, 'public_info/admin_curriculum.html', {'curricula': curricula})


@login_required
@admin_required
def admin_programs(request):
    programs = _apply_sorting(
        SpecialProgram.objects.all(), request,
        {'name': 'name', 'target_level': 'target_level', 'is_active': 'is_active'},
        default_field='name',
    )
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            SpecialProgram.objects.create(
                name=request.POST.get('name'),
                description=request.POST.get('description'),
                target_level=request.POST.get('target_level', ''),
            )
            messages.success(request, 'Program khusus berhasil ditambah.')
        elif action == 'delete':
            SpecialProgram.objects.filter(id=request.POST.get('pk')).delete()
            messages.success(request, 'Program khusus berhasil dihapus.')
        elif action == 'toggle':
            prog = get_object_or_404(SpecialProgram, id=request.POST.get('pk'))
            prog.is_active = not prog.is_active
            prog.save()
        return redirect('public_info:admin_programs')
    return render(request, 'public_info/admin_programs.html', {'programs': programs})


@login_required
@admin_required
def admin_teachers(request, pk=None):
    teachers = _apply_sorting(
        TeacherProfile.objects.all(), request,
        {'name': 'full_name', 'level': 'level', 'position': 'position', 'nip': 'nip'},
        default_field='order',
    )
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            TeacherProfile.objects.create(
                full_name=request.POST.get('full_name'),
                qualification=request.POST.get('qualification'),
                subjects=request.POST.get('subjects'),
                position=request.POST.get('position', ''),
                email=request.POST.get('email', ''),
                phone=request.POST.get('phone', ''),
                level=request.POST.get('level', ''),
                nip=request.POST.get('nip', ''),
                photo=request.FILES.get('photo'),
            )
            messages.success(request, 'Guru berhasil ditambah.')
            return redirect('public_info:admin_teachers')
        elif action == 'edit':
            t = get_object_or_404(TeacherProfile, id=request.POST.get('pk'))
            t.full_name = request.POST.get('full_name')
            t.qualification = request.POST.get('qualification')
            t.subjects = request.POST.get('subjects')
            t.position = request.POST.get('position', '')
            t.email = request.POST.get('email', '')
            t.phone = request.POST.get('phone', '')
            t.level = request.POST.get('level', '')
            t.nip = request.POST.get('nip', '')
            if request.FILES.get('photo'):
                t.photo = request.FILES['photo']
            t.save()
            messages.success(request, 'Guru berhasil diperbarui.')
            return redirect('public_info:admin_teachers')
        elif action == 'toggle':
            t = get_object_or_404(TeacherProfile, id=request.POST.get('pk'))
            t.is_active = not t.is_active
            t.save()
            messages.success(request, 'Status guru berhasil diubah.')
            return redirect('public_info:admin_teachers')
        elif action == 'delete':
            TeacherProfile.objects.filter(id=request.POST.get('pk')).delete()
            messages.success(request, 'Guru berhasil dihapus.')
            return redirect('public_info:admin_teachers')
    context = {'teachers': teachers}
    if pk:
        context['edit_teacher'] = get_object_or_404(TeacherProfile, id=pk)
    return render(request, 'public_info/admin_teachers.html', context)


@login_required
@admin_required
def admin_teacher_structure(request):
    if request.method == 'POST':
        level = request.POST.get('level')
        image = request.FILES.get('image')
        if level and image:
            TeacherStructure.objects.update_or_create(
                level=level,
                defaults={'image': image},
            )
            messages.success(request, 'Struktur organisasi berhasil disimpan.')
        else:
            messages.error(request, 'Pilih jenjang dan unggah gambar struktur.')
        return redirect('public_info:admin_teacher_structure')
    structures = TeacherStructure.objects.all()
    return render(request, 'public_info/admin_teacher_structure.html', {
        'structures': structures,
    })


@login_required
@admin_required
def download_teacher_template_xlsx(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Template Guru'
    headers = ['full_name', 'nip', 'qualification', 'subjects', 'position', 'email', 'phone', 'level']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="template_import_guru.xlsx"'
    wb.save(response)
    return response


@login_required
@admin_required
def import_teachers_xlsx(request):
    if request.method == 'POST':
        xlsx_file = request.FILES.get('file')
        if not xlsx_file:
            messages.error(request, 'Pilih file XLSX terlebih dahulu.')
            return redirect('public_info:admin_teachers')
        import tempfile
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        try:
            for chunk in xlsx_file.chunks():
                tmp.write(chunk)
            tmp.close()
            wb = openpyxl.load_workbook(tmp.name)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            created = 0
            for row in rows:
                full_name = row[0]
                if not full_name or not str(full_name).strip():
                    continue
                TeacherProfile.objects.create(
                    full_name=str(full_name).strip(),
                    nip=str(row[1]).strip() if row[1] else '',
                    qualification=str(row[2]).strip() if row[2] else '',
                    subjects=str(row[3]).strip() if row[3] else '',
                    position=str(row[4]).strip() if row[4] else '',
                    email=str(row[5]).strip() if row[5] else '',
                    phone=str(row[6]).strip() if row[6] else '',
                    level=str(row[7]).strip() if row[7] else '',
                )
                created += 1
            messages.success(request, f'{created} guru berhasil diimpor.')
        except Exception as e:
            messages.error(request, f'Gagal mengimpor file: {e}')
        finally:
            os.unlink(tmp.name)
        return redirect('public_info:admin_teachers')
    return redirect('public_info:admin_teachers')


@login_required
@admin_required
def admin_grading(request):
    gradings = _apply_sorting(
        GradingInfo.objects.all(), request,
        {'level': 'level', 'title': 'title', 'order': 'order'},
        default_field='order',
    )
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            GradingInfo.objects.create(
                level=request.POST.get('level'),
                title=request.POST.get('title'),
                description=request.POST.get('description'),
                grading_scale=request.POST.get('grading_scale', ''),
                report_guide=request.POST.get('report_guide', ''),
            )
            messages.success(request, 'Sistem penilaian berhasil ditambah.')
        elif action == 'delete':
            GradingInfo.objects.filter(id=request.POST.get('pk')).delete()
            messages.success(request, 'Sistem penilaian berhasil dihapus.')
        return redirect('public_info:admin_grading')
    return render(request, 'public_info/admin_grading.html', {'gradings': gradings})


def public_verify_invoice(request):
    from payments.models import PaymentTransaction
    from finance.models import SPPBill
    from accounts.models import User
    import hashlib
    from django.utils import timezone
    result = None
    result_type = ''
    code = ''
    if request.method == 'POST':
        code = request.POST.get('code', '').strip().upper()
        tx = PaymentTransaction.objects.filter(verification_code=code).first()
        if tx:
            result = tx
            result_type = 'payment'
        else:
            bill = SPPBill.objects.filter(verification_code=code).first()
            if bill:
                result = bill
                result_type = 'bill'
        if not result:
            # Fallback: try to match as SPP bill code using SHA256 formula
            from finance.models import SPPBill as SPPBillModel
            for bill in SPPBillModel.objects.filter(verification_code='').order_by('-id')[:100]:
                computed = hashlib.sha256(f"spp-bill-{bill.id}-{bill.student.nis}".encode()).hexdigest()[:12].upper()
                if computed == code:
                    bill.verification_code = code
                    bill.save(update_fields=['verification_code'])
                    result = bill
                    result_type = 'bill'
                    break
        if not result:
            # Check export codes: SHA256("export-{user_id}-{YYYYMMDD}")[:12] for today
            today = timezone.localtime(timezone.now()).strftime('%Y%m%d')
            for user in User.objects.all():
                computed = hashlib.sha256(f"export-{user.id}-{today}".encode()).hexdigest()[:12].upper()
                if computed == code:
                    result = {'type': 'export', 'user': user.username, 'date': timezone.localtime(timezone.now()).strftime('%d/%m/%Y')}
                    result_type = 'export'
                    break
    return render(request, 'public_info/verify_invoice.html', {'result': result, 'code': code, 'result_type': result_type})
