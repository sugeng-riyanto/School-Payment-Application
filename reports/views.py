from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, Count
from django.template.loader import render_to_string
from finance.models import SPPBill, DPP, PaymentReminder, Compensation
from eca.models import ECARegistration, ECAPayment
from payments.models import PaymentProof
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime


def get_export_data(user):
    if user.role == 'parent':
        children = user.children.all()
        spp_bills = SPPBill.objects.filter(student__in=children)
        eca_regs = ECARegistration.objects.filter(student__in=children)
        dpps = DPP.objects.filter(student__in=children)
    elif user.role in ['kepsek', 'vp_activity'] and user.assigned_level:
        lvl = user.assigned_level
        spp_bills = SPPBill.objects.filter(student__class_grade__grade__level=lvl)
        eca_regs = ECARegistration.objects.filter(student__class_grade__grade__level=lvl)
        dpps = DPP.objects.filter(student__class_grade__grade__level=lvl)
    else:
        spp_bills = SPPBill.objects.all()
        eca_regs = ECARegistration.objects.all()
        dpps = DPP.objects.all()
    return spp_bills, eca_regs, dpps


def render_to_pdf(template_src, context_dict):
    from xhtml2pdf import pisa
    html = render_to_string(template_src, context_dict)
    response = HttpResponse(content_type='application/pdf')
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('PDF generation error', status=500)
    return response


XLSX_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


@login_required
def export_spp_xlsx(request):
    spp_bills, _, _ = get_export_data(request.user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan SPP"

    headers = ['No', 'Nama Siswa', 'Kelas', 'Bulan', 'Tahun', 'Jumlah', 'Status', 'Tgl Bayar']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for i, bill in enumerate(spp_bills, 1):
        ws.append([
            i, bill.student.full_name, str(bill.student.class_grade),
            bill.get_month_display(), bill.year, int(bill.amount),
            'Lunas' if bill.is_paid else 'Belum',
            bill.paid_at.strftime('%d/%m/%Y') if bill.paid_at else '-',
        ])

    response = HttpResponse(content_type=XLSX_TYPE)
    response['Content-Disposition'] = f'attachment; filename=laporan_spp_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
def export_spp_pdf(request):
    spp_bills, _, _ = get_export_data(request.user)
    rows = []
    for i, bill in enumerate(spp_bills, 1):
        rows.append({'values': [
            i, bill.student.full_name, str(bill.student.class_grade),
            bill.get_month_display(), bill.year,
            f'Rp{int(bill.amount):,}',
            'Lunas' if bill.is_paid else 'Belum'
        ]})
    return render_to_pdf('reports/pdf_generic.html', {
        'rows': rows, 'title': 'Laporan SPP',
        'headers': ['No', 'Nama', 'Kelas', 'Bulan', 'Tahun', 'Jumlah', 'Status'],
    })


@login_required
def export_eca_xlsx(request):
    _, eca_regs, _ = get_export_data(request.user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan ECA"

    headers = ['No', 'Nama Siswa', 'Program', 'Harga', 'Status Bayar', 'Tgl Daftar']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for i, reg in enumerate(eca_regs, 1):
        ws.append([
            i, reg.student.full_name, reg.program.name,
            int(reg.program.price),
            'Lunas' if reg.is_paid else 'Belum',
            reg.registered_at.strftime('%d/%m/%Y'),
        ])

    response = HttpResponse(content_type=XLSX_TYPE)
    response['Content-Disposition'] = f'attachment; filename=laporan_eca_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
def export_eca_pdf(request):
    _, eca_regs, _ = get_export_data(request.user)
    rows = []
    for i, reg in enumerate(eca_regs, 1):
        rows.append({'values': [
            i, reg.student.full_name, reg.program.name,
            f'Rp{int(reg.program.price):,}',
            'Lunas' if reg.is_paid else 'Belum'
        ]})
    return render_to_pdf('reports/pdf_generic.html', {
        'rows': rows, 'title': 'Laporan ECA',
        'headers': ['No', 'Nama', 'Program', 'Harga', 'Status'],
    })


@login_required
def export_dpp_xlsx(request):
    _, _, dpps = get_export_data(request.user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan DPP"

    headers = ['No', 'Nama Siswa', 'Kategori', 'Jumlah', 'Status', 'Tgl Bayar']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for i, dpp in enumerate(dpps, 1):
        ws.append([
            i, dpp.student.full_name, dpp.get_category_display(),
            int(dpp.amount),
            'Lunas' if dpp.is_paid else 'Belum',
            dpp.paid_at.strftime('%d/%m/%Y') if dpp.paid_at else '-',
        ])

    response = HttpResponse(content_type=XLSX_TYPE)
    response['Content-Disposition'] = f'attachment; filename=laporan_dpp_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
def export_dpp_pdf(request):
    _, _, dpps = get_export_data(request.user)
    rows = []
    for i, dpp in enumerate(dpps, 1):
        rows.append({'values': [
            i, dpp.student.full_name, dpp.get_category_display(),
            f'Rp{int(dpp.amount):,}',
            'Lunas' if dpp.is_paid else 'Belum'
        ]})
    return render_to_pdf('reports/pdf_generic.html', {
        'rows': rows, 'title': 'Laporan DPP Pembangunan',
        'headers': ['No', 'Nama', 'Kategori', 'Jumlah', 'Status'],
    })


@login_required
def export_reminder_xlsx(request):
    reminders = PaymentReminder.objects.all().order_by('-sent_at')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Peringatan"
    headers = ['No', 'Siswa', 'Tipe', 'Bulan', 'Tahun Ajaran', 'Tgl Kirim', 'Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for i, r in enumerate(reminders, 1):
        ws.append([
            i, r.student.full_name, r.get_reminder_type_display(),
            r.get_month_display() if r.month else '-',
            str(r.academic_year),
            r.sent_at.strftime('%d/%m/%Y %H:%M') if r.sent_at else '-',
            'Dibaca' if r.is_read else 'Belum',
        ])
    response = HttpResponse(content_type=XLSX_TYPE)
    response['Content-Disposition'] = f'attachment; filename=laporan_peringatan_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
def export_reminder_pdf(request):
    reminders = PaymentReminder.objects.all().order_by('-sent_at')
    rows = []
    for i, r in enumerate(reminders, 1):
        rows.append({'values': [
            i, r.student.full_name, r.get_reminder_type_display(),
            r.get_month_display() if r.month else '-',
            r.sent_at.strftime('%d/%m/%Y') if r.sent_at else '-',
            'Dibaca' if r.is_read else 'Belum',
        ]})
    return render_to_pdf('reports/pdf_generic.html', {
        'rows': rows, 'title': 'Laporan Peringatan Pembayaran',
        'headers': ['No', 'Nama', 'Tipe', 'Bulan', 'Tgl Kirim', 'Status'],
    })


@login_required
def export_compensation_xlsx(request):
    compensations = Compensation.objects.all().order_by('-created_at')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kompensasi"
    headers = ['No', 'Siswa', 'Tipe', 'Asli', 'Potongan', 'Akhir', 'Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for i, c in enumerate(compensations, 1):
        ws.append([
            i, c.student.full_name, c.get_comp_type_display(),
            int(c.original_amount), int(c.comp_amount), int(c.final_amount),
            'Aktif' if c.is_active else 'Nonaktif',
        ])
    response = HttpResponse(content_type=XLSX_TYPE)
    response['Content-Disposition'] = f'attachment; filename=laporan_kompensasi_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
def export_compensation_pdf(request):
    compensations = Compensation.objects.all().order_by('-created_at')
    rows = []
    for i, c in enumerate(compensations, 1):
        rows.append({'values': [
            i, c.student.full_name, c.get_comp_type_display(),
            f'Rp{int(c.original_amount):,}',
            f'Rp{int(c.comp_amount):,}',
            f'Rp{int(c.final_amount):,}',
        ]})
    return render_to_pdf('reports/pdf_generic.html', {
        'rows': rows, 'title': 'Laporan Kompensasi Pembayaran',
        'headers': ['No', 'Nama', 'Tipe', 'Asli', 'Potongan', 'Akhir'],
    })


@login_required
def export_proof_xlsx(request):
    proofs = PaymentProof.objects.all().order_by('-uploaded_at')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bukti Pembayaran"
    headers = ['No', 'Siswa', 'Tagihan', 'Tgl Upload', 'Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for i, p in enumerate(proofs, 1):
        ws.append([
            i, p.student.full_name,
            f"{p.spp_bill.get_month_display()} {p.spp_bill.year}" if p.spp_bill else '-',
            p.uploaded_at.strftime('%d/%m/%Y %H:%M') if p.uploaded_at else '-',
            'Terverifikasi' if p.is_verified else 'Menunggu',
        ])
    response = HttpResponse(content_type=XLSX_TYPE)
    response['Content-Disposition'] = f'attachment; filename=laporan_bukti_bayar_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
def export_proof_pdf(request):
    proofs = PaymentProof.objects.all().order_by('-uploaded_at')
    rows = []
    for i, p in enumerate(proofs, 1):
        rows.append({'values': [
            i, p.student.full_name,
            f"{p.spp_bill.get_month_display()} {p.spp_bill.year}" if p.spp_bill else '-',
            p.uploaded_at.strftime('%d/%m/%Y') if p.uploaded_at else '-',
            'Terverifikasi' if p.is_verified else 'Menunggu',
        ]})
    return render_to_pdf('reports/pdf_generic.html', {
        'rows': rows, 'title': 'Laporan Bukti Pembayaran',
        'headers': ['No', 'Nama', 'Tagihan', 'Tgl Upload', 'Status'],
    })
