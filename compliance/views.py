from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, Alignment
from django.template.loader import render_to_string


def compliance_required(view_func):
    def wrapper(request, *args, **kwargs):
        if request.user.role not in ['admin', 'kepsek']:
            messages.error(request, 'Akses ditolak.')
            return redirect('accounts:dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
@compliance_required
def compliance_dashboard(request):
    from .models import AuditLog, SecurityEvent, ComplianceChecklist, PrivacyConsent, DataRetentionPolicy
    from accounts.models import User

    # Level filter for kepsek
    level_usernames = []
    if request.user.role == 'kepsek' and request.user.assigned_level:
        level_usernames = list(User.objects.filter(
            assigned_level=request.user.assigned_level
        ).values_list('username', flat=True))
        level_usernames.append(request.user.username)
    level_filter = {'username__in': level_usernames} if level_usernames else {}

    now = timezone.now()
    thirty_days_ago = now - timedelta(days=30)

    # ── Audit Log Stats ──
    all_logs = AuditLog.objects.filter(**level_filter)
    total_audit_logs = all_logs.count()
    audit_logs_30d = all_logs.filter(created_at__gte=thirty_days_ago).count()
    audit_by_action = list(
        all_logs.filter(created_at__gte=thirty_days_ago)
        .values('action')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    unique_users_audited = all_logs.filter(created_at__gte=thirty_days_ago).values('username').distinct().count()

    # Daily audit volume for chart (last 30 days)
    daily_audit = []
    daily_labels = []
    for i in range(29, -1, -1):
        day = now - timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        count = all_logs.filter(created_at__gte=day_start, created_at__lt=day_end).count()
        daily_labels.append(day.strftime('%d/%m'))
        daily_audit.append(count)

    # ── Security Events (filtered by user__username__in for kepsek) ──
    events_qs = SecurityEvent.objects.filter(user__username__in=level_usernames) if level_usernames else SecurityEvent.objects
    total_security_events = events_qs.count()
    unresolved_events = events_qs.filter(is_resolved=False).count()
    high_critical_events = events_qs.filter(severity__in=['high', 'critical'], is_resolved=False).count()
    events_by_severity = list(
        SecurityEvent.objects.values('severity').annotate(count=Count('id')).order_by('severity')
    )
    recent_events = events_qs.filter(is_resolved=False).order_by('-created_at')[:10]

    # ── Compliance Checklist ──
    checklists = ComplianceChecklist.objects.all()
    total_controls = checklists.count()
    compliant_count = checklists.filter(status='compliant').count()
    non_compliant_count = checklists.filter(status='non_compliant').count()
    partial_count = checklists.filter(status='partial').count()
    not_assessed = checklists.filter(status='not_assessed').count()

    # Per-framework stats for chart
    frameworks_data = {}
    for fw_code, fw_label in ComplianceChecklist.FRAMEWORK_CHOICES:
        qs = checklists.filter(framework=fw_code)
        total = qs.count()
        compliant = qs.filter(status='compliant').count()
        frameworks_data[fw_code] = {
            'label': fw_label,
            'total': total,
            'compliant': compliant,
            'pct': round(compliant / total * 100, 1) if total else 0,
        }

    # ── Privacy Consents ──
    total_users = User.objects.count()
    consent_count = PrivacyConsent.objects.filter(is_granted=True).values('user').distinct().count()
    consent_by_type = list(
        PrivacyConsent.objects.filter(is_granted=True)
        .values('consent_type')
        .annotate(count=Count('id', distinct=True))
    )

    # ── Data Retention ──
    retention_policies = DataRetentionPolicy.objects.filter(is_active=True)

    # ── Recent Audit Logs ──
    recent_logs = AuditLog.objects.select_related('user').all()[:50]

    context = {
        'total_audit_logs': total_audit_logs,
        'audit_logs_30d': audit_logs_30d,
        'audit_by_action': audit_by_action,
        'unique_users_audited': unique_users_audited,
        'daily_labels': daily_labels,
        'daily_audit': daily_audit,
        'total_security_events': total_security_events,
        'unresolved_events': unresolved_events,
        'high_critical_events': high_critical_events,
        'events_by_severity': events_by_severity,
        'recent_events': recent_events,
        'total_controls': total_controls,
        'compliant_count': compliant_count,
        'non_compliant_count': non_compliant_count,
        'partial_count': partial_count,
        'not_assessed': not_assessed,
        'frameworks_data': frameworks_data,
        'total_users': total_users,
        'consent_count': consent_count,
        'consent_by_type': consent_by_type,
        'retention_policies': retention_policies,
        'recent_logs': recent_logs,
    }
    return render(request, 'compliance/dashboard.html', context)


@login_required
@compliance_required
def audit_log_list(request):
    from .models import AuditLog
    from accounts.models import User

    # Level filter for kepsek
    level_filter = {}
    if request.user.role == 'kepsek' and request.user.assigned_level:
        level_usernames = list(User.objects.filter(
            assigned_level=request.user.assigned_level
        ).values_list('username', flat=True))
        level_usernames.append(request.user.username)
        level_filter = {'username__in': level_usernames}

    query = request.GET.get('q', '')
    action = request.GET.get('action', '')
    username = request.GET.get('username', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    logs = AuditLog.objects.select_related('user').filter(**level_filter)

    if query:
        logs = logs.filter(
            Q(description__icontains=query) |
            Q(username__icontains=query) |
            Q(model_name__icontains=query) |
            Q(object_repr__icontains=query)
        )
    if action:
        logs = logs.filter(action=action)
    if username:
        logs = logs.filter(username__icontains=username)
    if date_from:
        logs = logs.filter(created_at__gte=date_from)
    if date_to:
        logs = logs.filter(created_at__lte=date_to + 'T23:59:59')

    logs = logs[:200]
    action_choices = AuditLog.ACTION_CHOICES

    context = {
        'logs': logs,
        'action_choices': action_choices,
        'filter_action': action,
        'filter_username': username,
        'filter_date_from': date_from,
        'filter_date_to': date_to,
        'filter_query': query,
    }
    return render(request, 'compliance/audit_log_list.html', context)


@login_required
@compliance_required
def export_audit_log_xlsx(request):
    from .models import AuditLog
    from accounts.models import User

    level_filter = {}
    if request.user.role == 'kepsek' and request.user.assigned_level:
        level_usernames = list(User.objects.filter(assigned_level=request.user.assigned_level).values_list('username', flat=True))
        level_usernames.append(request.user.username)
        level_filter = {'username__in': level_usernames}

    logs = AuditLog.objects.select_related('user').filter(**level_filter)[:1000]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Log"

    headers = ['Timestamp', 'User', 'Action', 'Model', 'Object', 'Description', 'IP Address', 'Method', 'Path']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for log in logs:
        ws.append([
            log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else '',
            log.username,
            log.get_action_display(),
            log.model_name,
            log.object_repr,
            log.description,
            log.ip_address or '',
            log.request_method,
            log.request_path,
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=audit_log_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


@login_required
@compliance_required
def export_audit_log_pdf(request):
    from .models import AuditLog
    from accounts.models import User

    level_filter = {}
    if request.user.role == 'kepsek' and request.user.assigned_level:
        level_usernames = list(User.objects.filter(assigned_level=request.user.assigned_level).values_list('username', flat=True))
        level_usernames.append(request.user.username)
        level_filter = {'username__in': level_usernames}

    logs = AuditLog.objects.select_related('user').filter(**level_filter)[:200]
    rows = []
    for log in logs:
        rows.append({'values': [
            log.created_at.strftime('%Y-%m-%d %H:%M') if log.created_at else '',
            log.username,
            log.get_action_display(),
            log.model_name[:20],
            log.description[:50],
        ]})

    html = render_to_string('compliance/pdf_audit_log.html', {
        'rows': rows,
        'title': 'Audit Trail Log',
        'headers': ['Timestamp', 'User', 'Action', 'Model', 'Description'],
        'generated_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
    })
    from xhtml2pdf import pisa
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=audit_log_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Gagal membuat PDF', status=500)
    return response


@login_required
@compliance_required
def export_compliance_xlsx(request):
    from .models import ComplianceChecklist

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compliance Status"

    headers = ['Framework', 'Control ID', 'Control Name', 'Status', 'Notes', 'Last Assessed']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for c in ComplianceChecklist.objects.all().order_by('framework', 'control_id'):
        ws.append([
            c.get_framework_display(),
            c.control_id,
            c.control_name,
            c.get_status_display(),
            c.notes,
            c.last_assessed.strftime('%Y-%m-%d') if c.last_assessed else '',
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=compliance_status_{timezone.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
@compliance_required
def seed_compliance_checklist(request):
    from .models import ComplianceChecklist

    SOC2_CONTROLS = [
        ('CC1.1', 'Board/management oversight of security', 'soc2'),
        ('CC1.2', 'Formal security roles and responsibilities', 'soc2'),
        ('CC1.3', 'Security policies and procedures', 'soc2'),
        ('CC2.1', 'System monitoring and detection', 'soc2'),
        ('CC2.2', 'Communication of security policies', 'soc2'),
        ('CC3.1', 'Risk assessment process', 'soc2'),
        ('CC3.2', 'Risk mitigation activities', 'soc2'),
        ('CC4.1', 'Monitoring of controls', 'soc2'),
        ('CC4.2', 'Remediation of deficiencies', 'soc2'),
        ('CC5.1', 'Access provisioning and de-provisioning', 'soc2'),
        ('CC5.2', 'Logical and physical access controls', 'soc2'),
        ('CC6.1', 'System monitoring and logging', 'soc2'),
        ('CC6.2', 'Data encryption in transit and at rest', 'soc2'),
        ('CC6.3', 'Secure system development lifecycle', 'soc2'),
        ('CC7.1', 'Incident response plan', 'soc2'),
        ('CC7.2', 'Business continuity and disaster recovery', 'soc2'),
    ]
    PDP_CONTROLS = [
        ('PDP-1', 'Privacy policy published and accessible', 'pdp'),
        ('PDP-2', 'Data processing consent mechanism', 'pdp'),
        ('PDP-3', 'Data inventory and mapping', 'pdp'),
        ('PDP-4', 'Data subject access request (SAR) procedure', 'pdp'),
        ('PDP-5', 'Data retention and deletion policy', 'pdp'),
        ('PDP-6', 'Data breach notification procedure', 'pdp'),
        ('PDP-7', 'Data Protection Officer (DPO) appointment', 'pdp'),
        ('PDP-8', 'Data processing agreement with third parties', 'pdp'),
        ('PDP-9', 'Cross-border data transfer safeguards', 'pdp'),
        ('PDP-10', 'Privacy impact assessment (DPIA)', 'pdp'),
    ]
    PCI_CONTROLS = [
        ('PCI-1', 'Firewall and network segmentation', 'pci_dss'),
        ('PCI-2', 'Secure configuration of systems', 'pci_dss'),
        ('PCI-3', 'Cardholder data protection at rest', 'pci_dss'),
        ('PCI-4', 'Encryption of cardholder data in transit', 'pci_dss'),
        ('PCI-5', 'Anti-malware protection', 'pci_dss'),
        ('PCI-6', 'Access control to cardholder data', 'pci_dss'),
        ('PCI-7', 'Unique user IDs for access', 'pci_dss'),
        ('PCI-8', 'Physical security of systems', 'pci_dss'),
        ('PCI-9', 'Logging and monitoring of access', 'pci_dss'),
        ('PCI-10', 'Regular security testing', 'pci_dss'),
        ('PCI-11', 'Information security policy', 'pci_dss'),
        ('PCI-12', 'Vendor/third-party risk management', 'pci_dss'),
    ]

    created = 0
    for controls in [SOC2_CONTROLS, PDP_CONTROLS, PCI_CONTROLS]:
        for cid, name, fw in controls:
            _, was_created = ComplianceChecklist.objects.get_or_create(
                framework=fw,
                control_id=cid,
                defaults={
                    'control_name': name,
                    'description': f'Control {cid}: {name}',
                    'status': 'partial' if 'soc2' in fw else 'not_assessed' if 'pdp' in fw else 'partial' if 'pci_dss' in fw else 'not_assessed',
                }
            )
            if was_created:
                created += 1

    messages.success(request, f'{created} compliance controls seeded successfully.')
    return redirect('compliance:dashboard')
