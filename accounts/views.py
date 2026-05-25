from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q
from .models import AcademicYear
import openpyxl


def login_view(request):
    if request.user.is_authenticated:
        return redirect('accounts:dashboard')
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            from compliance.middleware import AuditMiddleware
            AuditMiddleware.log_action(request, 'login', 'User', user.id, user.username, f'User {user.username} logged in')
            return redirect('accounts:dashboard')
        from compliance.middleware import AuditMiddleware
        AuditMiddleware.log_action(request, 'login_failed', 'User', '', username, f'Failed login attempt for {username}')
        messages.error(request, 'Username atau password salah.')
    return render(request, 'accounts/login.html')


def logout_view(request):
    from compliance.middleware import AuditMiddleware
    if request.user.is_authenticated:
        AuditMiddleware.log_action(request, 'logout', 'User', request.user.id, request.user.username, f'User {request.user.username} logged out')
    logout(request)
    return redirect('accounts:login')


@login_required
def dashboard(request):
    role = request.user.role
    template_map = {
        'admin': 'reports/admin_dashboard.html',
        'tu': 'reports/tu_dashboard.html',
        'kepsek': 'reports/kepsek_dashboard.html',
        'vp_activity': 'reports/vp_activity_dashboard.html',
        'pic_teacher': 'reports/pic_teacher_dashboard.html',
        'eca_director': 'reports/eca_director_dashboard.html',
        'parent': 'reports/parent_dashboard.html',
    }
    template = template_map.get(role, 'reports/parent_dashboard.html')
    ctx = {'user': request.user}

    if role in ['admin', 'tu', 'kepsek']:
        from finance.models import SPPReport, DPPReport, Compensation, SPPBill, DPP
        from eca.models import ECAReport, ECARegistration, ECAPayment
        from payments.models import BlastEmailLog
        
        # Level-aware queries for pending reports
        level = request.user.assigned_level
        allowed = []
        if level and level != 'sd_smp_sma':
            level_map = {'sd': ['sd'], 'smp': ['smp'], 'sma': ['sma'],
                'sd_smp': ['sd','smp'], 'smp_sma': ['smp','sma']}
            allowed = level_map.get(level, [level])
        if allowed:
            ctx['spp_reports_pending'] = SPPReport.objects.filter(status='submitted', grade__level__in=allowed).count()
            ctx['dpp_reports_pending'] = DPPReport.objects.filter(status='submitted', grade__level__in=allowed).count()
            ctx['eca_reports_pending'] = ECAReport.objects.filter(status='submitted', program__grade__level__in=allowed).count()
            ctx['active_compensations'] = Compensation.objects.filter(is_active=True, student__class_grade__grade__level__in=allowed).count()
            ctx['total_unread_blast'] = BlastEmailLog.objects.filter(is_read=False, student__class_grade__grade__level__in=allowed).count()
        else:
            ctx['spp_reports_pending'] = SPPReport.objects.filter(status='submitted').count()
            ctx['dpp_reports_pending'] = DPPReport.objects.filter(status='submitted').count()
            ctx['eca_reports_pending'] = ECAReport.objects.filter(status='submitted').count()
            ctx['active_compensations'] = Compensation.objects.filter(is_active=True).count()
            ctx['total_unread_blast'] = BlastEmailLog.objects.filter(is_read=False).count()
        
        if role == 'admin':
            from accounts.models import User, Student
            from payments.models import PaymentTransaction
            from django.db.models import Sum, Q
            from django.utils import timezone

            ctx['total_users'] = User.objects.count()
            ctx['total_students'] = Student.objects.filter(is_active=True).count()
            ctx['total_paid'] = SPPBill.objects.filter(is_paid=True).count()
            ctx['total_unpaid'] = SPPBill.objects.filter(is_paid=False).count()

            now = timezone.now()
            current_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            ctx['paid_count'] = PaymentTransaction.objects.filter(status='success', created_at__gte=current_month_start).count()
            ctx['unpaid_count'] = PaymentTransaction.objects.filter(status__in=['pending', 'failed'], created_at__gte=current_month_start).count()

            # Monthly payment time series
            current_year = timezone.now().year
            monthly_data = []
            for m in range(1, 13):
                total = PaymentTransaction.objects.filter(status='success', created_at__year=current_year, created_at__month=m).count()
                monthly_data.append(total)
            ctx['time_series_data'] = monthly_data
            
            # ECA real distribution
            eca_dist = {}
            for reg in ECARegistration.objects.filter(status='approved').select_related('program__eca_type'):
                tname = reg.program.eca_type.name
                eca_dist[tname] = eca_dist.get(tname, 0) + 1
            ctx['eca_labels'] = list(eca_dist.keys())
            ctx['eca_values'] = list(eca_dist.values())
        
        if role == 'kepsek':
            from django.db.models import Sum
            assigned = request.user.assigned_level
            level_filter = {}
            if assigned:
                level_filter = {'student__class_grade__grade__level': assigned}
            
            ctx['total_spp_income'] = SPPBill.objects.filter(is_paid=True, **level_filter).aggregate(total=Sum('amount'))['total'] or 0
            ctx['total_dpp_income'] = DPP.objects.filter(is_paid=True, **level_filter).aggregate(total=Sum('amount'))['total'] or 0
            ctx['total_eca_participants'] = ECARegistration.objects.filter(status='approved').count()
            
            # Monthly SPP timeline
            from django.utils import timezone
            current_year = timezone.now().year
            monthly_data = []
            for m in range(1, 13):
                total = SPPBill.objects.filter(is_paid=True, month=m, year=current_year, **level_filter).aggregate(t=Sum('amount'))['t'] or 0
                monthly_data.append(float(total))
            ctx['spp_timeline_data'] = monthly_data
            
            # ECA distribution
            eca_dist = {}
            for reg in ECARegistration.objects.filter(status='approved').select_related('program__eca_type'):
                tname = reg.program.eca_type.name
                eca_dist[tname] = eca_dist.get(tname, 0) + 1
            ctx['eca_labels'] = list(eca_dist.keys())
            ctx['eca_values'] = list(eca_dist.values())
    
    if role == 'tu':
        from finance.models import SPPBill, DPP, PaymentReminder
        from payments.models import PaymentTransaction
        from django.utils import timezone
        
        ctx['total_spp'] = SPPBill.objects.filter(is_paid=False).count()
        ctx['total_dpp'] = DPP.objects.filter(status='pending').count()
        ctx['pending_reminders'] = PaymentReminder.objects.filter(sent_at__isnull=True).count()
        
        now = timezone.now()
        current_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        ctx['paid_count'] = PaymentTransaction.objects.filter(
            status='success', 
            created_at__gte=current_month_start
        ).count()
        ctx['unpaid_count'] = PaymentTransaction.objects.filter(
            status__in=['pending', 'failed'], 
            created_at__gte=current_month_start
        ).count()

    if role == 'parent':
        from payments.models import PaymentTransaction
        from finance.models import SPPBill
        from eca.models import ECARegistration, ECAProgram, ECASetting
        from calendars.models import AcademicCalendar
        from django.utils import timezone
        from datetime import timedelta
        students = request.user.children.all()
        ctx['students'] = students
        if students:
            ctx['spp_bills'] = SPPBill.objects.filter(student__in=students).order_by('-year', '-month')[:12]
            ctx['paid_bills'] = SPPBill.objects.filter(student__in=students, is_paid=True).count()
            ctx['unpaid_bills'] = SPPBill.objects.filter(student__in=students, is_paid=False).count()
            eca_regs = ECARegistration.objects.filter(student__in=students)
            ctx['eca_registrations'] = eca_regs
            ctx['eca_count'] = eca_regs.count()
            ctx['eca_active_count'] = eca_regs.filter(status='approved').count()
            setting = ECASetting.objects.first()
            ctx['eca_max_selection'] = setting.max_selection if setting else 3
            ctx['eca_available'] = ECAProgram.objects.filter(is_open=True, grade__in=students.values('class_grade__grade')).count()
            recent_cal = AcademicCalendar.objects.filter(
                updated_at__gte=timezone.now() - timedelta(days=7)
            ).first()
            ctx['recent_calendar_update'] = recent_cal
            from payments.models import BlastEmailLog
            ctx['blast_logs'] = BlastEmailLog.objects.filter(student__in=students)[:5]
            ctx['unread_blast_count'] = BlastEmailLog.objects.filter(student__in=students, is_read=False).count()

    if role == 'vp_activity':
        from activities.models import ActivityReport
        from eca.models import ECAProgram, ECARegistration, ECAPayment
        from django.utils import timezone

        assigned = request.user.assigned_level
        level_filter = {}
        if assigned:
            level_filter = {'student__class_grade__grade__level': assigned}

        ctx['total_activities'] = ActivityReport.objects.filter(pic__assigned_level=assigned).count() if assigned else ActivityReport.objects.count()
        ctx['total_eca'] = ECAProgram.objects.filter(is_open=True).count()
        ctx['total_participants'] = ECARegistration.objects.filter(status='approved', **level_filter).count()
        ctx['recent_activities'] = ActivityReport.objects.filter(pic__assigned_level=assigned).order_by('-start_date')[:5] if assigned else ActivityReport.objects.order_by('-start_date')[:5]

        # Pending reviews for VP Activity (submitted/pic_approved in their level)

        # Pending reviews for VP Activity (submitted/pic_approved in their level)
        if assigned:
            ctx['pending_reviews'] = ActivityReport.objects.filter(
                status__in=['submitted', 'pic_approved'],
                pic__assigned_level=assigned
            ).order_by('-updated_at')[:10]
        else:
            ctx['pending_reviews'] = ActivityReport.objects.filter(
                status__in=['submitted', 'pic_approved']
            ).order_by('-updated_at')[:10]

        # ECA distribution chart data
        eca_dist = {}
        for reg in ECARegistration.objects.filter(status='approved', **level_filter).select_related('program__eca_type'):
            tname = reg.program.eca_type.name
            eca_dist[tname] = eca_dist.get(tname, 0) + 1
        ctx['vp_eca_labels'] = list(eca_dist.keys())
        ctx['vp_eca_values'] = list(eca_dist.values())

    if role == 'pic_teacher':
        from activities.models import ActivityReport, ActivityIncomeExpense, ActivityEvidence

        my_activities = ActivityReport.objects.filter(pic=request.user)
        ctx['my_activities'] = my_activities.count()
        ctx['activities'] = my_activities.order_by('-start_date')[:10]

        # Total income from all my activities
        from django.db.models import Sum
        income = ActivityIncomeExpense.objects.filter(
            activity__pic=request.user,
            transaction_type='income'
        ).aggregate(total=Sum('amount'))['total'] or 0
        ctx['total_income'] = float(income)

        # Total evidence files
        ctx['total_evidences'] = ActivityEvidence.objects.filter(activity__pic=request.user).count()

        # Feedback alerts (activities where someone left feedback)
        ctx['feedback_alerts'] = my_activities.filter(
            feedback_notes__gt='', feedback_by__isnull=False
        ).order_by('-feedback_at')[:10]

    if role == 'eca_director':
        from eca.models import ECAType, ECAProgram, ECARegistration, ECAPayment, ECAReport
        from django.db.models import Sum, Q

        assigned = request.user.assigned_level
        lf = {}
        if assigned:
            lf = {'grade__level': assigned}

        ctx['total_types'] = ECAType.objects.filter(is_active=True).count()
        ctx['total_programs'] = ECAProgram.objects.filter(is_open=True, **lf).count()
        ctx['total_registrations'] = ECARegistration.objects.filter(**{k.replace('grade__', 'program__grade__'): v for k, v in lf.items()}).count() if lf else ECARegistration.objects.count()
        ctx['eca_available'] = ECAProgram.objects.filter(is_open=True, **lf).aggregate(total=Sum('max_participants'))['total'] or 0

        # ECA Finances
        eca_regs_qs = ECARegistration.objects.filter(status='approved')
        if assigned:
            eca_regs_qs = eca_regs_qs.filter(student__class_grade__grade__level=assigned)

        # Verified income (is_paid = confirmed by TU)
        verified_income = eca_regs_qs.filter(is_paid=True).aggregate(total=Sum('program__price'))['total'] or 0
        ctx['verified_income'] = float(verified_income)

        # Total revenue (from ECAPayment)
        from eca.models import ECAPayment
        pay_qs = ECAPayment.objects.filter(is_paid=True, payment_status='approved')
        if assigned:
            pay_qs = pay_qs.filter(registration__student__class_grade__grade__level=assigned)
        total_revenue = pay_qs.aggregate(total=Sum('paid_amount'))['total'] or 0
        ctx['total_revenue'] = float(total_revenue)

        # Spent from approved ECA reports
        eca_reports_qs = ECAReport.objects.filter(status='approved')
        if assigned:
            eca_reports_qs = eca_reports_qs.filter(program__grade__level=assigned)
        total_spent = eca_reports_qs.aggregate(total=Sum('total_revenue'))['total'] or 0
        ctx['total_spent'] = float(total_spent)

        # Remaining budget
        ctx['remaining_budget'] = max(0, float(verified_income) - float(total_spent))

        # Registration chart data
        from django.db.models import Count
        chart_data = ECARegistration.objects.values('program__name').annotate(total=Count('id')).order_by('-total')
        if assigned:
            chart_data = ECARegistration.objects.filter(program__grade__level=assigned).values('program__name').annotate(total=Count('id')).order_by('-total')
        ctx['chart_labels'] = [d['program__name'] for d in chart_data]
        ctx['chart_data'] = [d['total'] for d in chart_data]

    return render(request, template, ctx)


@login_required
def personal_info_list(request):
    from payments.models import BlastEmailLog

    if request.user.role != 'parent':
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')

    students = request.user.children.all()
    if not students:
        messages.error(request, 'Data siswa tidak ditemukan.')
        return redirect('accounts:dashboard')

    blast_logs = BlastEmailLog.objects.filter(student__in=students)

    unread_logs = blast_logs.filter(is_read=False)
    unread_logs.update(is_read=True, read_at=timezone.now())

    from finance.models import SPPBill, DPP
    from eca.models import ECARegistration
    from finance.models import CambridgeAssessment
    spp_bills = SPPBill.objects.filter(student__in=students).order_by('-year', '-month')
    dpp_list = DPP.objects.filter(student__in=students).order_by('-created_at')
    eca_unpaid = ECARegistration.objects.filter(student__in=students, is_paid=False, status='approved').select_related('program')
    cam_unpaid = CambridgeAssessment.objects.filter(student__in=students, is_paid=False)

    return render(request, 'accounts/personal_info_list.html', {
        'blast_logs': blast_logs,
        'spp_bills': spp_bills,
        'dpp_list': dpp_list,
        'eca_unpaid': eca_unpaid,
        'cam_unpaid': cam_unpaid,
        'paid_count': spp_bills.filter(is_paid=True).count(),
        'unpaid_count': spp_bills.filter(is_paid=False).count(),
    })


@login_required
def profile(request):
    if request.method == 'POST':
        user = request.user
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        alamat = request.POST.get('alamat', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')

        if not email:
            messages.error(request, 'Email wajib diisi.')
            return render(request, 'accounts/profile.html', {'profile_user': user})

        user.email = email
        user.phone = phone
        user.alamat = alamat
        user.first_name = first_name
        user.last_name = last_name
        if request.FILES.get('avatar'):
            user.avatar = request.FILES['avatar']
        user.save()

        if password1 and password2 and password1 == password2:
            user.set_password(password1)
            user.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Password berhasil diubah.')

        messages.success(request, 'Profil berhasil diperbarui.')
        return redirect('accounts:profile')

    return render(request, 'accounts/profile.html', {'profile_user': request.user})


def download_template_xlsx(filename, headers):
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="004085")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64+i)].width = 22
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


@login_required
def download_template(request, template_type):
    if template_type == 'siswa':
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        headers = ['Username', 'Nama', 'Role', 'Jenjang', 'Email', 'HP', 'Password']
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = Font(bold=True, color="004085")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[chr(64+i)].width = 22
        # Sample row
        for i, v in enumerate(['contoh_user', 'Nama Lengkap', 'parent', 'sd', 'email@mail.com', '08123456789', 'siswa123'], 1):
            ws.cell(row=2, column=i, value=v).font = Font(color="888888", italic=True)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=template_users.xlsx'
        wb.save(response)
        return response

    if template_type == 'grade':
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jenjang"
        for i, h in enumerate(['Nama Jenjang', 'Level (sd/smp/sma)'], 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            ws.column_dimensions[chr(64+i)].width = 22
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=template_jenjang.xlsx'
        wb.save(response)
        return response

    if template_type == 'class_grade':
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kelas"
        for i, h in enumerate(['Nama Kelas', 'Nama Jenjang (cocokkan)', 'Nama Tahun Ajaran (cocokkan)'], 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            ws.column_dimensions[chr(64+i)].width = 30
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=template_kelas.xlsx'
        wb.save(response)
        return response

    templates = {
        'spp': ('template_spp.xlsx', ['Jenjang', 'Nominal']),
        'dpp': ('template_dpp.xlsx', ['NIS', 'Kategori', 'Jumlah', 'Keterangan']),
        'eca_types': ('template_eca_types.xlsx', ['Nama Jenis', 'Deskripsi']),
        'eca_programs': ('template_eca_programs.xlsx', ['Nama Program', 'Jenis ECA', 'Jenjang', 'Harga', 'Durasi', 'Max Peserta', 'Jadwal 1', 'Jadwal 2', 'Deskripsi', 'PIC']),
        'activity': ('template_laporan_kegiatan.xlsx', [
            'Judul Kegiatan', 'Tanggal Mulai', 'Tanggal Selesai', 'Deskripsi',
            'Link Google Drive 1', 'Link Google Drive 2', 'Nomor Invoice',
            'Pemasukan', 'Pengeluaran', 'Refleksi', 'Evaluasi'
        ]),
        'eca_scores': ('template_eca_scores.xlsx', ['Nama Siswa', 'Nama Program ECA', 'Score', 'Keterangan']),
    }
    if template_type in templates:
        filename, headers = templates[template_type]
        return download_template_xlsx(filename, headers)
    return HttpResponse('Template not found', status=404)


@login_required
def import_students(request):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')

    if request.method == 'POST' and request.FILES.get('file'):
        from .models import Student, User, Grade, ClassGrade, AcademicYear

        file = request.FILES['file']
        wb = openpyxl.load_workbook(file)

        academic_year = AcademicYear.objects.filter(is_active=True).first()
        if not academic_year:
            messages.error(request, 'Tidak ada tahun ajaran aktif.')
            return redirect('accounts:dashboard')

        imported = 0
        errors = []

        if 'Siswa' in wb.sheetnames:
            ws = wb['Siswa']
        else:
            ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                nisn, nis, full_name, parent_name, parent_username, grade_name, class_name = row[:7]
                if not nis or not full_name:
                    continue
                nis = str(nis).strip()
                full_name = str(full_name).strip()

                parent_user = None
                if parent_username:
                    pu = str(parent_username).strip()
                    parent_user = User.objects.filter(username=pu).first()
                    if not parent_user:
                        parent_user = User.objects.create_user(
                            username=pu,
                            password='siswa123',
                            role='parent',
                            first_name=str(parent_name or '').strip(),
                        )

                grade = Grade.objects.filter(name__iexact=str(grade_name).strip()).first() if grade_name else None
                if not grade:
                    errors.append(f'NIS {nis}: Grade "{grade_name}" tidak ditemukan')
                    continue
                class_grade, _ = ClassGrade.objects.get_or_create(
                    name=str(class_name).strip() if class_name else 'A',
                    grade=grade, academic_year=academic_year
                )
                student, created = Student.objects.update_or_create(
                    nis=nis,
                    defaults={
                        'nisn': str(nisn or '').strip(),
                        'full_name': full_name,
                        'parent': parent_user,
                        'class_grade': class_grade,
                        'is_active': True,
                    }
                )
                if created and not student.user:
                    user = User.objects.create_user(
                        username=nis,
                        password='siswa123',
                        role='parent',
                        first_name=full_name,
                    )
                    student.user = user
                    student.save()
                imported += 1
            except Exception as e:
                errors.append(f'Baris: {e}')

        if 'PIC Teachers' in wb.sheetnames:
            ws2 = wb['PIC Teachers']
            for row in ws2.iter_rows(min_row=2, values_only=True):
                try:
                    name, username, password, level = row[:4]
                    if not username:
                        continue
                    username = str(username).strip()
                    existing = User.objects.filter(username=username).first()
                    if not existing:
                        User.objects.create_user(
                            username=username,
                            password=str(password or 'pic123').strip(),
                            role='pic_teacher',
                            first_name=str(name or '').strip(),
                            assigned_level=str(level or '').strip(),
                        )
                except Exception as e:
                    errors.append(f'PIC: {e}')

        msg = f'Berhasil import {imported} siswa.'
        if errors:
            msg += f' {len(errors)} error.'
        messages.success(request, msg)
        if errors:
            for e in errors[:5]:
                messages.error(request, e)
        return redirect('accounts:dashboard')

    return render(request, 'accounts/import_students.html')


@login_required
def user_list(request):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import User as UserModel
    from django.db.models import Q

    users = UserModel.objects.all()

    # Level filter for TU
    if request.user.role == 'tu' and request.user.assigned_level:
        level_map = {'sd': ['sd'], 'smp': ['smp'], 'sma': ['sma'],
            'sd_smp': ['sd','smp'], 'smp_sma': ['smp','sma'], 'sd_smp_sma': ['sd','smp','sma']}
        allowed = level_map.get(request.user.assigned_level, [])
        users = users.filter(Q(assigned_level__in=allowed) | Q(assigned_level=''))

    # Search
    q = request.GET.get('q', '').strip()
    if q:
        users = users.filter(
            Q(username__icontains=q) |
            Q(first_name__icontains=q) |
            Q(email__icontains=q) |
            Q(role__icontains=q) |
            Q(assigned_level__icontains=q)
        )

    # Role filter
    role_filter = request.GET.get('role', '')
    if role_filter:
        users = users.filter(role=role_filter)

    # Sort
    sort = request.GET.get('sort', 'role')
    dir = request.GET.get('dir', 'asc')
    allowed_sorts = ['username', 'first_name', 'role', 'assigned_level', 'email']
    if sort not in allowed_sorts:
        sort = 'role'
    order = f'-{sort}' if dir == 'desc' else sort
    users = users.order_by(order)

    return render(request, 'accounts/user_list.html', {
        'users': users,
        'filter_q': q,
        'filter_role': role_filter,
        'sort': sort,
        'dir': dir,
        'sort_options': {
            'Username': 'username',
            'Nama': 'first_name',
            'Role': 'role',
            'Jenjang': 'assigned_level',
            'Email': 'email',
        },
    })


@login_required
def reset_password(request, pk):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import User as UserModel
    target = get_object_or_404(UserModel, id=pk)
    if target.role == 'parent':
        default_pass = 'siswa123'
    elif target.role == 'pic_teacher':
        default_pass = 'pic123'
    else:
        default_pass = 'default123'
    target.set_password(default_pass)
    target.save()
    messages.success(request, f'Password {target.username} direset ke: {default_pass}')
    return redirect('accounts:user_list')


@login_required
def create_user(request):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import User as UserModel
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        role = request.POST.get('role', '')
        first_name = request.POST.get('first_name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        assigned_level = request.POST.get('assigned_level', '')

        if not username or not role:
            messages.error(request, 'Username dan role wajib diisi.')
            return redirect('accounts:import_students')

        if UserModel.objects.filter(username=username).exists():
            messages.error(request, f'Username "{username}" sudah digunakan.')
            return redirect('accounts:import_students')

        default_pass = 'siswa123' if role == 'parent' else 'default123'
        if role == 'pic_teacher':
            default_pass = 'pic123'

        UserModel.objects.create_user(
            username=username,
            password=default_pass,
            role=role,
            first_name=first_name,
            email=email,
            phone=phone,
            assigned_level=assigned_level if role in ['kepsek', 'vp_activity', 'pic_teacher', 'eca_director', 'tu'] else '',
        )
        messages.success(request, f'User {username} ({role}) berhasil dibuat. Password default: {default_pass}')
        return redirect('accounts:import_students')

    return redirect('accounts:import_students')


@login_required
def edit_user(request, pk):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import User as UserModel
    target = get_object_or_404(UserModel, id=pk)
    if request.method == 'POST':
        target.first_name = request.POST.get('first_name', '').strip()
        target.email = request.POST.get('email', '').strip()
        target.phone = request.POST.get('phone', '').strip()
        target.role = request.POST.get('role', target.role)
        assigned_level = request.POST.get('assigned_level', '')
        target.assigned_level = assigned_level if assigned_level else target.assigned_level
        target.save()
        messages.success(request, f'User {target.username} berhasil diperbarui.')
        return redirect('accounts:user_list')
    return render(request, 'accounts/edit_user.html', {'u': target})


@login_required
def delete_user(request, pk):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import User as UserModel
    target = get_object_or_404(UserModel, id=pk)
    if target == request.user:
        messages.error(request, 'Tidak dapat menghapus akun sendiri.')
        return redirect('accounts:user_list')
    username = target.username
    target.delete()
    messages.success(request, f'User {username} berhasil dihapus.')
    return redirect('accounts:user_list')


@login_required
def export_users_xlsx(request):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"
    headers = ['Username', 'Nama', 'Role', 'Email', 'HP', 'Password']
    hfill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="004085")
        cell.fill = hfill
        ws.column_dimensions[chr(64+i)].width = 22
    from .models import User as UserModel
    for idx, u in enumerate(UserModel.objects.all().order_by('role', 'username'), 2):
        ws.cell(row=idx, column=1, value=u.username)
        ws.cell(row=idx, column=2, value=u.get_full_name() or '-')
        ws.cell(row=idx, column=3, value=u.get_role_display())
        ws.cell(row=idx, column=4, value=u.email or '-')
        ws.cell(row=idx, column=5, value=u.phone or '-')
        ws.cell(row=idx, column=6, value='***')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=users.xlsx'
    wb.save(response)
    return response


def import_view(model_class, field_map, request, template, redirect_url, title):
    if request.method == 'POST' and request.FILES.get('file'):
        wb = openpyxl.load_workbook(request.FILES['file'])
        ws = wb.active
        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            data = {}
            for i, key in enumerate(field_map):
                if i < len(row) and row[i] is not None:
                    data[key] = row[i]
            if data:
                model_class.objects.create(**data)
                imported += 1
        messages.success(request, f'Berhasil import {imported} {title}.')
        return redirect(redirect_url)
    return render(request, template, {'title': f'Import {title}'})


@login_required
def grade_list(request):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import Grade
    grades = Grade.objects.all()
    return render(request, 'accounts/grade_list.html', {'grades': grades})


@login_required
def create_grade(request):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    if request.method == 'POST':
        from .models import Grade
        name = request.POST.get('name')
        level = request.POST.get('level')
        if name and level:
            Grade.objects.create(name=name, level=level)
            messages.success(request, f'Jenjang {name} berhasil dibuat.')
        return redirect('accounts:grade_list')
    return render(request, 'accounts/create_grade.html')


@login_required
def class_grade_list(request):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import ClassGrade
    classes = ClassGrade.objects.all()
    return render(request, 'accounts/class_grade_list.html', {'classes': classes})


@login_required
def create_class_grade(request):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    if request.method == 'POST':
        from .models import ClassGrade, Grade, AcademicYear
        name = request.POST.get('name')
        grade_id = request.POST.get('grade')
        academic_year_id = request.POST.get('academic_year')
        if name and grade_id and academic_year_id:
            ClassGrade.objects.create(
                name=name,
                grade_id=grade_id,
                academic_year_id=academic_year_id,
            )
            messages.success(request, f'Kelas {name} berhasil dibuat.')
        return redirect('accounts:class_grade_list')
    from .models import Grade, AcademicYear
    grades = Grade.objects.all()
    academic_years = AcademicYear.objects.all()
    return render(request, 'accounts/create_class_grade.html', {'grades': grades, 'academic_years': academic_years})


@login_required
def import_grade(request):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    if request.method == 'POST' and request.FILES.get('file'):
        import openpyxl
        from .models import Grade
        wb = openpyxl.load_workbook(request.FILES['file'])
        ws = wb.active
        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, level = row[0], row[1]
            if name and level:
                Grade.objects.get_or_create(name=str(name), level=str(level))
                imported += 1
        messages.success(request, f'Berhasil import {imported} jenjang.')
        return redirect('accounts:grade_list')
    return render(request, 'accounts/import_grade.html')


@login_required
def delete_grade(request, pk):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import Grade
    grade = get_object_or_404(Grade, id=pk)
    grade.delete()
    messages.success(request, 'Jenjang berhasil dihapus.')
    return redirect('accounts:grade_list')


@login_required
def import_class_grade(request):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    if request.method == 'POST' and request.FILES.get('file'):
        import openpyxl
        from .models import ClassGrade, Grade, AcademicYear
        wb = openpyxl.load_workbook(request.FILES['file'])
        ws = wb.active
        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, grade_name, ay_name = (row[0], row[1], row[2]) if len(row) >= 3 else (row[0], row[1], '')
            if name and grade_name and ay_name:
                grade = Grade.objects.filter(name=str(grade_name)).first()
                ay = AcademicYear.objects.filter(name=str(ay_name)).first()
                if grade and ay:
                    ClassGrade.objects.get_or_create(name=str(name), grade=grade, academic_year=ay)
                    imported += 1
        messages.success(request, f'Berhasil import {imported} kelas.')
        return redirect('accounts:class_grade_list')
    return render(request, 'accounts/import_class_grade.html')


@login_required
def delete_class_grade(request, pk):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import ClassGrade
    cg = get_object_or_404(ClassGrade, id=pk)
    cg.delete()
    messages.success(request, 'Kelas berhasil dihapus.')
    return redirect('accounts:class_grade_list')


@login_required
def grade_promotion(request):
    if request.user.role not in ['admin', 'tu', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')

    from .models import Grade, ClassGrade, AcademicYear, Student

    grades = Grade.objects.all()
    classes = ClassGrade.objects.filter(grade__in=grades)
    active_year = AcademicYear.objects.filter(is_active=True).first()

    students = Student.objects.filter(is_active=True, class_grade__isnull=False).select_related('class_grade__grade', 'class_grade__academic_year')

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'promote_all':
            grade_order = {'sd': 'smp', 'smp': 'sma', 'sma': None}
            promoted = 0
            for student in students:
                if not student.class_grade:
                    continue
                current_level = student.class_grade.grade.level
                next_level_str = grade_order.get(current_level)
                if next_level_str:
                    next_grade = Grade.objects.filter(level=next_level_str).first()
                    if next_grade:
                        next_class = ClassGrade.objects.filter(grade=next_grade, academic_year=active_year).first()
                        if next_class:
                            student.class_grade = next_class
                            student.save()
                            promoted += 1
            messages.success(request, f'{promoted} siswa berhasil dinaikkan kelas.')
            return redirect('accounts:grade_promotion')

        elif action == 'promote_selected':
            student_ids = request.POST.getlist('student_ids')
            target_class_id = request.POST.get('target_class')
            target_class = get_object_or_404(ClassGrade, id=target_class_id)
            count = Student.objects.filter(pk__in=student_ids).update(class_grade=target_class)
            messages.success(request, f'{count} siswa dipindahkan ke {target_class}.')
            return redirect('accounts:grade_promotion')

        elif action == 'graduate_single':
            nisn = request.POST.get('student_nisn')
            Student.objects.filter(nisn=nisn).update(is_active=False, graduated_at=timezone.now())
            messages.success(request, 'Siswa dinyatakan lulus.')
            return redirect('accounts:grade_promotion')

    from collections import defaultdict
    grouped = defaultdict(list)
    for s in students:
        key = str(s.class_grade)
        grouped[key].append(s)

    return render(request, 'accounts/grade_promotion.html', {
        'grades': grades,
        'classes': classes,
        'grouped': dict(grouped),
        'active_year': active_year,
    })


@login_required
def graduation_list(request):
    if request.user.role not in ['admin', 'tu', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import Student, Grade
    final_grades = Grade.objects.filter(name__in=['Kelas 6', 'Kelas 9', 'Kelas 12'])
    students = Student.objects.filter(
        class_grade__grade__in=final_grades,
        is_active=True,
        graduated_at__isnull=True
    ).select_related('class_grade__grade')
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'graduate_selected':
            student_ids = request.POST.getlist('student_ids')
            now = timezone.now()
            Student.objects.filter(nisn__in=student_ids).update(is_active=False, graduated_at=now)
            messages.success(request, f'{len(student_ids)} siswa dinyatakan lulus.')
            return redirect('accounts:graduation_list')
        elif action == 'graduate_all':
            now = timezone.now()
            count = students.update(is_active=False, graduated_at=now)
            messages.success(request, f'{count} siswa dinyatakan lulus.')
            return redirect('accounts:graduation_list')
    return render(request, 'accounts/graduation_list.html', {
        'students': students,
    })


@login_required
def academic_year_list(request):
    if request.user.role not in ['admin', 'tu', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    years = AcademicYear.objects.all().order_by('-start_date')
    return render(request, 'accounts/academic_year_list.html', {'years': years})


@login_required
def academic_year_create(request):
    if request.user.role not in ['admin', 'tu', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    if request.method == 'POST':
        name = request.POST.get('name')
        start = request.POST.get('start_date')
        end = request.POST.get('end_date')
        is_active = request.POST.get('is_active') == 'on'
        AcademicYear.objects.create(name=name, start_date=start, end_date=end, is_active=is_active)
        messages.success(request, 'Tahun ajaran berhasil dibuat.')
        return redirect('accounts:academic_year_list')
    return render(request, 'accounts/academic_year_create.html')


@login_required
def academic_year_edit(request, pk):
    if request.user.role not in ['admin', 'tu', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    year = get_object_or_404(AcademicYear, id=pk)
    if request.method == 'POST':
        year.name = request.POST.get('name')
        year.start_date = request.POST.get('start_date')
        year.end_date = request.POST.get('end_date')
        is_active = request.POST.get('is_active') == 'on'
        year.is_active = is_active
        year.save()
        messages.success(request, 'Tahun ajaran berhasil diupdate.')
        return redirect('accounts:academic_year_list')
    return render(request, 'accounts/academic_year_create.html', {'year': year})


@login_required
def academic_year_delete(request, pk):
    if request.user.role not in ['admin', 'tu', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    AcademicYear.objects.filter(id=pk).delete()
    messages.success(request, 'Tahun ajaran berhasil dihapus.')
    return redirect('accounts:academic_year_list')


from .models import InternalInfo


@login_required
def internal_info_list(request):
    user = request.user
    from payments.models import BlastEmailLog
    from django.db.models import Count
    if user.role in ['admin', 'tu', 'vp_activity', 'kepsek']:
        infos = InternalInfo.objects.all()
        blast_logs = BlastEmailLog.objects.select_related('student').all()
        blast_stats = BlastEmailLog.objects.aggregate(
            total=Count('id'),
            read=Count('id', filter=Q(is_read=True)),
            unread=Count('id', filter=Q(is_read=False)),
        )
    elif user.role == 'parent':
        students = user.children.all()
        if students:
            level = students[0].class_grade.grade.level
            infos = InternalInfo.objects.filter(is_published=True).filter(
                Q(level='all') | Q(level__contains=level))
            blast_logs = BlastEmailLog.objects.filter(student__in=students)
        else:
            infos = InternalInfo.objects.filter(is_published=True, level='all')
            blast_logs = BlastEmailLog.objects.none()
        blast_stats = {}
    elif user.role in ['pic_teacher', 'eca_director']:
        if user.assigned_level:
            levels = user.assigned_level.split(',') if ',' in user.assigned_level else [user.assigned_level]
            q_filter = Q(level='all')
            for lv in levels:
                q_filter |= Q(level__contains=lv.strip())
            infos = InternalInfo.objects.filter(is_published=True).filter(q_filter)
        else:
            infos = InternalInfo.objects.filter(is_published=True, level='all')
        blast_logs = BlastEmailLog.objects.none()
        blast_stats = {}
    else:
        infos = InternalInfo.objects.filter(is_published=True)
        blast_logs = BlastEmailLog.objects.none()
        blast_stats = {}

    return render(request, 'accounts/internal_info_list.html', {
        'infos': infos,
        'blast_logs': blast_logs,
        'blast_stats': blast_stats,
    })


@login_required
def internal_info_create(request):
    if request.user.role not in ['admin', 'tu', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    if request.method == 'POST':
        title = request.POST.get('title')
        content = request.POST.get('content')
        levels = request.POST.getlist('level')
        level = ','.join(levels) if levels else 'all'
        file = request.FILES.get('file')
        google_drive_link = request.POST.get('google_drive_link', '')
        info = InternalInfo.objects.create(
            title=title, content=content, level=level,
            file=file, google_drive_link=google_drive_link,
            created_by=request.user,
        )
        if request.user.role == 'admin':
            info.is_published = True
            info.reviewed_by = request.user
            info.reviewed_at = timezone.now()
            info.save()
            messages.success(request, 'Informasi internal berhasil dipublikasikan.')
        else:
            messages.info(request, 'Informasi internal dibuat, menunggu persetujuan Kepala Sekolah.')
        return redirect('accounts:internal_info_list')
    return render(request, 'accounts/internal_info_create.html')


@login_required
def internal_info_edit(request, pk):
    if request.user.role not in ['admin', 'tu', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    info = get_object_or_404(InternalInfo, id=pk)
    if request.method == 'POST':
        info.title = request.POST.get('title')
        info.content = request.POST.get('content')
        levels = request.POST.getlist('level')
        info.level = ','.join(levels) if levels else 'all'
        if request.FILES.get('file'):
            info.file = request.FILES['file']
        info.google_drive_link = request.POST.get('google_drive_link', '')
        info.is_published = False
        info.reviewed_by = None
        info.reviewed_at = None
        info.save()
        messages.info(request, 'Informasi diupdate, menunggu persetujuan ulang.')
        return redirect('accounts:internal_info_list')
    return render(request, 'accounts/internal_info_create.html', {'info': info})


@login_required
def internal_info_delete(request, pk):
    if request.user.role not in ['admin', 'tu', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    InternalInfo.objects.filter(id=pk).delete()
    messages.success(request, 'Informasi berhasil dihapus.')
    return redirect('accounts:internal_info_list')


@login_required
def internal_info_approve(request, pk):
    if request.user.role not in ['admin', 'kepsek']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    info = get_object_or_404(InternalInfo, id=pk)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'approve':
            info.is_published = True
            info.reviewed_by = request.user
            info.reviewed_at = timezone.now()
            info.save()
            messages.success(request, 'Informasi disetujui dan dipublikasikan.')
        elif action == 'reject':
            info.delete()
            messages.info(request, 'Informasi ditolak dan dihapus.')
        return redirect('accounts:internal_info_list')
    return render(request, 'accounts/internal_info_approve.html', {'info': info})


@login_required
def blast_statistics(request):
    if request.user.role not in ['admin', 'kepsek', 'vp_activity', 'eca_director']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from payments.models import BlastEmailLog
    from django.db.models import Count, Q
    from accounts.models import User, Student

    user = request.user
    base_qs = BlastEmailLog.objects.all()

    if user.role != 'admin':
        if user.assigned_level:
            level_map = {'sd': ['sd'], 'smp': ['smp'], 'sma': ['sma'],
                'sd_smp': ['sd','smp'], 'smp_sma': ['smp','sma'], 'sd_smp_sma': ['sd','smp','sma']}
            allowed = level_map.get(user.assigned_level, [])
            base_qs = base_qs.filter(student__class_grade__grade__level__in=allowed)

    total = base_qs.count()
    read_count = base_qs.filter(is_read=True).count()
    unread_count = base_qs.filter(is_read=False).count()

    per_student = base_qs.values(
        'student_id', 'student__full_name', 'student__nisn', 'student__class_grade__name'
    ).annotate(
        total_sent=Count('id'),
        read=Count('id', filter=Q(is_read=True)),
        unread=Count('id', filter=Q(is_read=False)),
    ).order_by('-total_sent')

    by_payment_type = base_qs.values('payment_type').annotate(
        total=Count('id'),
        read=Count('id', filter=Q(is_read=True)),
    ).order_by('payment_type')

    return render(request, 'accounts/blast_statistics.html', {
        'total': total,
        'read_count': read_count,
        'unread_count': unread_count,
        'per_student': per_student,
        'by_payment_type': by_payment_type,
    })
