from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from .models import ECAType, ECAProgram, ECARegistration, ECAReport
from accounts.models import AcademicYear, Grade


def director_required(view_func):
    def wrapper(request, *args, **kwargs):
        if request.user.role not in ['admin', 'eca_director']:
            messages.error(request, 'Akses ditolak.')
            return redirect('accounts:dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
@director_required
def eca_type_list(request):
    types = ECAType.objects.all()
    return render(request, 'eca/type_list.html', {'types': types})


@login_required
@director_required
def create_eca_type(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        ECAType.objects.create(name=name, description=description)
        messages.success(request, 'Jenis ECA berhasil dibuat.')
        return redirect('eca:type_list')
    return render(request, 'eca/create_type.html')


@login_required
@director_required
def edit_eca_type(request, pk):
    t = get_object_or_404(ECAType, id=pk)
    if request.method == 'POST':
        t.name = request.POST.get('name')
        t.description = request.POST.get('description', '')
        t.save()
        messages.success(request, 'Jenis ECA berhasil diupdate.')
        return redirect('eca:type_list')
    return render(request, 'eca/edit_type.html', {'type_obj': t})


@login_required
@director_required
def delete_eca_type(request, pk):
    t = get_object_or_404(ECAType, id=pk)
    t.delete()
    messages.success(request, 'Jenis ECA berhasil dihapus.')
    return redirect('eca:type_list')


@login_required
@director_required
@login_required
@director_required
def program_list(request):
    programs = ECAProgram.objects.all()
    return render(request, 'eca/program_list.html', {'programs': programs})


@login_required
@director_required
def edit_program(request, pk):
    program = get_object_or_404(ECAProgram, id=pk)
    if request.method == 'POST':
        program.eca_type_id = request.POST.get('eca_type')
        program.name = request.POST.get('name')
        program.description = request.POST.get('description', '')
        program.grade_id = request.POST.get('grade')
        program.price = request.POST.get('price')
        program.duration = request.POST.get('duration')
        program.max_participants = request.POST.get('max_participants', 0)
        program.schedule = request.POST.get('schedule', '')
        program.schedule2 = request.POST.get('schedule2', '')
        program.pic = request.POST.get('pic', '')
        program.coach_profile = request.POST.get('coach_profile', '')
        program.rules = request.POST.get('rules', '')
        program.additional_info = request.POST.get('additional_info', '')
        program.location = request.POST.get('location', '')
        program.meeting_link = request.POST.get('meeting_link', '')
        program.is_open = request.POST.get('is_open') == 'on'
        program.save()
        messages.success(request, 'Program ECA berhasil diupdate.')
        return redirect('eca:program_list')
    types = ECAType.objects.filter(is_active=True)
    grades = Grade.objects.all()
    return render(request, 'eca/create_program.html', {
        'types': types, 'grades': grades, 'program': program, 'edit_mode': True
    })


@login_required
@director_required
def create_program(request):
    if request.method == 'POST':
        eca_type_id = request.POST.get('eca_type')
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        grade_id = request.POST.get('grade')
        price = request.POST.get('price')
        duration = request.POST.get('duration')
        max_participants = request.POST.get('max_participants', 0)
        schedule = request.POST.get('schedule', '')
        pic = request.POST.get('pic', '')

        academic_year = AcademicYear.objects.filter(is_active=True).first()
        if not academic_year:
            messages.error(request, 'Tidak ada tahun ajaran aktif.')
            return redirect('eca:program_list')

        ECAProgram.objects.create(
            eca_type_id=eca_type_id,
            name=name,
            description=description,
            grade_id=grade_id,
            academic_year=academic_year,
            price=price,
            duration=duration,
            max_participants=max_participants,
            schedule=schedule,
            schedule2=request.POST.get('schedule2', ''),
            pic=pic,
            coach_profile=request.POST.get('coach_profile', ''),
            rules=request.POST.get('rules', ''),
            additional_info=request.POST.get('additional_info', ''),
            location=request.POST.get('location', ''),
            meeting_link=request.POST.get('meeting_link', ''),
            is_approved=(request.user.role == 'admin'),
            approved_by=request.user if request.user.role == 'admin' else None,
        )
        messages.success(request, 'Program ECA berhasil dibuat.')
        return redirect('eca:program_list')

    types = ECAType.objects.filter(is_active=True)
    grades = Grade.objects.all()
    return render(request, 'eca/create_program.html', {'types': types, 'grades': grades})


@login_required
def registration_list(request):
    from accounts.models import Student
    import json
    student = None
    students = None
    if request.user.role == 'parent':
        students = request.user.children.all()
        if not students:
            messages.error(request, 'Data siswa tidak ditemukan. Hubungi Tata Usaha.')
            return redirect('accounts:dashboard')
        registrations = ECARegistration.objects.filter(student__in=students)
    elif request.user.role in ['admin', 'eca_director']:
        registrations = ECARegistration.objects.all()
    else:
        registrations = ECARegistration.objects.none()
    open_programs = ECAProgram.objects.filter(is_open=True, is_approved=True) if students else ECAProgram.objects.none()

    # Chart data
    from django.db.models import Count
    from django.db.models.functions import TruncDate
    prog_data = ECARegistration.objects.values('program__name').annotate(total=Count('id')).order_by('-total')
    chart_programs = [d['program__name'] for d in prog_data]
    chart_prog_counts = [d['total'] for d in prog_data]

    status_data = ECARegistration.objects.values('status').annotate(total=Count('id'))
    chart_status_labels = {'pending': 'Pending', 'approved': 'Disetujui', 'rejected': 'Ditolak'}
    chart_status = [chart_status_labels.get(d['status'], d['status']) for d in status_data]
    chart_status_counts = [d['total'] for d in status_data]

    time_data = ECARegistration.objects.annotate(date=TruncDate('registered_at')).values('date').annotate(total=Count('id')).order_by('date')
    chart_dates = [d['date'].strftime('%d/%m') if d['date'] else '' for d in time_data]
    chart_time_counts = [d['total'] for d in time_data]

    kelas_data = ECARegistration.objects.values('student__class_grade__grade__name').annotate(total=Count('id')).order_by('student__class_grade__grade__name')
    chart_kelas = [d['student__class_grade__grade__name'] for d in kelas_data]
    chart_kelas_counts = [d['total'] for d in kelas_data]

    # Bulk payment info
    unpaid_total = 0
    unpaid_count = 0
    reg_prices = []
    if students:
        unpaid_regs = ECARegistration.objects.filter(student__in=students, status='approved', is_paid=False)
        unpaid_total = sum(r.program.price for r in unpaid_regs)
        unpaid_count = unpaid_regs.count()
        reg_prices = [{'reg_id': r.id, 'program': r.program.name, 'price': r.program.price} for r in unpaid_regs]

    return render(request, 'eca/registration_list.html', {
        'registrations': registrations,
        'students': students,
        'open_programs': open_programs,
        'unpaid_total': unpaid_total,
        'unpaid_count': unpaid_count,
        'reg_prices': reg_prices,
        'chart_programs': json.dumps(chart_programs),
        'chart_prog_counts': json.dumps(chart_prog_counts),
        'chart_status': json.dumps(chart_status),
        'chart_status_counts': json.dumps(chart_status_counts),
        'chart_dates': json.dumps(chart_dates),
        'chart_time_counts': json.dumps(chart_time_counts),
        'chart_kelas': json.dumps(chart_kelas),
        'chart_kelas_counts': json.dumps(chart_kelas_counts),
    })


@login_required
@director_required
def approve_registration(request, reg_id):
    registration = get_object_or_404(ECARegistration, id=reg_id)
    if registration.status == 'approved':
        registration.status = 'pending'
        registration.reviewed_by = request.user
        registration.reviewed_at = timezone.now()
        registration.save()
        messages.warning(request, 'Persetujuan pendaftaran dibatalkan.')
    else:
        registration.status = 'approved'
        registration.reviewed_by = request.user
        registration.reviewed_at = timezone.now()
        registration.save()
        messages.success(request, 'Pendaftaran disetujui.')
    return redirect('eca:registration_list')


@login_required
@director_required
def approve_all_registrations(request):
    count = ECARegistration.objects.filter(status='pending').update(
        status='approved', reviewed_by=request.user, reviewed_at=timezone.now()
    )
    messages.success(request, f'{count} pendaftaran berhasil disetujui.')
    return redirect('eca:registration_list')


@login_required
def registration_management(request):
    if request.user.role not in ['admin', 'eca_director', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    status_filter = request.GET.get('status', 'pending')
    qs = ECARegistration.objects.select_related('student', 'program', 'reviewed_by').all()
    
    # Level filter for TU
    if request.user.role == 'tu' and request.user.assigned_level:
        level_map = {'sd': ['sd'], 'smp': ['smp'], 'sma': ['sma'],
            'sd_smp': ['sd','smp'], 'smp_sma': ['smp','sma'], 'sd_smp_sma': ['sd','smp','sma']}
        allowed = level_map.get(request.user.assigned_level, [])
        if allowed:
            qs = qs.filter(student__class_grade__grade__level__in=allowed)
    
    if status_filter in ['pending', 'approved', 'rejected']:
        qs = qs.filter(status=status_filter)

    if request.method == 'POST':
        reg_id = request.POST.get('reg_id')
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')
        registration = get_object_or_404(ECARegistration, id=reg_id)
        if action == 'approve':
            if registration.status == 'approved':
                registration.status = 'pending'
                registration.rejection_notes = ''
                registration.reviewed_by = None
                registration.reviewed_at = None
                registration.save()
                messages.info(request, f'Pendaftaran {registration.student.full_name} - {registration.program.name} dikembalikan ke pending.')
            else:
                registration.status = 'approved'
                registration.rejection_notes = ''
                registration.reviewed_by = request.user
                registration.reviewed_at = timezone.now()
                registration.save()
                messages.success(request, f'Pendaftaran {registration.student.full_name} - {registration.program.name} disetujui.')
        elif action == 'reject':
            registration.status = 'rejected'
            registration.rejection_notes = notes
            registration.reviewed_by = request.user
            registration.reviewed_at = timezone.now()
            registration.save()
            messages.success(request, f'Pendaftaran {registration.student.full_name} - {registration.program.name} ditolak.')
        elif action == 'delete':
            registration.delete()
            messages.success(request, 'Pendaftaran dihapus.')
        elif action == 'mark_paid':
            registration.is_paid = True
            registration.paid_at = timezone.now()
            registration.save()
            from payments.models import PaymentTransaction
            import uuid
            order_id = f'ECA-MANUAL-{registration.student.nisn}-{uuid.uuid4().hex[:8].upper()}'
            PaymentTransaction.objects.create(
                student=registration.student,
                transaction_type='eca',
                amount=registration.program.price,
                status='success',
                midtrans_order_id=order_id,
                paid_at=timezone.now(),
                payment_method=request.POST.get('payment_method', 'manual'),
                invoice_number=f'INV/ECA/{registration.student.nisn}/{uuid.uuid4().hex[:6].upper()}',
            )
            messages.success(request, f'Pembayaran ECA {registration.student.full_name} dikonfirmasi lunas.')
        elif action == 'mark_unpaid':
            registration.is_paid = False
            registration.paid_at = None
            registration.save()
            messages.info(request, f'Pembayaran ECA {registration.student.full_name} dikembalikan ke belum bayar.')
        elif action == 'upload_proof':
            if request.FILES.get('payment_proof'):
                from payments.models import PaymentProof
                from django.core.files.storage import default_storage
                proof = PaymentProof.objects.create(
                    student=registration.student,
                    image=request.FILES['payment_proof'],
                    description=notes,
                )
                registration.is_paid = True
                registration.paid_at = timezone.now()
                registration.save()
                messages.success(request, f'Bukti bayar ECA {registration.student.full_name} diupload & dikonfirmasi.')
            else:
                messages.error(request, 'Pilih file bukti pembayaran.')
        return redirect('eca:registration_management')

    context = {
        'registrations': qs,
        'current_status': status_filter,
        'status_count': {
            'pending': ECARegistration.objects.filter(status='pending').count(),
            'approved': ECARegistration.objects.filter(status='approved').count(),
            'rejected': ECARegistration.objects.filter(status='rejected').count(),
        },
    }
    return render(request, 'eca/registration_management.html', context)


@login_required
@director_required
def approve_program(request, pk):
    program = get_object_or_404(ECAProgram, id=pk)
    if program.is_approved:
        program.is_approved = False
        program.approved_by = None
        messages.warning(request, f'Persetujuan program {program.name} dibatalkan.')
    else:
        program.is_approved = True
        program.approved_by = request.user
        messages.success(request, f'Program {program.name} disetujui.')
    program.save()
    return redirect('eca:program_list')


@login_required
@director_required
def eca_setting(request):
    from .models import ECASetting
    setting = ECASetting.objects.first()
    if not setting:
        setting = ECASetting.objects.create(max_selection=3)
    if request.method == 'POST':
        max_sel = request.POST.get('max_selection', 3)
        setting.max_selection = int(max_sel)
        setting.save()
        messages.success(request, 'Pengaturan ECA berhasil diupdate.')
        return redirect('eca:registration_list')
    return render(request, 'eca/eca_setting.html', {'setting': setting})


@login_required
def register_eca(request):
    if request.user.role != 'parent':
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from accounts.models import Student
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        program_id = request.POST.get('program_id')
        student = get_object_or_404(Student, pk=student_id)
        program = get_object_or_404(ECAProgram, id=program_id, is_approved=True, is_open=True)

        if ECARegistration.objects.filter(student=student, program=program).exists():
            messages.error(request, 'Sudah terdaftar di program ini.')
            return redirect('eca:registration_list')

        from .models import ECASetting
        setting = ECASetting.objects.first()
        max_sel = setting.max_selection if setting else 3
        current_approved = ECARegistration.objects.filter(student=student, status='approved').count()
        if current_approved >= max_sel:
            messages.error(request, f'Maksimal {max_sel} program ECA per siswa. Selesaikan atau batalkan program lain dahulu.')
            return redirect('eca:registration_list')

        approved_count = ECARegistration.objects.filter(program=program, status='approved').count()
        if program.max_participants > 0 and approved_count >= program.max_participants:
            messages.error(request, f'Kuota program {program.name} sudah penuh.')
            return redirect('eca:registration_list')

        # Jika sudah punya approved ECA → pending + butuh approval director
        has_approved = ECARegistration.objects.filter(student=student, status='approved').exists()
        if has_approved:
            ECARegistration.objects.create(student=student, program=program, status='pending')
            messages.info(request, 'Pendaftaran perlu disetujui ECA Director.')
            return redirect('eca:registration_list')

        # First ECA → auto-approve + generate VA payment
        ECARegistration.objects.create(student=student, program=program, status='approved')
        from payments.models import PaymentTransaction
        from uuid import uuid4
        approved_regs = ECARegistration.objects.filter(student=student, status='approved', is_paid=False)
        total_amount = sum(r.program.price for r in approved_regs)
        va = f'VA-ECA-{student.nis}-{uuid4().hex[:6].upper()}'

        existing_tx = PaymentTransaction.objects.filter(
            student=student, transaction_type='eca', status='pending'
        ).first()
        if existing_tx:
            existing_tx.amount = total_amount
            existing_tx.virtual_account = va
            existing_tx.save()
        else:
            order_id = f'ECA-{student.nisn}-{uuid4().hex[:8].upper()}'
            PaymentTransaction.objects.create(
                student=student,
                transaction_type='eca',
                amount=total_amount,
                midtrans_order_id=order_id,
                virtual_account=va,
                status='pending',
            )
        messages.success(request, f'Pendaftaran {program.name} berhasil otomatis disetujui.')
        return redirect('eca:registration_list')
    return redirect('eca:registration_list')


@login_required
def export_eca_report_xlsx(request):
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan ECA"
    headers = ['Program', 'Periode', 'Siswa', 'Lunas', 'Revenue', 'Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in ECAReport.objects.all():
        ws.append([
            r.program.name, r.get_period_display(),
            r.total_students, r.total_paid,
            int(r.total_revenue or 0), r.get_status_display()
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=laporan_eca_komprehensif.xlsx'
    wb.save(response)
    return response


@login_required
@director_required
def download_eca_report(request):
    return render(request, 'eca/download_report.html')


@login_required
def export_participants_xlsx(request):
    if request.user.role not in ['admin', 'eca_director', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import ECAProgram, ECARegistration
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    lf = {}
    if request.user.assigned_level:
        level_map = {'sd': ['sd'], 'smp': ['smp'], 'sma': ['sma'],
            'sd_smp': ['sd','smp'], 'smp_sma': ['smp','sma'], 'sd_smp_sma': ['sd','smp','sma']}
        allowed = level_map.get(request.user.assigned_level, [])
        if allowed:
            lf = {'grade__level__in': allowed}

    wb = openpyxl.Workbook()
    programs = ECAProgram.objects.filter(**lf).prefetch_related('registrations__student__class_grade__grade')

    for prog in programs:
        ws = wb.create_sheet(title=prog.name[:31]) if programs.count() > 1 else wb.active
        if programs.first() == prog and programs.count() == 1:
            ws.title = prog.name[:31]
        ws.cell(row=1, column=1, value=f'Peserta {prog.name}').font = Font(bold=True, size=12)
        ws.cell(row=2, column=1, value=f'Jenjang: {prog.grade.get_level_display()}').font = Font(italic=True, color='666666')
        ws.cell(row=3, column=1, value=f'Harga: Rp{int(prog.price):,} | Jadwal: {prog.schedule} | PIC: {prog.pic}').font = Font(italic=True, color='666666')

        headers = ['No', 'Nama Siswa', 'NIS', 'NISN', 'Kelas', 'Tgl Daftar', 'Status', 'Pembayaran']
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        regs = prog.registrations.select_related('student__class_grade__grade').all()
        for i, reg in enumerate(regs, 1):
            row = i + 5
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=reg.student.full_name)
            ws.cell(row=row, column=3, value=reg.student.nis)
            ws.cell(row=row, column=4, value=reg.student.nisn)
            ws.cell(row=row, column=5, value=str(reg.student.class_grade))
            ws.cell(row=row, column=6, value=reg.registered_at.strftime('%d/%m/%Y') if reg.registered_at else '-')
            ws.cell(row=row, column=7, value=reg.get_status_display())
            ws.cell(row=row, column=8, value='Lunas' if reg.is_paid else 'Belum')

        for col in range(1, 9):
            ws.column_dimensions[chr(64 + col)].width = 20

    if wb.sheetnames[0] == 'Sheet' and len(wb.sheetnames) == 1:
        del wb['Sheet']

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=peserta_eca_all.xlsx'
    wb.save(response)
    return response


@login_required
def export_participants_pdf(request):
    if request.user.role not in ['admin', 'eca_director', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import ECAProgram, ECARegistration
    from django.template.loader import render_to_string
    from xhtml2pdf import pisa
    from django.utils import timezone

    lf = {}
    if request.user.assigned_level:
        level_map = {'sd': ['sd'], 'smp': ['smp'], 'sma': ['sma'],
            'sd_smp': ['sd','smp'], 'smp_sma': ['smp','sma'], 'sd_smp_sma': ['sd','smp','sma']}
        allowed = level_map.get(request.user.assigned_level, [])
        if allowed:
            lf = {'grade__level__in': allowed}

    programs = ECAProgram.objects.filter(**lf).prefetch_related('registrations__student__class_grade__grade')
    sections = []
    for prog in programs:
        rows = []
        for i, reg in enumerate(prog.registrations.all(), 1):
            rows.append({'no': i, 'nama': reg.student.full_name, 'nis': reg.student.nis,
                'kelas': str(reg.student.class_grade), 'tgl': reg.registered_at.strftime('%d/%m/%Y') if reg.registered_at else '-',
                'status': reg.get_status_display(), 'bayar': 'Lunas' if reg.is_paid else 'Belum'})
        sections.append({'name': prog.name, 'grade': prog.grade.get_level_display(),
            'schedule': prog.schedule, 'price': f'Rp{int(prog.price):,}', 'rows': rows})

    html = render_to_string('eca/pdf_participants.html', {
        'sections': sections,
        'generated_at': timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M'),
    })
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=peserta_eca_all.pdf'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Gagal membuat PDF', status=500)
    return response


@login_required
@director_required
def import_types(request):
    if request.method == 'POST' and request.FILES.get('file'):
        import openpyxl
        wb = openpyxl.load_workbook(request.FILES['file'])
        ws = wb.active
        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            name, desc = row[0], row[1] if len(row) > 1 else ''
            ECAType.objects.update_or_create(
                name=str(name),
                defaults={'description': str(desc or ''), 'is_active': True}
            )
            imported += 1
        messages.success(request, f'Berhasil import {imported} jenis ECA.')
        return redirect('eca:type_list')
    return render(request, 'eca/import_types.html')


@login_required
@director_required
def import_programs(request):
    if request.method == 'POST' and request.FILES.get('file'):
        import openpyxl
        from accounts.models import Grade, AcademicYear

        wb = openpyxl.load_workbook(request.FILES['file'])
        ws = wb.active
        academic_year = AcademicYear.objects.filter(is_active=True).first()
        if not academic_year:
            messages.error(request, 'Tidak ada tahun ajaran aktif.')
            return redirect('eca:program_list')

        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            name = row[0]
            eca_type_name = row[1] if len(row) > 1 else ''
            grade_name = row[2] if len(row) > 2 else ''
            price = row[3] if len(row) > 3 else 0
            duration = row[4] if len(row) > 4 else '1month'
            max_part = row[5] if len(row) > 5 else 0
            schedule = row[6] if len(row) > 6 else ''
            schedule2 = row[7] if len(row) > 7 else ''
            description = row[8] if len(row) > 8 else ''
            pic = row[9] if len(row) > 9 else ''

            eca_type = ECAType.objects.filter(name__iexact=str(eca_type_name)).first()
            grade = Grade.objects.filter(name__iexact=str(grade_name)).first()
            if name and eca_type and grade and price:
                ECAProgram.objects.create(
                    eca_type=eca_type,
                    name=str(name),
                    grade=grade,
                    academic_year=academic_year,
                    price=int(price),
                    duration=str(duration or '1month'),
                    max_participants=int(max_part or 0),
                    schedule=str(schedule or ''),
                    schedule2=str(schedule2 or ''),
                    description=str(description or ''),
                    pic=str(pic or ''),
                )
                imported += 1
        messages.success(request, f'Berhasil import {imported} program ECA.')
        return redirect('eca:program_list')
    return render(request, 'eca/import_programs.html')


@login_required
def score_list(request):
    programs = ECAProgram.objects.all()
    from .models import ECAScore
    total_scores = ECAScore.objects.count()
    total_students = ECAScore.objects.values('student').distinct().count()
    return render(request, 'eca/score_list.html', {
        'programs': programs,
        'total_scores': total_scores,
        'total_students': total_students,
    })


@login_required
def score_program(request, program_id):
    program = get_object_or_404(ECAProgram, id=program_id)
    from .models import ECAScore

    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        score = request.POST.get('score')
        notes = request.POST.get('notes', '')
        if student_id and score:
            ECAScore.objects.update_or_create(
                student_id=student_id,
                program=program,
                defaults={'score': score, 'notes': notes}
            )
            messages.success(request, 'Nilai berhasil disimpan.')

    from .models import ECAScore
    registrations = ECARegistration.objects.filter(program=program, status='approved')
    score_map = {s.student_id: s for s in ECAScore.objects.filter(program=program)}
    rows = []
    for i, reg in enumerate(registrations, 1):
        s = score_map.get(reg.student_id)
        rows.append({
            'num': i,
            'student': reg.student,
            'grade': str(reg.student.class_grade.grade),
            'duration': program.get_duration_display(),
            'score': s.score if s else None,
            'notes': s.notes if s else '',
            'student_id': reg.student_id,
        })
    return render(request, 'eca/score_program.html', {
        'program': program,
        'rows': rows,
    })


@login_required
@director_required
def download_score_template(request, program_id):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    program = get_object_or_404(ECAProgram, id=program_id)
    registrations = ECARegistration.objects.filter(program=program, status='approved')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Scores {program.name}"

    headers = ['Nama Siswa', 'Jenjang', 'Nama ECA', 'Waktu', 'Score', 'Keterangan']
    hfill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="004085")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64+i)].width = 22

    for i, reg in enumerate(registrations, 2):
        ws.cell(row=i, column=1, value=reg.student.full_name)
        ws.cell(row=i, column=2, value=str(reg.student.class_grade.grade))
        ws.cell(row=i, column=3, value=program.name)
        ws.cell(row=i, column=4, value=program.get_duration_display())
        ws.cell(row=i, column=5, value='')
        ws.cell(row=i, column=6, value='')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=scores_{program.name}.xlsx'
    wb.save(response)
    return response


@login_required
@director_required
def import_all_scores(request):
    if request.method == 'POST' and request.FILES.get('file'):
        import openpyxl
        from .models import ECAScore
        from accounts.models import Student

        wb = openpyxl.load_workbook(request.FILES['file'])
        ws = wb.active
        imported = 0
        errors = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            nama, nama_program, score, keterangan = (row[0], row[1], row[2], row[3]) if len(row) >= 4 else (row[0], row[1], row[2], '')
            if not nama or not nama_program or score is None:
                errors.append(f'Baris tidak lengkap: {nama}, {nama_program}, {score}')
                continue
            student = Student.objects.filter(full_name__iexact=str(nama).strip()).first()
            if not student:
                errors.append(f'Siswa tidak ditemukan: {nama}')
                continue
            program = ECAProgram.objects.filter(name__iexact=str(nama_program).strip()).first()
            if not program:
                errors.append(f'Program tidak ditemukan: {nama_program}')
                continue
            ECAScore.objects.update_or_create(
                student=student,
                program=program,
                defaults={'score': float(score), 'notes': str(keterangan or '')}
            )
            imported += 1
        msg = f'Berhasil import {imported} nilai.'
        if errors:
            msg += f' {len(errors)} error: ' + '; '.join(errors[:5])
        messages.success(request, msg)
        return redirect('eca:score_list')
    return render(request, 'eca/import_scores_all.html')


@login_required
@director_required
def download_all_scores_template(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from .models import ECAScore

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Nilai Semua ECA"

    headers = ['Nama Siswa', 'Nama Program ECA', 'Score', 'Keterangan']
    hfill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="004085")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64+i)].width = 30

    programs = ECAProgram.objects.filter(is_open=True)
    row = 2
    for p in programs:
        regs = ECARegistration.objects.filter(program=p, status='approved')
        for reg in regs:
            ws.cell(row=row, column=1, value=reg.student.full_name)
            ws.cell(row=row, column=2, value=p.name)
            ws.cell(row=row, column=3, value='')
            ws.cell(row=row, column=4, value='')
            row += 1

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=template_nilai_semua_eca.xlsx'
    wb.save(response)
    return response


@login_required
@director_required
def generate_dummy_scores(request):
    from .models import ECAScore
    import random
    programs = ECAProgram.objects.all()
    count = 0
    for p in programs:
        regs = ECARegistration.objects.filter(program=p, status='approved')
        for reg in regs:
            if not ECAScore.objects.filter(student=reg.student, program=p).exists():
                ECAScore.objects.create(
                    student=reg.student,
                    program=p,
                    score=random.uniform(60, 100),
                    notes='Nilai dummy'
                )
                count += 1
    messages.success(request, f'{count} nilai dummy berhasil dibuat.')
    return redirect('eca:score_list')


@login_required
@director_required
def import_scores(request, program_id):
    program = get_object_or_404(ECAProgram, id=program_id)
    if request.method == 'POST' and request.FILES.get('file'):
        import openpyxl
        from .models import ECAScore
        from accounts.models import Student

        wb = openpyxl.load_workbook(request.FILES['file'])
        ws = wb.active
        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            nama, jenjang, nama_eca, waktu, score, keterangan = row[0], row[1], row[2], row[3], row[4], row[5] if len(row) > 5 else ''
            if not nama or not score:
                continue
            student = Student.objects.filter(full_name__iexact=str(nama).strip()).first()
            if student:
                ECAScore.objects.update_or_create(
                    student=student,
                    program=program,
                    defaults={'score': float(score), 'notes': str(keterangan or '')}
                )
                imported += 1
        messages.success(request, f'Berhasil import {imported} nilai.')
        return redirect('eca:score_program', program_id=program.id)
    return render(request, 'eca/import_scores.html', {'program': program})


def _check_level_access(user, grade):
    if user.role == 'admin' or not user.assigned_level:
        return True
    level_map = {
        'sd': ['sd'], 'smp': ['smp'], 'sma': ['sma'],
        'sd_smp': ['sd', 'smp'], 'smp_sma': ['smp', 'sma'],
        'sd_smp_sma': ['sd', 'smp', 'sma'],
    }
    allowed = level_map.get(user.assigned_level, [])
    if not allowed:
        return True
    return grade.level in allowed


@login_required
def eca_report_list(request):
    if request.user.role not in ['admin', 'kepsek', 'vp_activity', 'eca_director', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from eca.models import ECAPayment
    from django.db.models import Sum, Q
    reports = ECAReport.objects.all()
    if request.user.role in ['kepsek', 'vp_activity']:
        reports = [r for r in reports if _check_level_access(request.user, r.program.grade)]

    import json

    # Build filtered QS for level-aware queries
    eca_qs = ECAReport.objects.all()
    if request.user.role in ['kepsek', 'vp_activity'] and request.user.assigned_level:
        level_ids = [r.id for r in eca_qs if _check_level_access(request.user, r.program.grade)]
        eca_qs = ECAReport.objects.filter(id__in=level_ids)
        pay_qs = ECAPayment.objects.filter(registration__program__grade__level=request.user.assigned_level)
    else:
        pay_qs = ECAPayment.objects.all()

    # Financial overview
    total_revenue = eca_qs.filter(status='approved').aggregate(t=Sum('total_revenue'))['t'] or 0
    total_billed = pay_qs.aggregate(t=Sum('amount'))['t'] or 0
    total_collected = pay_qs.filter(is_paid=True, payment_status='approved').aggregate(t=Sum('paid_amount'))['t'] or 0
    outstanding = total_billed - total_collected

    # Contribution per program (doughnut)
    contrib = pay_qs.filter(is_paid=True, payment_status='approved').values('registration__program__name').annotate(total=Sum('paid_amount')).order_by('-total')
    contrib_programs = [d['registration__program__name'] for d in contrib]
    contrib_amounts = [float(d['total']) for d in contrib]

    # Comparison per program (bar)
    comp = pay_qs.values('registration__program__name').annotate(
        billed=Sum('amount'),
        collected=Sum('paid_amount', filter=Q(is_paid=True, payment_status='approved'))
    ).order_by('registration__program__name')
    comp_programs = [d['registration__program__name'] for d in comp]
    comp_billed = [float(d['billed'] or 0) for d in comp]
    comp_collected = [float(d['collected'] or 0) for d in comp]

    # Chart: revenue per program from reports
    prog_rev = eca_qs.values('program__name').annotate(total_rev=Sum('total_revenue')).order_by('-total_rev')
    rev_programs = [d['program__name'] for d in prog_rev]
    rev_amounts = [float(d['total_rev'] or 0) for d in prog_rev]

    # Chart: status distribution
    status_counts = {}
    for r in reports:
        s = r.get_status_display()
        status_counts[s] = status_counts.get(s, 0) + 1
    chart_rpt_status = list(status_counts.keys())
    chart_rpt_counts = list(status_counts.values())

    # Chart: total students per program
    prog_students = eca_qs.values('program__name').annotate(total=Sum('total_students')).order_by('-total')
    stud_programs = [d['program__name'] for d in prog_students]
    stud_counts = [int(d['total'] or 0) for d in prog_students]

    context = {
        'reports': reports,
        'rev_programs': json.dumps(rev_programs),
        'rev_amounts': json.dumps(rev_amounts),
        'chart_rpt_status': json.dumps(chart_rpt_status),
        'chart_rpt_counts': json.dumps(chart_rpt_counts),
        'stud_programs': json.dumps(stud_programs),
        'stud_counts': json.dumps(stud_counts),
        'total_revenue': total_revenue,
        'total_billed': total_billed,
        'total_collected': total_collected,
        'outstanding': outstanding,
        'contrib_programs': json.dumps(contrib_programs),
        'contrib_amounts': json.dumps(contrib_amounts),
        'comp_programs': json.dumps(comp_programs),
        'comp_billed': json.dumps(comp_billed),
        'comp_collected': json.dumps(comp_collected),
    }
    return render(request, 'eca/report_list.html', context)


@login_required
def review_eca_report(request, pk):
    if request.user.role not in ['admin', 'kepsek', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    report = get_object_or_404(ECAReport, id=pk)
    if not _check_level_access(request.user, report.program.grade):
        messages.error(request, 'Akses terbatas sesuai jenjang.')
        return redirect('eca:eca_report_list')

    if request.method == 'POST':
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')
        from django.utils import timezone
        if action == 'approved':
            report.status = 'approved'
            report.reviewed_by = request.user
            report.reviewed_at = timezone.now()
            messages.success(request, 'Laporan ECA disetujui.')
        elif action == 'revised':
            report.status = 'revised'
            report.rejection_notes = notes
            messages.success(request, 'Laporan ECA perlu direvisi.')
        elif action == 'rejected':
            report.status = 'rejected'
            report.rejection_notes = notes
            report.reviewed_by = request.user
            report.reviewed_at = timezone.now()
            messages.success(request, 'Laporan ECA ditolak.')
        report.save()
        return redirect('eca:eca_report_list')

    return render(request, 'finance/review_report.html', {'report': report, 'type': 'ECA', 'back_url': 'eca:eca_report_list'})


@login_required
def create_eca_report(request):
    if request.user.role not in ['admin', 'eca_director', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    if request.method == 'POST':
        program_id = request.POST.get('program')
        period = request.POST.get('period')
        program = get_object_or_404(ECAProgram, id=program_id)
        registrations = ECARegistration.objects.filter(program=program, status='approved')
        total = registrations.count()
        paid = registrations.filter(is_paid=True).count()
        revenue = sum(r.program.price for r in registrations if r.is_paid)
        ECAReport.objects.create(
            academic_year=program.academic_year,
            program=program,
            period=period,
            total_students=total,
            total_paid=paid,
            total_revenue=revenue,
            status='submitted',
        )
        messages.success(request, 'Laporan ECA berhasil dibuat dan diajukan.')
        return redirect('eca:eca_report_list')
    programs = ECAProgram.objects.all()
    return render(request, 'eca/create_report.html', {'programs': programs})


@login_required
def eca_payment_review(request, pay_id):
    if request.user.role not in ['admin', 'eca_director', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import ECAPayment
    payment = get_object_or_404(ECAPayment, id=pay_id)
    if request.method == 'POST':
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')
        paid_amount = request.POST.get('paid_amount', payment.paid_amount)
        from django.utils import timezone
        payment.payment_notes = notes
        payment.paid_amount = paid_amount
        payment.reviewed_by = request.user
        payment.reviewed_at = timezone.now()
        price = payment.registration.program.price
        diff = int(paid_amount) - int(price)
        payment.under_over_amount = diff
        if action == 'approved':
            payment.payment_status = 'approved'
            payment.registration.is_paid = True
            payment.registration.save()
            if diff > 0:
                notes += f' [Kelebihan bayar Rp{diff:,}]'
            elif diff < 0:
                notes += f' [Kurang bayar Rp{abs(diff):,}]'
            payment.save()
            messages.success(request, f'Pembayaran diterima.{notes}')
        elif action == 'revised':
            payment.payment_status = 'revised'
            payment.save()
            if diff > 0:
                messages.warning(request, f'Kelebihan bayar Rp{diff:,}. {notes}')
            elif diff < 0:
                messages.warning(request, f'Kurang bayar Rp{abs(diff):,}. Silakan generate VA untuk kekurangan. {notes}')
            else:
                messages.warning(request, f'Perlu revisi: {notes}')
        elif action == 'rejected':
            payment.payment_status = 'rejected'
            payment.save()
            messages.error(request, f'Pembayaran ditolak: {notes}')
        return redirect('eca:registration_list')
    return render(request, 'eca/payment_review.html', {'payment': payment})


@login_required
def eca_payment_list(request):
    if request.user.role not in ['admin', 'tu', 'eca_director']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    filter_param = request.GET.get('filter', 'unpaid')
    qs = ECAPayment.objects.select_related('registration__student', 'registration__program').all()
    if filter_param == 'paid':
        qs = qs.filter(is_paid=True)
    elif filter_param == 'unpaid':
        qs = qs.filter(is_paid=False)
    return render(request, 'eca/payment_list.html', {'payments': qs, 'filter': filter_param})


@login_required
def download_eca_invoice(request, inv_id):
    from .models import ECAInvoice
    invoice = get_object_or_404(ECAInvoice, id=inv_id)
    if request.user.role == 'parent':
        if invoice.student not in request.user.children.all():
            messages.error(request, 'Akses ditolak.')
            return redirect('accounts:dashboard')
    if invoice.pdf:
        response = HttpResponse(invoice.pdf.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{invoice.invoice_number}.pdf"'
        return response
    # Generate PDF on the fly if no file stored
    from django.template.loader import render_to_string
    from xhtml2pdf import pisa
    html = render_to_string('eca/invoice_pdf.html', {'inv': invoice})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{invoice.invoice_number}.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('PDF error', status=500)
    return response
