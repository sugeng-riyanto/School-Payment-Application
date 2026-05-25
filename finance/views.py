from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum, Count
from .models import SPP, SPPBill, DPP, PaymentReminder, SPPReport, DPPReport, Compensation, Discount, CambridgeAssessment, CambridgeFee
from accounts.models import Student, AcademicYear, Grade
from payments.models import PaymentTransaction
from uuid import uuid4


def tu_required(view_func):
    def wrapper(request, *args, **kwargs):
        if request.user.role not in ['admin', 'tu']:
            messages.error(request, 'Akses ditolak.')
            return redirect('accounts:dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


def role_required(allowed_roles=None):
    def decorator(view_func):
        @login_required
        def wrapper(request, *args, **kwargs):
            if allowed_roles and request.user.role not in allowed_roles:
                messages.error(request, 'Akses ditolak.')
                return redirect('accounts:dashboard')
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


def _get_grade_level_filter(user):
    """Returns a filter dict for grade-based queries, or empty dict for admin/all-access."""
    if user.role == 'admin' or not user.assigned_level:
        return {}
    level_map = {'sd': ['sd'], 'smp': ['smp'], 'sma': ['sma'],
        'sd_smp': ['sd', 'smp'], 'smp_sma': ['smp', 'sma'], 'sd_smp_sma': ['sd', 'smp', 'sma']}
    levels = level_map.get(user.assigned_level, [])
    if not levels:
        return {}
    return {'grade__level__in': levels}


def _get_student_level_filter(user):
    """Returns a filter dict for student-based queries, or empty dict for admin/all-access."""
    if user.role == 'admin' or not user.assigned_level:
        return {}
    level_map = {'sd': ['sd'], 'smp': ['smp'], 'sma': ['sma'],
        'sd_smp': ['sd', 'smp'], 'smp_sma': ['smp', 'sma'], 'sd_smp_sma': ['sd', 'smp', 'sma']}
    levels = level_map.get(user.assigned_level, [])
    if not levels:
        return {}
    return {'student__class_grade__grade__level__in': levels}


@login_required
@tu_required
def spp_list(request):
    from .models import SPP, AcademicYear
    year_id = request.GET.get('year', '')
    lf = _get_grade_level_filter(request.user)
    spps = SPP.objects.select_related('academic_year', 'grade').filter(**lf).order_by('academic_year', 'grade__level', 'grade__name')
    if year_id:
        spps = spps.filter(academic_year_id=year_id)
    years = AcademicYear.objects.all()
    return render(request, 'finance/spp_list.html', {'spps': spps, 'years': years, 'selected_year': year_id})


@login_required
@tu_required
def spp_create(request):
    from .models import SPP, AcademicYear
    from accounts.models import Grade
    if request.method == 'POST':
        academic_year_id = request.POST.get('academic_year')
        grade_id = request.POST.get('grade')
        amount = request.POST.get('amount')
        SPP.objects.create(
            academic_year_id=academic_year_id,
            grade_id=grade_id,
            amount=amount,
        )
        messages.success(request, 'Tarif SPP berhasil ditambahkan.')
        return redirect('finance:spp_list')
    grades = Grade.objects.all()
    years = AcademicYear.objects.all()
    return render(request, 'finance/spp_create.html', {'grades': grades, 'years': years})


@login_required
@tu_required
def toggle_spp(request, pk):
    spp = get_object_or_404(SPP, id=pk)
    spp.is_active = not spp.is_active
    spp.save()
    status = 'diaktifkan' if spp.is_active else 'dinonaktifkan'
    messages.success(request, f'Tarif SPP untuk {spp.grade} berhasil {status}.')
    return redirect('finance:spp_list')


@login_required
@tu_required
def edit_spp(request, pk):
    spp = get_object_or_404(SPP, id=pk)
    if request.method == 'POST':
        spp.amount = request.POST.get('amount')
        spp.is_active = request.POST.get('is_active') == 'on'
        spp.save()
        messages.success(request, 'Tarif SPP berhasil diupdate.')
        return redirect('finance:spp_list')
    from accounts.models import Grade, AcademicYear
    grades = Grade.objects.all()
    years = AcademicYear.objects.all()
    return render(request, 'finance/edit_spp.html', {'spp': spp, 'grades': grades, 'years': years})


@login_required
@tu_required
def delete_spp(request, pk):
    SPP.objects.filter(id=pk).delete()
    messages.success(request, 'Tarif SPP berhasil dihapus.')
    return redirect('finance:spp_list')


@login_required
def spp_bill_list(request):
    if request.user.role not in ['admin', 'tu', 'kepsek', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import SPPBill

    lf = _get_student_level_filter(request.user)

    status = request.GET.get('status', 'all')
    sort = request.GET.get('sort', '-year')
    download = request.GET.get('download')

    qs = SPPBill.objects.select_related('student__class_grade__grade').filter(**lf)
    if status == 'paid':
        qs = qs.filter(is_paid=True)
    elif status == 'unpaid':
        qs = qs.filter(is_paid=False)

    if sort in ['name', '-name']:
        qs = qs.order_by('student__full_name' if sort == 'name' else '-student__full_name')
    else:
        qs = qs.order_by('-year', '-month')

    student = request.GET.get('student', '')
    if student:
        qs = qs.filter(student__full_name__icontains=student)

    if download == 'pdf':
        from django.http import HttpResponse
        from django.template.loader import render_to_string
        from xhtml2pdf import pisa
        html = render_to_string('finance/spp_bills_pdf.html', {'bills': qs, 'status': status})
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=spp_bills_{status}.pdf'
        pisa.CreatePDF(html, dest=response)
        return response

    if download == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SPP Bills"
        headers = ['NISN', 'Nama', 'Kelas', 'Bulan', 'Tahun', 'Nominal', 'Status']
        fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True)
            c.fill = fill
        for row, b in enumerate(qs, 2):
            ws.cell(row=row, column=1, value=b.student.nisn)
            ws.cell(row=row, column=2, value=b.student.full_name)
            ws.cell(row=row, column=3, value=str(b.student.class_grade))
            ws.cell(row=row, column=4, value=b.get_month_display())
            ws.cell(row=row, column=5, value=b.year)
            ws.cell(row=row, column=6, value=int(b.amount))
            ws.cell(row=row, column=7, value='Lunas' if b.is_paid else 'Belum')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=spp_bills_{status}.xlsx'
        wb.save(response)
        return response

    # Map bill -> transaction for receipt invoice links
    from payments.models import PaymentTransaction
    bill_ids = list(qs.values_list('id', flat=True))
    tx_qs = PaymentTransaction.objects.filter(
        transaction_type='spp',
        student_id__in=qs.values_list('student_id', flat=True),
    ).values('student_id', 'amount', 'id', 'invoice_number')
    bill_tx_map = {}
    for bill in qs:
        matching = [t for t in tx_qs if t['student_id'] == bill.student_id and t['amount'] == bill.amount]
        if matching:
            bill_tx_map[bill.id] = matching[0]

    ctx = {
        'bills': qs,
        'current_status': status,
        'current_sort': sort,
        'counts': {
            'all': SPPBill.objects.count(),
            'paid': SPPBill.objects.filter(is_paid=True).count(),
            'unpaid': SPPBill.objects.filter(is_paid=False).count(),
        },
        'bill_tx_map': bill_tx_map,
    }
    return render(request, 'finance/spp_bill_list.html', ctx)


@login_required
def confirm_all_spp(request):
    if request.user.role not in ['admin', 'tu']:
        from django.http import JsonResponse
        return JsonResponse({'success': False, 'message': 'Akses ditolak.'}, status=403)
    from .models import SPPBill
    from payments.models import PaymentTransaction
    from django.utils import timezone
    import uuid
    lf = _get_student_level_filter(request.user)
    bills = SPPBill.objects.filter(is_paid=False, **lf).select_related('student')
    count = 0
    for bill in bills:
        bill.is_paid = True
        bill.paid_at = timezone.now()
        bill.payment_method = 'manual'
        bill.save()
        order_id = f'BULK-SPP-{bill.student.nisn}-{bill.year}{bill.month:02d}-{uuid.uuid4().hex[:6].upper()}'
        PaymentTransaction.objects.get_or_create(
            student=bill.student, transaction_type='spp', amount=bill.amount,
            midtrans_order_id=order_id,
            defaults={
                'status': 'success', 'paid_at': timezone.now(),
                'payment_method': 'manual',
                'invoice_number': f'INV/SPP/{bill.student.class_grade.grade.level.upper()}/{bill.student.nisn}/{bill.year}{bill.month:02d}/{uuid.uuid4().hex[:4].upper()}',
            }
        )
        count += 1
    from django.http import JsonResponse
    return JsonResponse({'success': True, 'message': f'{count} tagihan berhasil dikonfirmasi.'})


@login_required
@tu_required
def edit_spp_bill(request, pk):
    bill = get_object_or_404(SPPBill, id=pk)
    if request.method == 'POST':
        bill.amount = request.POST.get('amount')
        bill.virtual_account = request.POST.get('virtual_account', '')
        bill.is_paid = request.POST.get('is_paid') == 'on'
        bill.save()
        messages.success(request, 'Tagihan SPP berhasil diupdate.')
        return redirect('finance:spp_bill_list')
    return render(request, 'finance/edit_spp_bill.html', {'bill': bill})


@login_required
@tu_required
@login_required
@tu_required
def delete_spp_bill(request, pk):
    SPPBill.objects.filter(id=pk).delete()
    messages.success(request, 'Tagihan SPP berhasil dihapus.')
    return redirect('finance:spp_bill_list')


@login_required
@tu_required
def export_spp_bills_xlsx(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SPP Bills"
    headers = ['NIS', 'Nama', 'Bulan', 'Tahun', 'Jumlah', 'VA', 'Status']
    hfill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="004085")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64+i)].width = 22
    for idx, bill in enumerate(SPPBill.objects.select_related('student').all().order_by('-year', '-month'), 2):
        ws.cell(row=idx, column=1, value=bill.student.nis)
        ws.cell(row=idx, column=2, value=bill.student.full_name)
        ws.cell(row=idx, column=3, value=bill.get_month_display())
        ws.cell(row=idx, column=4, value=bill.year)
        ws.cell(row=idx, column=5, value=int(bill.amount))
        ws.cell(row=idx, column=6, value=bill.virtual_account or '-')
        ws.cell(row=idx, column=7, value='Lunas' if bill.is_paid else 'Belum')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=spp_bills.xlsx'
    wb.save(response)
    return response


@login_required
def export_spp_bills_pdf(request):
    from payments.models import PaymentProof
    from datetime import datetime
    bills = SPPBill.objects.select_related('student').all().order_by('-year', '-month')
    rows = []
    for i, bill in enumerate(bills, 1):
        proof = PaymentProof.objects.filter(spp_bill=bill).first()
        rows.append({'values': [
            i, bill.student.full_name, bill.student.nis, bill.get_month_display(), bill.year,
            f'Rp{int(bill.amount):,}', bill.virtual_account or '-',
            'Lunas' if bill.is_paid else 'Belum',
            bill.paid_at.strftime('%d/%m/%Y') if bill.paid_at else '-',
            'Ada' if proof else '-',
            'Disetujui' if proof and proof.is_verified else ('Menunggu' if proof else '-'),
        ]})
    from reports.views import render_to_pdf
    return render_to_pdf('reports/pdf_generic.html', {
        'rows': rows, 'title': 'Tagihan SPP',
        'headers': ['No', 'Nama', 'NIS', 'Bulan', 'Tahun', 'Nominal', 'VA', 'Status', 'Tgl Bayar', 'Bukti', 'Verifikasi'],
    })


@login_required
@tu_required
def import_spp_bills(request):
    if request.method == 'POST' and request.FILES.get('file'):
        import openpyxl
        from accounts.models import Student

        wb = openpyxl.load_workbook(request.FILES['file'])
        ws = wb.active
        imported = 0
        errors = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            nis, month, year, amount = (row[0], row[1], row[2], row[3]) if len(row) >= 4 else (row[0], None, None, None)
            if not nis or not month or not year or not amount:
                errors.append(f'Baris tidak lengkap: {row}')
                continue
            student = Student.objects.filter(nis=str(nis).strip()).first()
            if not student:
                errors.append(f'Siswa NIS {nis} tidak ditemukan')
                continue
            try:
                spp = SPP.objects.filter(grade=student.class_grade.grade, is_active=True).first()
                SPPBill.objects.get_or_create(
                    student=student,
                    month=int(month),
                    year=int(year),
                    defaults={
                        'spp': spp,
                        'amount': int(amount),
                    }
                )
                imported += 1
            except Exception as e:
                errors.append(f'NIS {nis}: {e}')
        msg = f'Berhasil import {imported} tagihan.'
        if errors:
            msg += f' {len(errors)} error: ' + '; '.join(errors[:5])
        messages.success(request, msg)
        return redirect('finance:spp_bill_list')
    return render(request, 'finance/import_spp_bills.html')


@login_required
def confirm_spp_payment(request, bill_id):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    bill = get_object_or_404(SPPBill, id=bill_id)
    if request.method == 'POST':
        from django.utils import timezone
        action = request.POST.get('action', 'approve')
        if action == 'revisi':
            from payments.models import PaymentProof
            notes = request.POST.get('notes', '')
            PaymentProof.objects.filter(student=bill.student, spp_bill=bill).update(
                is_verified=False, description=notes
            )
            messages.warning(request, f'SPP {bill.get_month_display()} {bill.year} - {bill.student.full_name}: revisi diminta.')
            return redirect('finance:spp_bill_list')
        elif action == 'rejected':
            from payments.models import PaymentProof
            notes = request.POST.get('notes', '')
            PaymentProof.objects.filter(student=bill.student, spp_bill=bill).update(
                is_verified=False, description=f'Ditolak: {notes}'
            )
            messages.error(request, f'SPP {bill.get_month_display()} {bill.year} - {bill.student.full_name}: ditolak.')
            return redirect('finance:spp_bill_list')
        bill.is_paid = True
        bill.paid_at = timezone.now()
        bill.virtual_account = request.POST.get('virtual_account', bill.virtual_account)
        bill.payment_method = request.POST.get('payment_method', 'manual')
        bill.save()
        order_id = f'MANUAL-SPP-{bill.student.nisn}-{bill.year}{bill.month:02d}-{uuid4().hex[:6].upper()}'
        PaymentTransaction.objects.get_or_create(
            student=bill.student,
            transaction_type='spp',
            amount=bill.amount,
            midtrans_order_id=order_id,
            defaults={
                'status': 'success',
                'paid_at': timezone.now(),
                'payment_method': bill.payment_method,
                'invoice_number': f'INV/SPP/{bill.student.class_grade.grade.level.upper()}/{bill.student.nisn}/{bill.year}{bill.month:02d}/{uuid4().hex[:4].upper()}',
            }
        )
        messages.success(request, f'Pembayaran SPP {bill.student.full_name} - {bill.get_month_display()} {bill.year} dikonfirmasi.')
        return redirect('finance:spp_bill_list')
    return render(request, 'finance/confirm_spp.html', {'bill': bill})


@login_required
@tu_required
def create_spp_bills(request):
    # Original function body moved to Blast Email — kept for legacy URL redirect.
    messages.info(request, 'Pembuatan tagihan SPP sudah dipindahkan ke Blast Email.')
    return redirect('finance:blast_email')


@login_required
@tu_required
def generate_virtual_accounts(request):
    bills = SPPBill.objects.filter(is_paid=False, virtual_account='')
    for bill in bills:
        bill.virtual_account = f'VA-{bill.student.nis}-{bill.year}{bill.month:02d}'
        bill.save()
    messages.success(request, f'VA berhasil dibuat untuk {bills.count()} tagihan.')
    return redirect('finance:spp_list')


@login_required
@tu_required
def dpp_list(request):
    lf = _get_student_level_filter(request.user)
    dpps = DPP.objects.filter(category='pembangunan', **lf)
    filter_param = request.GET.get('filter', 'all')
    if filter_param == 'paid':
        dpps = dpps.filter(is_paid=True)
    elif filter_param == 'unpaid':
        dpps = dpps.filter(is_paid=False)
    return render(request, 'finance/dpp_list.html', {'dpps': dpps, 'filter': filter_param})


@login_required
@tu_required
def kegiatan_list(request):
    lf = _get_student_level_filter(request.user)
    dpps = DPP.objects.filter(category='kegiatan', **lf)
    filter_param = request.GET.get('filter', 'all')
    if filter_param == 'paid':
        dpps = dpps.filter(is_paid=True)
    elif filter_param == 'unpaid':
        dpps = dpps.filter(is_paid=False)
    return render(request, 'finance/kegiatan_list.html', {'dpps': dpps, 'filter': filter_param})


@login_required
@tu_required
def create_dpp(request):
    if request.method == 'POST':
        student_id = request.POST.get('student')
        category = request.POST.get('category')
        amount = request.POST.get('amount')
        description = request.POST.get('description', '')

        student = get_object_or_404(Student, pk=student_id)
        DPP.objects.create(
            student=student,
            category=category,
            amount=amount,
            description=description,
        )
        messages.success(request, 'DPP berhasil dibuat.')
        return redirect('finance:dpp_list')

    students = Student.objects.filter(is_active=True)
    return render(request, 'finance/create_dpp.html', {'students': students})


@login_required
@tu_required
def create_reminder(request):
    messages.info(request, 'Pembuatan peringatan sudah dipindahkan ke Blast Email.')
    return redirect('finance:blast_email')


@login_required
@tu_required
def import_spp(request):
    messages.info(request, 'Import Data SPP sudah dipindahkan ke halaman Blast Email.')
    return redirect('finance:blast_email')


@login_required
@tu_required
def edit_dpp(request, pk):
    dpp = get_object_or_404(DPP, id=pk)
    if request.method == 'POST':
        dpp.amount = request.POST.get('amount')
        dpp.description = request.POST.get('description', '')
        dpp.save()
        messages.success(request, 'DPP berhasil diupdate.')
        return redirect('finance:dpp_list')
    students = Student.objects.filter(is_active=True)
    return render(request, 'finance/edit_dpp.html', {'dpp': dpp, 'students': students})


@login_required
@tu_required
@login_required
@tu_required
def delete_dpp(request, pk):
    DPP.objects.filter(id=pk).delete()
    messages.success(request, 'DPP berhasil dihapus.')
    return redirect('finance:dpp_list')


@login_required
@tu_required
def edit_kegiatan(request, pk):
    dpp = get_object_or_404(DPP, id=pk)
    if request.method == 'POST':
        dpp.amount = request.POST.get('amount')
        dpp.description = request.POST.get('description', '')
        dpp.save()
        messages.success(request, 'Uang Kegiatan berhasil diupdate.')
        return redirect('finance:kegiatan_list')
    students = Student.objects.filter(is_active=True)
    return render(request, 'finance/edit_kegiatan.html', {'dpp': dpp, 'students': students})


@login_required
@tu_required
def delete_kegiatan(request, pk):
    DPP.objects.filter(id=pk).delete()
    messages.success(request, 'Uang Kegiatan berhasil dihapus.')
    return redirect('finance:kegiatan_list')


@login_required
@tu_required
@login_required
@tu_required
def approve_dpp(request, pk):
    dpp = get_object_or_404(DPP, id=pk)
    if request.method == 'POST':
        from django.utils import timezone
        action = request.POST.get('action')
        if action == 'approve' and request.FILES.get('payment_proof'):
            dpp.status = 'approved'
            dpp.is_paid = True
            dpp.paid_at = timezone.now()
            dpp.payment_method = request.POST.get('payment_method', 'manual')
            dpp.payment_proof = request.FILES['payment_proof']
            dpp.save()
            order_id = f'MANUAL-DPP-{dpp.student.nisn}-{uuid4().hex[:8].upper()}'
            PaymentTransaction.objects.create(
                student=dpp.student,
                transaction_type='kegiatan' if dpp.category == 'kegiatan' else 'dpp',
                amount=dpp.amount,
                status='success',
                midtrans_order_id=order_id,
                paid_at=dpp.paid_at,
                payment_method=dpp.payment_method,
                invoice_number=f'INV/DPP/{dpp.student.class_grade.grade.level.upper()}/{dpp.student.nisn}/{uuid4().hex[:6].upper()}',
            )
            messages.success(request, f'DPP {dpp.student.full_name} disetujui.')
        elif action == 'approved_no_proof':
            dpp.status = 'approved'
            dpp.is_paid = True
            dpp.paid_at = timezone.now()
            dpp.payment_method = request.POST.get('payment_method', 'manual')
            dpp.save()
            order_id = f'MANUAL-DPP-{dpp.student.nisn}-{uuid4().hex[:8].upper()}'
            PaymentTransaction.objects.create(
                student=dpp.student,
                transaction_type='kegiatan' if dpp.category == 'kegiatan' else 'dpp',
                amount=dpp.amount,
                status='success',
                midtrans_order_id=order_id,
                paid_at=dpp.paid_at,
                payment_method=dpp.payment_method,
                invoice_number=f'INV/DPP/{dpp.student.class_grade.grade.level.upper()}/{dpp.student.nisn}/{uuid4().hex[:6].upper()}',
            )
            messages.success(request, f'DPP {dpp.student.full_name} disetujui (tanpa bukti).')
        elif action == 'revisi':
            notes = request.POST.get('notes', '')
            dpp.status = 'revised'
            dpp.description = (dpp.description + '\n[Revisi: ' + notes + ']').strip()
            dpp.save()
            messages.warning(request, f'DPP {dpp.student.full_name} direvisi: {notes}')
        elif action == 'rejected':
            notes = request.POST.get('notes', '')
            dpp.status = 'rejected'
            dpp.description = (dpp.description + '\n[Ditolak: ' + notes + ']').strip()
            dpp.save()
            messages.error(request, f'DPP {dpp.student.full_name} ditolak: {notes}')
        return redirect('finance:dpp_list')
    return render(request, 'finance/approve_dpp.html', {'dpp': dpp})


@login_required
@tu_required
def approve_all_dpp(request):
    from django.utils import timezone
    count = DPP.objects.filter(is_paid=False).count()
    if request.method == 'POST' and request.FILES.get('payment_proof'):
        wb = openpyxl.load_workbook(request.FILES['payment_proof'])
        ws = wb.active
        rows_seen = set()
        approved = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            nis = str(row[0]).strip() if row[0] else ''
            if not nis or nis in rows_seen:
                continue
            rows_seen.add(nis)
            student = Student.objects.filter(nis=nis).first()
            if student:
                dpps = DPP.objects.filter(student=student, is_paid=False)
                for dpp in dpps:
                    dpp.status = 'approved'
                    dpp.is_paid = True
                    dpp.paid_at = timezone.now()
                    dpp.payment_method = 'import'
                    dpp.save()
                    approved += 1
        messages.success(request, f'{approved} DPP berhasil disetujui massal.')
        return redirect('finance:dpp_list')
    return render(request, 'finance/approve_all_dpp.html', {'count': count})


@login_required
@login_required
def reminder_list(request):
    if request.user.role not in ['admin', 'tu', 'kepsek']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import PaymentReminder
    qs = PaymentReminder.objects.select_related('student', 'academic_year').all().order_by('-sent_at')
    user = request.user
    if user.role == 'kepsek' and user.assigned_level:
        level_map = {'sd': ['sd'], 'smp': ['smp'], 'sma': ['sma'],
            'sd_smp': ['sd','smp'], 'smp_sma': ['smp','sma'], 'sd_smp_sma': ['sd','smp','sma']}
        allowed = level_map.get(user.assigned_level, [])
        qs = qs.filter(student__class_grade__grade__level__in=allowed)
    return render(request, 'finance/reminder_list.html', {'reminders': qs})

@login_required
def export_reminder_xlsx(request):
    if request.user.role not in ['admin', 'tu', 'kepsek']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import PaymentReminder
    import openpyxl
    from openpyxl.styles import Font
    from datetime import datetime
    qs = PaymentReminder.objects.select_related('student', 'academic_year').all().order_by('-sent_at')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Peringatan"
    headers = ['No', 'Siswa', 'NISN', 'Kelas', 'Jenis Peringatan', 'Bulan', 'Tahun Ajaran', 'Catatan', 'Tgl Kirim', 'Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for i, r in enumerate(qs, 1):
        ws.append([
            i, r.student.full_name, r.student.nisn, str(r.student.class_grade),
            r.get_reminder_type_display(),
            r.get_month_display() if r.month else '-',
            str(r.academic_year) if r.academic_year else '-',
            r.notes or '-',
            r.sent_at.strftime('%d/%m/%Y %H:%M') if r.sent_at else '-',
            'Sudah Dibaca' if r.is_read else 'Belum Dibaca',
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=laporan_peringatan_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response

@login_required
def export_reminder_pdf(request):
    if request.user.role not in ['admin', 'tu', 'kepsek']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import PaymentReminder
    qs = PaymentReminder.objects.select_related('student', 'academic_year').all().order_by('-sent_at')
    rows = []
    for i, r in enumerate(qs, 1):
        rows.append({'values': [
            i, r.student.full_name, r.student.nisn, str(r.student.class_grade),
            r.get_reminder_type_display(),
            r.get_month_display() if r.month else '-',
            r.notes or '-',
            r.sent_at.strftime('%d/%m/%Y') if r.sent_at else '-',
            'Sudah Dibaca' if r.is_read else 'Belum Dibaca',
        ]})
    from reports.views import render_to_pdf
    return render_to_pdf('reports/pdf_generic.html', {
        'rows': rows, 'title': 'Laporan Peringatan Pembayaran',
        'headers': ['No', 'Siswa', 'NISN', 'Kelas', 'Jenis', 'Bulan', 'Catatan', 'Tgl Kirim', 'Status'],
    })


@login_required
@tu_required
def import_dpp(request):
    messages.info(request, 'Import Data DPP sudah dipindahkan ke halaman Blast Email.')
    return redirect('finance:blast_email')


@login_required
@tu_required
def import_kegiatan(request):
    if request.method == 'POST' and request.FILES.get('file'):
        import openpyxl
        from .models import DPP
        from accounts.models import Student

        wb = openpyxl.load_workbook(request.FILES['file'])
        ws = wb.active
        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            nis, kegiatan_name, month, amount, desc = (row[0], row[1], row[2], row[3], row[4]) if len(row) >= 5 else (row[0], '', row[1], row[2], '')
            student = Student.objects.filter(nis=str(nis)).first()
            if student and kegiatan_name and amount:
                DPP.objects.create(
                    student=student,
                    category='kegiatan',
                    amount=int(amount),
                    description=f"{kegiatan_name} - {desc}".strip(' -'),
                )
                imported += 1
        messages.success(request, f'Berhasil import {imported} Uang Kegiatan.')
        return redirect('finance:kegiatan_list')
    return render(request, 'finance/import_kegiatan.html')


@login_required
def payment_recap(request):
    if request.user.role not in ['admin', 'tu', 'kepsek']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from accounts.models import AcademicYear, Grade, ClassGrade
    year = request.GET.get('year', '')
    grade_level = request.GET.get('grade_level', '')
    class_grade_id = request.GET.get('class_grade_id', '')
    month = request.GET.get('month', '')
    status = request.GET.get('status', 'all')

    qs = SPPBill.objects.select_related('student__class_grade__grade').all()
    # Level restriction for kepsek/vp_activity
    if request.user.role in ['kepsek', 'vp_activity'] and request.user.assigned_level:
        qs = qs.filter(student__class_grade__grade__level=request.user.assigned_level)
    if year:
        qs = qs.filter(year=year)
    if grade_level:
        qs = qs.filter(student__class_grade__grade__level=grade_level)
    if class_grade_id:
        qs = qs.filter(student__class_grade_id=class_grade_id)
    if month:
        qs = qs.filter(month=month)
    if status == 'paid':
        qs = qs.filter(is_paid=True)
    elif status == 'unpaid':
        qs = qs.filter(is_paid=False)
    qs = qs.order_by('-year', '-month')

    total_paid = qs.filter(is_paid=True).aggregate(Sum('amount'))['amount__sum'] or 0
    total_unpaid = qs.filter(is_paid=False).aggregate(Sum('amount'))['amount__sum'] or 0
    paid_count = qs.filter(is_paid=True).count()
    unpaid_count = qs.filter(is_paid=False).count()

    years = AcademicYear.objects.all()
    grade_levels = Grade.LEVEL_CHOICES
    class_grades = ClassGrade.objects.filter(grade__level=grade_level) if grade_level else ClassGrade.objects.all()
    months = SPP.MONTH_CHOICES

    context = {
        'spp_bills': qs,
        'total_paid': total_paid,
        'total_unpaid': total_unpaid,
        'paid_count': paid_count,
        'unpaid_count': unpaid_count,
        'years': years,
        'grade_levels': grade_levels,
        'class_grades': class_grades,
        'months': months,
        'selected_year': year,
        'selected_grade_level': grade_level,
        'selected_class_grade_id': class_grade_id,
        'selected_month': month,
        'selected_status': status,
    }
    return render(request, 'finance/payment_recap.html', context)


@login_required
def export_recap_xlsx(request):
    from accounts.models import AcademicYear, Grade, ClassGrade
    import openpyxl
    from openpyxl.styles import Font
    from datetime import datetime
    year = request.GET.get('year', '')
    grade_level = request.GET.get('grade_level', '')
    class_grade_id = request.GET.get('class_grade_id', '')
    month = request.GET.get('month', '')
    status = request.GET.get('status', 'all')
    qs = SPPBill.objects.select_related('student__class_grade__grade').all()
    if year: qs = qs.filter(year=year)
    if grade_level: qs = qs.filter(student__class_grade__grade__level=grade_level)
    if class_grade_id: qs = qs.filter(student__class_grade_id=class_grade_id)
    if month: qs = qs.filter(month=month)
    if status == 'paid': qs = qs.filter(is_paid=True)
    elif status == 'unpaid': qs = qs.filter(is_paid=False)
    qs = qs.order_by('-year', '-month')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Pembayaran"
    headers = ['No', 'Siswa', 'NISN', 'Kelas', 'Bulan', 'Tahun', 'Nominal', 'Status', 'Tgl Bayar', 'Metode']
    ws.append(headers)
    for cell in ws[1]: cell.font = Font(bold=True)
    for i, b in enumerate(qs, 1):
        ws.append([i, b.student.full_name, b.student.nisn, str(b.student.class_grade),
            b.get_month_display(), b.year, int(b.amount),
            'Lunas' if b.is_paid else 'Belum',
            b.paid_at.strftime('%d/%m/%Y') if b.paid_at else '-', b.payment_method or '-'])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=rekap_pembayaran_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
def export_recap_pdf(request):
    from datetime import datetime
    year = request.GET.get('year', '')
    grade_level = request.GET.get('grade_level', '')
    class_grade_id = request.GET.get('class_grade_id', '')
    month = request.GET.get('month', '')
    status = request.GET.get('status', 'all')
    qs = SPPBill.objects.select_related('student__class_grade__grade').all()
    if year: qs = qs.filter(year=year)
    if grade_level: qs = qs.filter(student__class_grade__grade__level=grade_level)
    if class_grade_id: qs = qs.filter(student__class_grade_id=class_grade_id)
    if month: qs = qs.filter(month=month)
    if status == 'paid': qs = qs.filter(is_paid=True)
    elif status == 'unpaid': qs = qs.filter(is_paid=False)
    qs = qs.order_by('-year', '-month')

    rows = []
    for i, b in enumerate(qs, 1):
        rows.append({'values': [i, b.student.full_name, b.student.nisn, str(b.student.class_grade),
            b.get_month_display(), b.year, f'Rp{int(b.amount):,}',
            'Lunas' if b.is_paid else 'Belum',
            b.paid_at.strftime('%d/%m/%Y') if b.paid_at else '-']})
    from reports.views import render_to_pdf
    return render_to_pdf('reports/pdf_generic.html', {
        'rows': rows, 'title': 'Rekap Pembayaran SPP',
        'headers': ['No', 'Siswa', 'NISN', 'Kelas', 'Bulan', 'Tahun', 'Nominal', 'Status', 'Tgl Bayar'],
    })


@login_required
def import_spp_va(request):
    messages.info(request, 'Import VA SPP sudah dipindahkan ke halaman Blast Email.')
    return redirect('finance:blast_email')


@login_required
def discount_list(request):
    if request.user.role not in ['admin', 'kepsek', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    discounts = Discount.objects.filter(**_get_student_level_filter(request.user)).order_by('-created_at')
    return render(request, 'finance/discount_list.html', {'discounts': discounts})


@login_required
def create_discount(request):
    if request.user.role not in ['admin', 'kepsek', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    if request.method == 'POST':
        student_id = request.POST.get('student')
        discount_type = request.POST.get('discount_type')
        amount = request.POST.get('amount')
        description = request.POST.get('description', '')
        valid_until = request.POST.get('valid_until')
        student = get_object_or_404(Student, pk=student_id)
        Discount.objects.create(
            student=student,
            discount_type=discount_type,
            amount=amount,
            description=description,
            valid_until=valid_until,
            created_by=request.user,
        )
        messages.success(request, 'Diskon berhasil dibuat.')
        return redirect('finance:discount_list')
    students = Student.objects.filter(is_active=True)
    return render(request, 'finance/create_discount.html', {'students': students})


@login_required
@tu_required
def edit_cambridge(request, pk):
    ca = get_object_or_404(CambridgeAssessment, id=pk)
    if request.method == 'POST':
        ca.exam_type = request.POST.get('exam_type')
        ca.subject = request.POST.get('subject')
        ca.amount = request.POST.get('amount')
        ca.notes = request.POST.get('notes', '')
        ca.save()
        messages.success(request, 'Cambridge Assessment berhasil diupdate.')
        return redirect('finance:cambridge_list')
    students = Student.objects.filter(is_active=True)
    return render(request, 'finance/edit_cambridge.html', {'ca': ca, 'students': students})


@login_required
@tu_required
def delete_cambridge(request, pk):
    CambridgeAssessment.objects.filter(id=pk).delete()
    messages.success(request, 'Cambridge Assessment berhasil dihapus.')
    return redirect('finance:cambridge_list')


@login_required
@tu_required
def confirm_cambridge(request, pk):
    ca = get_object_or_404(CambridgeAssessment, id=pk)
    if request.method == 'POST':
        from django.utils import timezone
        action = request.POST.get('action')
        if action == 'generate_va':
            order_id = f'VA-CAM-{ca.student.nisn}-{uuid4().hex[:8].upper()}'
            PaymentTransaction.objects.create(
                student=ca.student,
                transaction_type='cambridge',
                amount=ca.amount,
                status='pending',
                midtrans_order_id=order_id,
                virtual_account=f'VA-CAM-{ca.student.nis}-{uuid4().hex[:6].upper()}',
                invoice_number=f'INV/CAM/{ca.student.class_grade.grade.level.upper()}/{ca.student.nisn}/{uuid4().hex[:6].upper()}',
            )
            ca.is_paid = True
            ca.paid_at = timezone.now()
            ca.payment_method = 'va_midtrans'
            ca.save()
            messages.success(request, f'VA Midtrans berhasil dibuat untuk {ca.student.full_name} - {ca.get_exam_type_display()} {ca.subject}.')
            return redirect('finance:cambridge_list')
        else:
            ca.is_paid = True
            ca.paid_at = timezone.now()
            ca.payment_method = request.POST.get('payment_method', 'manual')
            ca.save()
            order_id = f'MANUAL-CAM-{ca.student.nisn}-{uuid4().hex[:8].upper()}'
            PaymentTransaction.objects.create(
                student=ca.student,
                transaction_type='cambridge',
                amount=ca.amount,
                status='success',
                midtrans_order_id=order_id,
                paid_at=timezone.now(),
                payment_method=ca.payment_method,
                invoice_number=f'INV/CAM/{ca.student.class_grade.grade.level.upper()}/{ca.student.nisn}/{uuid4().hex[:6].upper()}',
            )
            messages.success(request, f'Pembayaran Cambridge {ca.student.full_name} - {ca.get_exam_type_display()} {ca.subject} dikonfirmasi.')
            return redirect('finance:cambridge_list')
    return render(request, 'finance/confirm_cambridge.html', {'ca': ca})


@login_required
def cambridge_fee_list(request):
    if request.user.role not in ['admin', 'tu', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    fees = CambridgeFee.objects.all()
    return render(request, 'finance/cambridge_fee_list.html', {'fees': fees})


@login_required
def cambridge_fee_create(request, pk=None):
    if request.user.role not in ['admin', 'tu', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    fee = None
    if pk:
        fee = get_object_or_404(CambridgeFee, id=pk)
    if request.method == 'POST':
        exam_type = request.POST.get('exam_type')
        subject = request.POST.get('subject')
        if subject == 'Lainnya':
            subject = request.POST.get('custom_subject', '').strip()
        defaults = {'amount': request.POST.get('amount')}
        if fee:
            fee.exam_type = exam_type
            fee.subject = subject
            fee.amount = request.POST.get('amount')
            fee.save()
        else:
            CambridgeFee.objects.update_or_create(
                exam_type=exam_type, subject=subject,
                defaults=defaults
            )
        messages.success(request, 'Tarif Cambridge berhasil disimpan.')
        return redirect('finance:cambridge_fee_list')
    return render(request, 'finance/cambridge_fee_create.html', {'fee': fee})


@login_required
def cambridge_fee_delete(request, pk):
    if request.user.role not in ['admin', 'tu', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    CambridgeFee.objects.filter(id=pk).delete()
    messages.success(request, 'Tarif Cambridge berhasil dihapus.')
    return redirect('finance:cambridge_fee_list')


@login_required
def cambridge_fee_toggle(request, pk):
    if request.user.role not in ['admin', 'tu', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    fee = get_object_or_404(CambridgeFee, id=pk)
    fee.is_active = not fee.is_active
    fee.save()
    status = 'diaktifkan' if fee.is_active else 'dinonaktifkan'
    messages.success(request, f'Tarif Cambridge {fee.subject} berhasil {status}.')
    return redirect('finance:cambridge_fee_list')


@login_required
def export_cambridge_fee_template(request):
    if request.user.role not in ['admin', 'tu', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tarif Cambridge"
    headers = ['Exam Type', 'Subject', 'Amount']
    hfill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="004085")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64+i)].width = 25
    ws.cell(row=2, column=1, value='checkpoints / igcse / as_level / a_level')
    ws.cell(row=2, column=2, value='Mathematics / English / Physics / ...')
    ws.cell(row=2, column=3, value=500000)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=cambridge_fee_template.xlsx'
    wb.save(response)
    return response


@login_required
def import_cambridge_fee(request):
    if request.user.role not in ['admin', 'tu', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    if request.method == 'POST' and request.FILES.get('file'):
        import openpyxl
        import os
        uploaded = request.FILES['file']
        wb = openpyxl.load_workbook(uploaded)
        ws = wb.active
        imported = 0
        errors = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            exam_type = str(row[0]).strip() if row[0] else ''
            subject = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            amount = row[2] if len(row) > 2 and row[2] else None
            if not exam_type or not subject or not amount:
                errors.append(f'Baris {idx}: data tidak lengkap')
                continue
            valid_types = dict(CambridgeFee.EXAM_CHOICES).keys()
            if exam_type not in valid_types:
                errors.append(f'Baris {idx}: exam_type "{exam_type}" tidak valid')
                continue
            CambridgeFee.objects.update_or_create(
                exam_type=exam_type, subject=subject,
                defaults={'amount': int(amount)}
            )
            imported += 1
        # Delete uploaded file
        if hasattr(uploaded, 'temporary_file_path'):
            try:
                os.unlink(uploaded.temporary_file_path())
            except (OSError, AttributeError):
                pass
        if errors:
            messages.warning(request, f'{imported} tarif diimport. {len(errors)} error: ' + '; '.join(errors[:5]))
        else:
            messages.success(request, f'Berhasil import {imported} tarif Cambridge.')
        return redirect('finance:cambridge_fee_list')
    return render(request, 'finance/import_cambridge_fee.html', {'exam_choices': CambridgeFee.EXAM_CHOICES})


@login_required
@tu_required
def import_cambridge(request):
    if request.method == 'POST' and request.FILES.get('file'):
        import openpyxl
        wb = openpyxl.load_workbook(request.FILES['file'])
        ws = wb.active
        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            nis, exam_type, subject, amount = row[0], row[1], row[2], row[3]
            student = Student.objects.filter(nis=str(nis).strip()).first()
            if student and exam_type and subject and amount:
                CambridgeAssessment.objects.create(
                    student=student,
                    exam_type=str(exam_type).strip(),
                    subject=str(subject).strip(),
                    amount=int(amount),
                )
                imported += 1
        messages.success(request, f'Berhasil import {imported} Cambridge Assessment.')
        return redirect('finance:cambridge_list')
    return render(request, 'finance/import_cambridge.html')


@login_required
@tu_required
def export_cambridge_xlsx(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cambridge Assessment"
    headers = ['NIS', 'Nama', 'Exam Type', 'Subject', 'Jumlah', 'Status', 'Tgl Bayar', 'Metode']
    hfill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="004085")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64+i)].width = 22
    for idx, ca in enumerate(CambridgeAssessment.objects.select_related('student').all().order_by('-created_at'), 2):
        ws.cell(row=idx, column=1, value=ca.student.nis)
        ws.cell(row=idx, column=2, value=ca.student.full_name)
        ws.cell(row=idx, column=3, value=ca.get_exam_type_display())
        ws.cell(row=idx, column=4, value=ca.subject)
        ws.cell(row=idx, column=5, value=int(ca.amount))
        ws.cell(row=idx, column=6, value='Lunas' if ca.is_paid else 'Belum')
        ws.cell(row=idx, column=7, value=ca.paid_at.strftime('%d/%m/%Y') if ca.paid_at else '-')
        ws.cell(row=idx, column=8, value=ca.payment_method or '-')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=cambridge_assessment.xlsx'
    wb.save(response)
    return response


@login_required
def export_cambridge_pdf(request):
    from payments.models import PaymentProof
    from datetime import datetime
    cas = CambridgeAssessment.objects.select_related('student').all().order_by('-created_at')
    rows = []
    for i, ca in enumerate(cas, 1):
        proof = PaymentProof.objects.filter(transaction__student=ca.student, transaction__transaction_type='cambridge').first()
        rows.append({'values': [
            i, ca.student.full_name, ca.student.nis, ca.get_exam_type_display(), ca.subject,
            f'Rp{int(ca.amount):,}',
            'Lunas' if ca.is_paid else 'Belum',
            ca.paid_at.strftime('%d/%m/%Y') if ca.paid_at else '-',
            'Ada' if proof else '-',
            'Disetujui' if proof and proof.is_verified else ('Menunggu' if proof else '-'),
        ]})
    from reports.views import render_to_pdf
    return render_to_pdf('reports/pdf_generic.html', {
        'rows': rows, 'title': 'Cambridge Assessment',
        'headers': ['No', 'Nama', 'NIS', 'Exam Type', 'Subject', 'Nominal', 'Status', 'Tgl Bayar', 'Bukti', 'Verifikasi'],
    })


@login_required
def compensation_list(request):
    if request.user.role not in ['admin', 'tu', 'kepsek']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    compensations = Compensation.objects.filter(**_get_student_level_filter(request.user)).order_by('-created_at')
    return render(request, 'finance/compensation_list.html', {'compensations': compensations})


@login_required
def create_compensation(request):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    if request.method == 'POST':
        student_id = request.POST.get('student')
        comp_type = request.POST.get('comp_type')
        discount_method = request.POST.get('discount_method', 'nominal')
        description = request.POST.get('description')
        original_amount = request.POST.get('original_amount')
        comp_amount = request.POST.get('comp_amount')
        student = get_object_or_404(Student, pk=student_id)
        Compensation.objects.create(
            student=student,
            comp_type=comp_type,
            discount_method=discount_method,
            description=description,
            original_amount=original_amount,
            comp_amount=comp_amount,
            created_by=request.user,
        )
        messages.success(request, 'Kompensasi berhasil dibuat.')
        return redirect('finance:compensation_list')
    students = Student.objects.all()
    return render(request, 'finance/create_compensation.html', {'students': students})


@login_required
@tu_required
def import_compensation(request):
    if request.method == 'POST' and request.FILES.get('file'):
        import openpyxl
        wb = openpyxl.load_workbook(request.FILES['file'])
        ws = wb.active
        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            nis, comp_type, desc, orig, comp = row[0], row[1], row[2], row[3], row[4]
            student = Student.objects.filter(nis=str(nis)).first()
            if student and comp_type and orig and comp:
                original_amount = int(orig)
                comp_amount = int(comp)
                final_amount = original_amount - comp_amount
                Compensation.objects.create(
                    student=student,
                    comp_type=str(comp_type),
                    description=str(desc or ''),
                    original_amount=original_amount,
                    comp_amount=comp_amount,
                    created_by=request.user,
                )
                imported += 1
        messages.success(request, f'Berhasil import {imported} kompensasi.')
        return redirect('finance:compensation_list')
    return render(request, 'finance/import_compensation.html')


@login_required
@tu_required
def cambridge_list(request):
    from .models import CambridgeAssessment
    status = request.GET.get('status', 'all')
    sort = request.GET.get('sort', '-created_at')
    download = request.GET.get('download')

    qs = CambridgeAssessment.objects.select_related('student__class_grade__grade').filter(**_get_student_level_filter(request.user))
    if status == 'paid':
        qs = qs.filter(is_paid=True)
    elif status == 'unpaid':
        qs = qs.filter(is_paid=False)

    if sort in ['name', '-name']:
        qs = qs.order_by('student__full_name' if sort == 'name' else '-student__full_name')
    else:
        qs = qs.order_by('-created_at')

    student = request.GET.get('student', '')
    if student:
        qs = qs.filter(student__full_name__icontains=student)

    if download == 'pdf':
        from django.http import HttpResponse
        from django.template.loader import render_to_string
        from xhtml2pdf import pisa
        html = render_to_string('finance/cambridge_pdf.html', {'exams': qs, 'status': status})
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=cambridge_{status}.pdf'
        pisa.CreatePDF(html, dest=response)
        return response

    if download == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cambridge"
        headers = ['NISN', 'Nama', 'Kelas', 'Exam', 'Subject', 'Nominal', 'Status']
        fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True)
            c.fill = fill
        for row, ca in enumerate(qs, 2):
            ws.cell(row=row, column=1, value=ca.student.nisn)
            ws.cell(row=row, column=2, value=ca.student.full_name)
            ws.cell(row=row, column=3, value=str(ca.student.class_grade))
            ws.cell(row=row, column=4, value=ca.get_exam_type_display())
            ws.cell(row=row, column=5, value=ca.subject)
            ws.cell(row=row, column=6, value=int(ca.amount))
            ws.cell(row=row, column=7, value='Lunas' if ca.is_paid else 'Belum')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=cambridge_{status}.xlsx'
        wb.save(response)
        return response

    ctx = {
        'exams': qs,
        'current_status': status,
        'current_sort': sort,
        'counts': {
            'all': CambridgeAssessment.objects.count(),
            'paid': CambridgeAssessment.objects.filter(is_paid=True).count(),
            'unpaid': CambridgeAssessment.objects.filter(is_paid=False).count(),
        },
    }
    return render(request, 'finance/cambridge_list.html', ctx)


@login_required
@tu_required
def create_cambridge(request):
    if request.method == 'POST':
        student_id = request.POST.get('student')
        exam_type = request.POST.get('exam_type')
        subject = request.POST.get('subject')
        amount = request.POST.get('amount')
        notes = request.POST.get('notes', '')
        student = get_object_or_404(Student, pk=student_id)
        CambridgeAssessment.objects.create(
            student=student,
            exam_type=exam_type,
            subject=subject,
            amount=amount,
            notes=notes,
        )
        messages.success(request, 'Cambridge Assessment berhasil dibuat.')
        return redirect('finance:cambridge_list')
    students = Student.objects.filter(is_active=True)
    return render(request, 'finance/create_cambridge.html', {'students': students})


@login_required
def spp_report_list(request):
    if request.user.role not in ['admin', 'tu', 'kepsek']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    reports = SPPReport.objects.all()
    if request.user.role == 'kepsek' and request.user.assigned_level:
        from eca.views import _check_level_access
        reports = [r for r in reports if _check_level_access(request.user, r.grade)]
    reports = sorted(reports, key=lambda r: (-r.year, -r.month))
    return render(request, 'finance/spp_report_list.html', {'reports': reports})


@login_required
def dpp_report_list(request):
    if request.user.role not in ['admin', 'tu', 'kepsek']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    reports = DPPReport.objects.all()
    if request.user.role == 'kepsek' and request.user.assigned_level:
        from eca.views import _check_level_access
        reports = [r for r in reports if _check_level_access(request.user, r.grade)]
    return render(request, 'finance/dpp_report_list.html', {'reports': reports})


@login_required
def review_spp_report(request, pk):
    if request.user.role not in ['admin', 'kepsek', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    report = get_object_or_404(SPPReport, id=pk)
    from eca.views import _check_level_access
    if request.user.role == 'kepsek' and not _check_level_access(request.user, report.grade):
        messages.error(request, 'Akses terbatas sesuai jenjang.')
        return redirect('finance:spp_report_list')
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'approve':
            report.status = 'approved'
            report.reviewed_by = request.user
            report.reviewed_at = timezone.now()
            messages.success(request, 'Laporan SPP disetujui.')
        elif action == 'revise':
            report.status = 'revised'
            report.rejection_notes = request.POST.get('notes', '')
            messages.info(request, 'Laporan SPP diminta revisi.')
        elif action == 'reject':
            report.status = 'rejected'
            report.rejection_notes = request.POST.get('notes', '')
            messages.error(request, 'Laporan SPP ditolak.')
        report.save()
        return redirect('finance:spp_report_list')
    return render(request, 'finance/review_spp_report.html', {'report': report})


@login_required
def review_dpp_report(request, pk):
    if request.user.role not in ['admin', 'kepsek', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    report = get_object_or_404(DPPReport, id=pk)
    from eca.views import _check_level_access
    if request.user.role == 'kepsek' and not _check_level_access(request.user, report.grade):
        messages.error(request, 'Akses terbatas sesuai jenjang.')
        return redirect('finance:dpp_report_list')
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'approve':
            report.status = 'approved'
            report.reviewed_by = request.user
            report.reviewed_at = timezone.now()
            messages.success(request, 'Laporan DPP disetujui.')
        elif action == 'revise':
            report.status = 'revised'
            report.rejection_notes = request.POST.get('notes', '')
            messages.info(request, 'Laporan DPP diminta revisi.')
        elif action == 'reject':
            report.status = 'rejected'
            report.rejection_notes = request.POST.get('notes', '')
            messages.error(request, 'Laporan DPP ditolak.')
        report.save()
        return redirect('finance:dpp_report_list')
    return render(request, 'finance/review_dpp_report.html', {'report': report})


@login_required
def payment_history(request):
    from django.db.models import Sum, Count, Q

    years = AcademicYear.objects.all().order_by('-start_date')
    base_qs = PaymentTransaction.objects.all().select_related('student__class_grade__grade')

    # Role-based level filtering
    user = request.user
    if user.role == 'parent':
        base_qs = base_qs.filter(student__in=user.children.all())
    elif user.role == 'pic_teacher':
        base_qs = base_qs.filter(student__class_grade__grade__in=user.assigned_level.split(',') if user.assigned_level else [])
    elif user.role in ['kepsek', 'vp_activity']:
        if user.assigned_level:
            base_qs = base_qs.filter(student__class_grade__grade__level=user.assigned_level)

    # ── Filters ──
    q = request.GET.get('q', '').strip()
    transaction_type = request.GET.get('type', '')
    status = request.GET.get('status', '')
    year_id = request.GET.get('year', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    month = request.GET.get('month', '')

    if q:
        base_qs = base_qs.filter(
            Q(student__full_name__icontains=q) |
            Q(student__nisn__icontains=q) |
            Q(invoice_number__icontains=q) |
            Q(verification_code__icontains=q)
        )
    if transaction_type:
        base_qs = base_qs.filter(transaction_type=transaction_type)
    if status:
        base_qs = base_qs.filter(status=status)
    if year_id:
        try:
            ay = AcademicYear.objects.get(id=year_id)
            base_qs = base_qs.filter(created_at__date__gte=ay.start_date, created_at__date__lte=ay.end_date)
        except AcademicYear.DoesNotExist:
            pass
    if date_from:
        base_qs = base_qs.filter(created_at__date__gte=date_from)
    if date_to:
        base_qs = base_qs.filter(created_at__date__lte=date_to)
    if month:
        base_qs = base_qs.filter(created_at__month=month)

    # ── Sorting ──
    sort = request.GET.get('sort', 'created_at')
    dir = request.GET.get('dir', 'desc')
    allowed_sorts = ['created_at', 'amount', 'transaction_type', 'status', 'payment_method', 'student__full_name', 'student__nisn', 'student__class_grade__name']
    if sort not in allowed_sorts:
        sort = 'created_at'
    order = f'-{sort}' if dir == 'desc' else sort
    transactions = base_qs.order_by(order)[:500]

    # ── Aggregates ──
    by_type = base_qs.values('transaction_type').annotate(total=Sum('amount'), count=Count('id'))
    total_all = base_qs.aggregate(total=Sum('amount'))['total'] or 0
    filter_count = base_qs.count()

    context = {
        'transactions': transactions,
        'years': years,
        'by_type': by_type,
        'total_all': total_all,
        'filter_count': filter_count,
        'selected_year': year_id,
        'filter_q': q,
        'filter_type': transaction_type,
        'filter_status': status,
        'filter_date_from': date_from,
        'filter_date_to': date_to,
        'filter_month': month,
        'sort': sort,
        'dir': dir,
        'sort_options': {
            'Tgl': 'created_at',
            'Nama': 'student__full_name',
            'NISN': 'student__nisn',
            'Tipe': 'transaction_type',
            'Jumlah': 'amount',
            'Status': 'status',
        },
        'month_choices': [(1,'Jan'),(2,'Feb'),(3,'Mar'),(4,'Apr'),(5,'Mei'),(6,'Jun'),(7,'Jul'),(8,'Agu'),(9,'Sep'),(10,'Okt'),(11,'Nov'),(12,'Des')],
    }
    return render(request, 'finance/payment_history.html', context)


@login_required
def comparison_charts(request):
    if request.user.role not in ['admin', 'tu', 'kepsek', 'vp_activity', 'eca_director']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    
    # Level filter for kepsek/vp_activity
    level_filter = {}
    if request.user.role in ['kepsek', 'vp_activity'] and request.user.assigned_level:
        level_filter = {'student__class_grade__grade__level': request.user.assigned_level}
    
    years = AcademicYear.objects.all().order_by('start_date')
    spp_data = []
    for y in years:
        total = SPPBill.objects.filter(
            year__gte=y.start_date.year,
            year__lte=y.end_date.year,
            is_paid=True,
            **level_filter
        ).aggregate(total=Sum('amount'))['total'] or 0
        spp_data.append(float(total))
    dpp_data = []
    for y in years:
        total = DPP.objects.filter(
            paid_at__date__gte=y.start_date,
            paid_at__date__lte=y.end_date,
            is_paid=True,
            student__class_grade__grade__level=request.user.assigned_level
        ).aggregate(total=Sum('amount'))['total'] or 0 if request.user.role in ['kepsek', 'vp_activity'] and request.user.assigned_level else DPP.objects.filter(
            paid_at__date__gte=y.start_date,
            paid_at__date__lte=y.end_date,
            is_paid=True
        ).aggregate(total=Sum('amount'))['total'] or 0
        dpp_data.append(float(total))
    cam_data = []
    for y in years:
        total = CambridgeAssessment.objects.filter(
            paid_at__date__gte=y.start_date,
            paid_at__date__lte=y.end_date,
            is_paid=True
        ).aggregate(total=Sum('amount'))['total'] or 0
        cam_data.append(float(total))
    ay = AcademicYear.objects.filter(is_active=True).first()
    monthly_spp = []
    monthly_cam = []
    if ay:
        for m in range(7, 13):
            spp = SPPBill.objects.filter(month=m, year=ay.start_date.year, is_paid=True, **level_filter).aggregate(t=Sum('amount'))['t'] or 0
            cam = CambridgeAssessment.objects.filter(
                paid_at__month=m, paid_at__year=ay.start_date.year, is_paid=True
            ).aggregate(t=Sum('amount'))['t'] or 0
            monthly_spp.append(float(spp))
            monthly_cam.append(float(cam))
    year_labels = [y.name for y in years]
    return render(request, 'finance/comparison_charts.html', {
        'years': years,
        'year_labels': year_labels,
        'spp_data': spp_data,
        'dpp_data': dpp_data,
        'cam_data': cam_data,
        'monthly_spp': monthly_spp,
        'monthly_cam': monthly_cam,
        'month_labels': ['Jul', 'Ags', 'Sep', 'Okt', 'Nov', 'Des'],
    })


from django.http import JsonResponse
from datetime import datetime


BULAN_NAMES = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']


@login_required
def blast_email(request):
    if request.user.role not in ['admin', 'tu', 'kepsek', 'vp_activity', 'pic_teacher']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from accounts.models import ClassGrade, Student, AcademicYear, Grade
    from payments.models import BlastEmailTemplate
    from .models import SPP, DPP, SPPBill
    import openpyxl

    # Handle POST for imports and blast
    if request.method == 'POST':
        import_action = request.POST.get('import_action', '')

        if import_action == 'import_spp_va':
            if request.FILES.get('file'):
                wb = openpyxl.load_workbook(request.FILES['file'])
                ws = wb.active
                imported = 0
                errors = []
                headers = []
                for cell in ws[1]:
                    headers.append(str(cell.value).strip().lower() if cell.value else '')
                if not headers:
                    messages.error(request, 'File tidak memiliki header.')
                    return redirect('finance:blast_email')
                first_header = headers[0] if headers else ''
                if first_header.startswith('nis'):
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not any(row): continue
                        nis = str(row[0]).strip() if row[0] else ''
                        va = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                        if nis and va:
                            bills = SPPBill.objects.filter(student__nis=nis, is_paid=False, virtual_account='')
                            for bill in bills:
                                bill.virtual_account = va
                                bill.save()
                                imported += 1
                elif first_header.startswith('subject'):
                    from accounts.models import Student as Std
                    from django.utils import timezone
                    col_map = {h: idx for idx, h in enumerate(headers)}
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not any(row): continue
                        try:
                            customer_name = str(row[col_map.get('customer_name', 3)]).strip() if len(row) > col_map.get('customer_name', 3) and row[col_map.get('customer_name', 3)] else ''
                            grade_name = str(row[col_map.get('grade', 1)]).strip() if len(row) > col_map.get('grade', 1) and row[col_map.get('grade', 1)] else ''
                            va = str(row[col_map.get('virtual_account', 2)]).strip() if len(row) > col_map.get('virtual_account', 2) and row[col_map.get('virtual_account', 2)] else ''
                            trx_amount = row[col_map.get('trx_amount', 5)] if len(row) > col_map.get('trx_amount', 5) else None
                            if not customer_name and not grade_name:
                                errors.append('Baris tanpa nama/kelas'); continue
                            student = None
                            if customer_name:
                                student = Std.objects.filter(full_name__icontains=customer_name).first()
                            if not student and grade_name:
                                cg = ClassGrade.objects.filter(name__icontains=grade_name).first()
                                if cg: student = Std.objects.filter(class_grade=cg, is_active=True).first()
                            if not student:
                                errors.append(f'Siswa "{customer_name}" tak ditemukan'); continue
                            if trx_amount:
                                try: trx_amount = int(float(str(trx_amount).replace(',', '').replace('.', '')))
                                except: trx_amount = 0
                            bills = SPPBill.objects.filter(student=student, is_paid=False)
                            if bills.exists():
                                for bill in bills:
                                    bill.virtual_account = va; bill.save(); imported += 1
                            else:
                                errors.append(f'Tidak ada tagihan SPP untuk {student.full_name}')
                            from payments.models import PaymentTransaction
                            PaymentTransaction.objects.update_or_create(
                                student=student, transaction_type='spp', status='pending',
                                defaults={
                                    'amount': trx_amount or 0, 'virtual_account': va,
                                    'midtrans_order_id': f'VA-SPP-{student.nisn}-{uuid4().hex[:8].upper()}',
                                }
                            )
                        except Exception as e:
                            errors.append(f'Error: {e}')
                else:
                    messages.error(request, 'Format file tidak dikenali.')
                    return redirect('finance:blast_email')
                msg = f'VA berhasil diupload untuk {imported} tagihan.'
                if errors: msg += f' {len(errors)} error: ' + '; '.join(errors[:5])
                messages.success(request, msg)
                return redirect('finance:blast_email')
            messages.error(request, 'Pilih file XLSX terlebih dahulu.')
            return redirect('finance:blast_email')

        elif import_action == 'import_spp':
            if request.FILES.get('file'):
                wb = openpyxl.load_workbook(request.FILES['file'])
                ws = wb.active
                academic_year = AcademicYear.objects.filter(is_active=True).first()
                if not academic_year:
                    messages.error(request, 'Tidak ada tahun ajaran aktif.')
                    return redirect('finance:blast_email')
                imported = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row): continue
                    grade_name, amount = row[0], row[1]
                    grade = Grade.objects.filter(name__iexact=str(grade_name)).first()
                    if grade and amount:
                        SPP.objects.update_or_create(
                            academic_year=academic_year, grade=grade,
                            defaults={'amount': int(amount), 'is_active': True}
                        )
                        imported += 1
                messages.success(request, f'Berhasil import {imported} tarif SPP.')
                return redirect('finance:blast_email')
            messages.error(request, 'Pilih file XLSX terlebih dahulu.')
            return redirect('finance:blast_email')

        elif import_action == 'import_dpp':
            if request.FILES.get('file'):
                wb = openpyxl.load_workbook(request.FILES['file'])
                ws = wb.active
                imported = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row): continue
                    nis, category, amount, desc = row[0], row[1], row[2], row[3] if len(row) > 3 else ''
                    student = Student.objects.filter(nis=str(nis)).first()
                    if student and category and amount:
                        DPP.objects.create(
                            student=student, category=str(category),
                            amount=int(amount), description=str(desc or ''),
                        )
                        imported += 1
                messages.success(request, f'Berhasil import {imported} DPP.')
                return redirect('finance:blast_email')
            messages.error(request, 'Pilih file XLSX terlebih dahulu.')
            return redirect('finance:blast_email')

    from .models import BankAccount as BankAccountModel
    bank_options = []
    for b in BankAccountModel.objects.filter(status='approved', is_active=True).order_by('level', 'payment_type'):
        lvl = b.get_level_display()
        pt = b.get_payment_type_display()
        bank_options.append({'name': b.bank_name, 'account': b.account_number, 'owner': b.account_holder, 'label': f'{lvl} - {pt}: {b.bank_name} {b.account_number}'})
    if not bank_options:
        bank_options = [{'name': 'No bank', 'account': '', 'owner': '', 'label': 'Tidak ada rekening aktif. Atur di Rekening Bank.'}]

    default_body = """Kepada Yth. Orang Tua/Wali Murid {1} (Kelas {2})

Salam Hormat,

Kami sampaikan informasi tagihan SPP Bulan {3}.
Rincian Pembayaran:
• Nominal: {4}
• Batas Pembayaran: {5}
• Virtual Account BNI: {6}
• Rekening OCBC: 131010888990

Mohon pembayaran dilakukan sebelum tanggal jatuh tempo untuk menghindari denda administrasi sebesar 5%. Jika pembayaran sudah dilakukan, mohon abaikan pesan ini.

Untuk bukti transfer atau pertanyaan, silakan hubungi {7} di {8} atau klik tautan berikut untuk panduan: {9}.

Terima kasih.
School Admin"""

    from django.utils import timezone
    now = timezone.localtime(timezone.now())
    
    context = {
        'payment_types': ['spp', 'kegiatan', 'eca', 'cambridge', 'seragam', 'denda'],
        'class_grades': ClassGrade.objects.all().order_by('grade__level', 'name'),
        'students': Student.objects.filter(is_active=True).order_by('full_name'),
        'templates': BlastEmailTemplate.objects.all(),
        'bank_options': bank_options,
        'default_body': default_body,
        'recent_spp_bills': SPPBill.objects.select_related('student').all().order_by('-created_at')[:20],
        'recent_dpp': DPP.objects.select_related('student').all().order_by('-created_at')[:20],
        'recent_kegiatan': DPP.objects.filter(category='kegiatan').select_related('student').all().order_by('-created_at')[:20],
        'recent_spp_rates': SPP.objects.select_related('academic_year', 'grade').all().order_by('-id')[:20],
        'month_choices': [(1,'Jan'),(2,'Feb'),(3,'Mar'),(4,'Apr'),(5,'Mei'),(6,'Jun'),(7,'Jul'),(8,'Agu'),(9,'Sep'),(10,'Okt'),(11,'Nov'),(12,'Des')],
        'year_choices': [now.year - 1, now.year, now.year + 1],
        'current_month': now.month,
        'current_year': now.year,
        'spp_rate': 'Rp 500,000',
        'day_choices': range(1, 32),
    }
    return render(request, 'finance/blast_email.html', context)


@login_required
@tu_required
def blast_email_send(request):
    if request.user.role not in ['admin', 'tu', 'kepsek', 'vp_activity', 'pic_teacher']:
        return JsonResponse({'success': False, 'error': 'Akses ditolak.'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    mode = request.POST.get('mode', 'preview')

    from accounts.models import Student
    from eca.models import ECARegistration

    payment_type = request.POST.get('payment_type', '')
    target = request.POST.get('target', '')
    body_template = request.POST.get('body', '')
    student_id = request.POST.get('student_id')
    class_grade_id = request.POST.get('class_grade_id')
    tu_name = request.POST.get('tu_name', 'Tata Usaha')
    tu_wa = request.POST.get('tu_wa', '')
    due_day = request.POST.get('due_day', '')
    due_month = request.POST.get('due_month', '')
    due_year = request.POST.get('due_year', '')
    google_drive_link = request.POST.get('google_drive_link', '')

    due_date = f'{due_day}/{due_month}/{due_year}' if due_day and due_month and due_year else ''

    academic_year = AcademicYear.objects.filter(is_active=True).first()
    recipients = []
    lf = _get_student_level_filter(request.user)

    auto_subject_map = {
        'spp': 'Informasi Tagihan SPP Bulan {3}',
        'kegiatan': 'Informasi Uang Kegiatan',
        'eca': 'Informasi Pembayaran ECA',
        'cambridge': 'Informasi Pembayaran Cambridge Assessment',
        'seragam': 'Informasi Pembayaran Seragam',
        'denda': 'Informasi Pembayaran Denda',
    }

    if payment_type == 'spp':
        bills = SPPBill.objects.filter(is_paid=False, **lf)
        if academic_year:
            bills = bills.filter(year=academic_year.start_date.year)
        if target == 'individual' and student_id:
            bills = bills.filter(student_id=student_id)
        elif target == 'class' and class_grade_id:
            bills = bills.filter(student__class_grade_id=class_grade_id)
        for bill in bills.select_related('student__class_grade__grade', 'student__parent'):
            student = bill.student
            recipients.append({
                'student': student,
                'amount': bill.amount,
                'va': bill.virtual_account or '',
                'month': bill.month,
                'year': bill.year,
                'email': student.parent.email if student.parent else student.user.email if student.user else '',
            })

    elif payment_type == 'kegiatan':
        dpks = DPP.objects.filter(category='kegiatan', status='pending', **lf)
        if target == 'individual' and student_id:
            dpks = dpks.filter(student_id=student_id)
        elif target == 'class' and class_grade_id:
            dpks = dpks.filter(student__class_grade_id=class_grade_id)
        for dpp in dpks.select_related('student__class_grade__grade', 'student__parent'):
            student = dpp.student
            recipients.append({
                'student': student,
                'amount': dpp.amount,
                'va': '',
                'month': 0,
                'year': 0,
                'email': student.parent.email if student.parent else student.user.email if student.user else '',
            })

    elif payment_type == 'eca':
        regs = ECARegistration.objects.filter(status='approved', **lf)
        if target == 'individual' and student_id:
            regs = regs.filter(student_id=student_id)
        elif target == 'class' and class_grade_id:
            regs = regs.filter(student__class_grade_id=class_grade_id)
        for reg in regs.select_related('student__class_grade__grade', 'student__parent', 'program'):
            student = reg.student
            recipients.append({
                'student': student,
                'amount': reg.program.price,
                'va': '',
                'month': 0,
                'year': 0,
                'email': student.parent.email if student.parent else student.user.email if student.user else '',
            })

    elif payment_type == 'cambridge':
        cas = CambridgeAssessment.objects.filter(is_paid=False, **lf)
        if target == 'individual' and student_id:
            cas = cas.filter(student_id=student_id)
        elif target == 'class' and class_grade_id:
            cas = cas.filter(student__class_grade_id=class_grade_id)
        for ca in cas.select_related('student__class_grade__grade', 'student__parent'):
            student = ca.student
            recipients.append({
                'student': student,
                'amount': ca.amount,
                'va': '',
                'month': 0,
                'year': 0,
                'email': student.parent.email if student.parent else student.user.email if student.user else '',
            })

    elif payment_type in ('seragam', 'denda'):
        students_qs = Student.objects.filter(is_active=True)
        if target == 'individual' and student_id:
            students_qs = students_qs.filter(id=student_id)
        elif target == 'class' and class_grade_id:
            students_qs = students_qs.filter(class_grade_id=class_grade_id)
        for student in students_qs.select_related('class_grade__grade', 'parent'):
            recipients.append({
                'student': student,
                'amount': 0,
                'va': '',
                'month': 0,
                'year': 0,
                'email': student.parent.email if student.parent else student.user.email if student.user else '',
            })

    formatted_emails = []
    auto_subject = auto_subject_map.get(payment_type, '')

    for r in recipients:
        bulan_name = BULAN_NAMES[r['month']] if 1 <= r['month'] <= 12 else ''
        tahun = str(r['year']) if r['year'] else ''
        nominal = f"Rp{int(r['amount']):,}" if r['amount'] else ''

        body = body_template.replace('{1}', r['student'].full_name)
        body = body.replace('{2}', str(r['student'].class_grade))
        body = body.replace('{3}', (bulan_name + ' ' + tahun).strip() if bulan_name else '')
        body = body.replace('{4}', nominal)
        body = body.replace('{5}', due_date)
        body = body.replace('{6}', r['va'])
        body = body.replace('{7}', tu_name)
        body = body.replace('{8}', tu_wa)
        body = body.replace('{9}', google_drive_link)

        subject = auto_subject.replace('{3}', (bulan_name + ' ' + tahun).strip() if bulan_name else '')

        formatted_emails.append({
            'subject': subject,
            'body': body,
            'student_name': r['student'].full_name,
            'class': str(r['student'].class_grade),
            'amount': nominal,
            'va': r['va'],
            'email': r['email'],
        })

    if mode == 'send':
        from django.core.mail import EmailMessage
        from payments.models import BlastEmailLog
        sent = 0
        failed = 0
        for i, item in enumerate(formatted_emails):
            if not item['email']:
                failed += 1
                continue
            try:
                msg = EmailMessage(
                    subject=item['subject'],
                    body=item['body'],
                    to=[item['email']],
                )
                msg.send()
                sent += 1
            except Exception:
                failed += 1
            try:
                BlastEmailLog.objects.create(
                    student_id=recipients[i]['student'].id,
                    payment_type=payment_type,
                    subject=item['subject'],
                    body=item['body'],
                    recipient_email=item['email'],
                )
            except Exception:
                pass
        return JsonResponse({
            'success': True,
            'sent': sent,
            'failed': failed,
            'count': len(formatted_emails),
        })

    preview = formatted_emails[:3]
    return JsonResponse({
        'success': True,
        'count': len(formatted_emails),
        'preview': preview,
    })


@login_required
@tu_required
def blast_email_template_list(request):
    from payments.models import BlastEmailTemplate
    templates = BlastEmailTemplate.objects.all()
    return render(request, 'finance/blast_email_template_list.html', {'templates': templates})


@login_required
@tu_required
def blast_email_template_create(request):
    from payments.models import BlastEmailTemplate
    from django.http import JsonResponse
    if request.method == 'POST':
        name = request.POST.get('name')
        payment_type = request.POST.get('payment_type')
        subject = request.POST.get('subject')
        body = request.POST.get('body')
        bank_account = request.POST.get('bank_account', '')
        if not name:
            return JsonResponse({'success': False, 'error': 'Nama template wajib diisi.'})
        BlastEmailTemplate.objects.create(
            name=name, payment_type=payment_type,
            subject=subject, body=body,
            bank_account=bank_account, created_by=request.user,
        )
        return JsonResponse({'success': True})
    return redirect('finance:blast_email')


@login_required
@tu_required
def blast_email_template_edit(request, pk):
    from payments.models import BlastEmailTemplate
    template = get_object_or_404(BlastEmailTemplate, id=pk)
    if request.method == 'POST':
        template.name = request.POST.get('name')
        template.payment_type = request.POST.get('payment_type')
        template.subject = request.POST.get('subject')
        template.body = request.POST.get('body')
        template.bank_account = request.POST.get('bank_account', '')
        template.save()
        messages.success(request, 'Template berhasil diupdate.')
        return redirect('finance:blast_email_template_list')
    return redirect('finance:blast_email')


@login_required
@tu_required
def blast_email_template_delete(request, pk):
    from payments.models import BlastEmailTemplate
    template = get_object_or_404(BlastEmailTemplate, id=pk)
    template.delete()
    messages.success(request, 'Template berhasil dihapus.')
    return redirect('finance:blast_email_template_list')


@login_required
@tu_required
def blast_email_load_template(request, pk):
    from payments.models import BlastEmailTemplate
    template = get_object_or_404(BlastEmailTemplate, id=pk)
    return JsonResponse({
        'name': template.name,
        'payment_type': template.payment_type,
        'subject': template.subject,
        'body': template.body,
        'bank_account': template.bank_account,
    })


@login_required
@tu_required
def download_annual_activity_template(request, level):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Kegiatan Tahunan"

    if level == 'smp':
        headers = ['No.', 'Activities', '7/CS1', '8/CS2', '9/CS3']
    else:
        headers = ['No.', 'Activities', '10/CS4', '11/JC1', '12/JC2']

    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64 + i)].width = 22

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=annual_activity_template_{level}.xlsx'
    wb.save(response)
    return response


@login_required
@tu_required
def download_spp_va_template(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SPP VA Template"

    headers = ['Subject', 'Grade', 'virtual_account', 'customer_name', 'customer_email', 'trx_amount', 'expired_date', 'expired_time', 'description', 'link']

    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=spp_va_template.xlsx'
    wb.save(response)
    return response


@login_required
def bank_account_list(request):
    if request.user.role not in ['admin', 'tu', 'kepsek']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import BankAccount
    lf = {}
    if request.user.assigned_level:
        level_map = {'sd': ['sd'], 'smp': ['smp'], 'sma': ['sma'],
            'sd_smp': ['sd','smp'], 'smp_sma': ['smp','sma'], 'sd_smp_sma': ['sd','smp','sma']}
        allowed = level_map.get(request.user.assigned_level, [])
        lf = {'level__in': allowed}
    accounts = BankAccount.objects.filter(**lf).order_by('level')
    return render(request, 'finance/bank_account_list.html', {'accounts': accounts})


@login_required
def bank_account_create(request):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import BankAccount
    if request.method == 'POST':
        level = request.POST.get('level')
        payment_type = request.POST.get('payment_type')
        if not payment_type:
            payment_type = 'spp'
        if request.user.role == 'tu' and request.user.assigned_level:
            level_map = {'sd': ['sd'], 'smp': ['smp'], 'sma': ['sma'],
                'sd_smp': ['sd','smp'], 'smp_sma': ['smp','sma'], 'sd_smp_sma': ['sd','smp','sma']}
            if level not in level_map.get(request.user.assigned_level, []):
                messages.error(request, 'Akses ditolak untuk jenjang ini.')
                return redirect('finance:bank_account_list')
        BankAccount.objects.create(
            level=level, payment_type=payment_type,
            bank_name=request.POST['bank_name'],
            account_number=request.POST['account_number'],
            account_holder=request.POST['account_holder'],
            status='pending', is_active=False, updated_by=request.user,
        )
        messages.success(request, f'Rekening {dict(BankAccount.PAYMENT_TYPE_CHOICES).get(payment_type)} untuk {level.upper()} diajukan. Menunggu persetujuan Kepala Sekolah.')
        return redirect('finance:bank_account_list')
    return render(request, 'finance/bank_account_form.html', {'action': 'Tambah'})


@login_required
def bank_account_edit(request, pk):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import BankAccount
    account = get_object_or_404(BankAccount, id=pk)
    if request.user.role == 'tu' and request.user.assigned_level:
        level_map = {'sd': ['sd'], 'smp': ['smp'], 'sma': ['sma'],
            'sd_smp': ['sd','smp'], 'smp_sma': ['smp','sma'], 'sd_smp_sma': ['sd','smp','sma']}
        if account.level not in level_map.get(request.user.assigned_level, []):
            messages.error(request, 'Akses ditolak untuk jenjang ini.')
            return redirect('finance:bank_account_list')
    if request.method == 'POST':
        account.bank_name = request.POST['bank_name']
        account.account_number = request.POST['account_number']
        account.account_holder = request.POST['account_holder']
        account.status = 'pending'
        account.is_active = False
        account.updated_by = request.user
        account.save()
        messages.success(request, 'Rekening diperbarui. Menunggu persetujuan Kepala Sekolah.')
        return redirect('finance:bank_account_list')
    return render(request, 'finance/bank_account_form.html', {'account': account, 'action': 'Edit'})


@login_required
def bank_account_delete(request, pk):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import BankAccount
    account = get_object_or_404(BankAccount, id=pk)
    if request.user.role == 'tu' and request.user.assigned_level:
        level_map = {'sd': ['sd'], 'smp': ['smp'], 'sma': ['sma'],
            'sd_smp': ['sd','smp'], 'smp_sma': ['smp','sma'], 'sd_smp_sma': ['sd','smp','sma']}
        if account.level not in level_map.get(request.user.assigned_level, []):
            messages.error(request, 'Akses ditolak.')
            return redirect('finance:bank_account_list')
    account.delete()
    messages.success(request, 'Rekening bank berhasil dihapus.')
    return redirect('finance:bank_account_list')


@login_required
def bank_account_approve(request, pk):
    if request.user.role not in ['admin', 'kepsek']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import BankAccount
    from django.utils import timezone
    account = get_object_or_404(BankAccount, id=pk)
    if request.user.role == 'kepsek' and request.user.assigned_level and account.level != request.user.assigned_level:
        messages.error(request, 'Anda hanya dapat menyetujui rekening untuk jenjang Anda.')
        return redirect('finance:bank_account_list')
    if request.method == 'POST':
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')
        if action == 'approved':
            account.status = 'approved'
            account.is_active = True
            account.approved_by = request.user
            account.approved_at = timezone.now()
            account.notes = notes
            messages.success(request, f'Rekening {account.get_level_display()} ({account.bank_name}) disetujui.')
        elif action == 'rejected':
            account.status = 'rejected'
            account.is_active = False
            account.notes = notes
            messages.error(request, f'Rekening {account.get_level_display()} ditolak.')
        account.save()
        return redirect('finance:bank_account_list')
    return render(request, 'finance/bank_account_approve.html', {'account': account})
