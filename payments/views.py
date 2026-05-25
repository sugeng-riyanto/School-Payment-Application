from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.template.loader import render_to_string
from django.http import HttpResponse
from .models import PaymentTransaction, PaymentProof, MidtransConfig
from finance.models import SPPBill
from accounts.models import Student
import uuid
import json
import logging

logger = logging.getLogger(__name__)


def get_student(user):
    if user.role == 'parent':
        return user.children.first()
    return None


def generate_order_id():
    return f"PAY-{uuid.uuid4().hex[:12].upper()}"


@login_required
def create_payment(request, type, type_id):
    if request.user.role != 'parent':
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')

    student = get_student(request.user)
    if not student:
        messages.error(request, 'Data siswa tidak ditemukan. Hubungi Tata Usaha.')
        return redirect('accounts:dashboard')

    amount = 0
    description = ''

    if type == 'spp':
        try:
            bill = SPPBill.objects.get(id=type_id, student=student)
        except SPPBill.DoesNotExist:
            messages.error(request, 'Tagihan tidak ditemukan atau sudah berubah. Silakan cek tagihan terbaru Anda.')
            return redirect('payments:history')
        amount = bill.amount
        description = f"SPP {bill.get_month_display()} {bill.year}"
    else:
        messages.error(request, 'Tipe pembayaran tidak valid.')
        return redirect('accounts:dashboard')

    order_id = generate_order_id()
    transaction = PaymentTransaction.objects.create(
        student=student,
        transaction_type=type,
        amount=amount,
        midtrans_order_id=order_id,
        status='pending',
    )

    config = MidtransConfig.get_config()
    if config.server_key and config.client_key:
        import midtransclient
        snap = midtransclient.Snap(
            is_production=config.is_production,
            server_key=config.server_key,
        )
        param = {
            "transaction_details": {
                "order_id": order_id,
                "gross_amount": int(amount),
            },
            "credit_card": {"secure": True},
            "customer_details": {
                "first_name": student.full_name,
                "email": request.user.email,
                "phone": request.user.phone,
            }
        }
        try:
            response = snap.create_transaction(param)
            transaction.midtrans_redirect_url = response.get('redirect_url', '')
            transaction.save()
        except Exception as e:
            messages.error(request, f'Gagal membuat transaksi: {e}')
            return redirect('payments:history')
    else:
        transaction.virtual_account = f'VA-{order_id}'
        transaction.save()

    return render(request, 'payments/checkout.html', {
        'transaction': transaction,
        'client_key': config.client_key,
    })


def midtrans_callback(request):
    if request.method == 'POST':
        from django.utils import timezone
        try:
            data = json.loads(request.body)
            order_id = data.get('order_id')
            transaction_status = data.get('transaction_status')
            transaction = get_object_or_404(PaymentTransaction, midtrans_order_id=order_id)

            if transaction_status in ('settlement', 'capture'):
                transaction.status = 'success'
                transaction.paid_at = timezone.now()
                transaction.midtrans_transaction_id = data.get('transaction_id', '')
                transaction.payment_method = data.get('payment_type', '')
                transaction.save()
                if transaction.transaction_type == 'eca':
                    from eca.models import ECARegistration, ECAPayment, ECAInvoice
                    import uuid
                    if order_id.startswith('ECA-BULK-'):
                        student_id = order_id.split('-')[2]
                        unpaid = ECARegistration.objects.filter(
                            student_id=student_id, status='approved', is_paid=False
                        )
                        for reg in unpaid:
                            reg.is_paid = True
                            reg.paid_at = transaction.paid_at
                            reg.save()
                            ECAPayment.objects.create(
                                registration=reg,
                                amount=reg.program.price,
                                paid_amount=reg.program.price,
                                is_paid=True,
                                payment_method='midtrans',
                                payment_status='approved',
                                paid_at=transaction.paid_at,
                            )
                            inv_num = f"INV/ECA/{reg.id}/{uuid.uuid4().hex[:6].upper()}"
                            ECAInvoice.objects.create(
                                registration=reg,
                                student=reg.student,
                                invoice_number=inv_num,
                                amount=reg.program.price,
                                paid_at=transaction.paid_at,
                            )
                    else:
                        reg_id = order_id.split('-')[1] if '-' in order_id else None
                        if reg_id:
                            try:
                                reg = ECARegistration.objects.get(id=reg_id)
                                reg.is_paid = True
                                reg.paid_at = transaction.paid_at
                                reg.save()
                                ECAPayment.objects.create(
                                    registration=reg,
                                    amount=reg.program.price,
                                    paid_amount=reg.program.price,
                                    is_paid=True,
                                    payment_method='midtrans',
                                    payment_status='approved',
                                    paid_at=transaction.paid_at,
                                )
                                inv_num = f"INV/ECA/{reg.id}/{uuid.uuid4().hex[:6].upper()}"
                                ECAInvoice.objects.create(
                                    registration=reg,
                                    student=reg.student,
                                    invoice_number=inv_num,
                                    amount=reg.program.price,
                                    paid_at=transaction.paid_at,
                                )
                            except ECARegistration.DoesNotExist:
                                pass
            elif transaction_status in ('deny', 'cancel', 'expire'):
                transaction.status = 'failed'
                transaction.save()
        except Exception:
            pass
    return render(request, 'payments/callback_response.html')


@login_required
def upload_proof(request, bill_id):
    if request.user.role != 'parent':
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')

    student = get_student(request.user)
    if not student:
        messages.error(request, 'Data siswa tidak ditemukan.')
        return redirect('accounts:dashboard')
    try:
        bill = SPPBill.objects.get(id=bill_id, student=student)
    except SPPBill.DoesNotExist:
        messages.error(request, 'Tagihan tidak ditemukan. Silakan cek tagihan terbaru Anda.')
        return redirect('payments:history')

    if request.method == 'POST' and request.FILES.get('proof_image'):
        existing = PaymentProof.objects.filter(student=student, spp_bill=bill).first()
        if existing:
            existing.image = request.FILES['proof_image']
            existing.description = request.POST.get('description', '')
            existing.is_verified = False
            existing.save()
            msg = 'Bukti pembayaran berhasil diperbarui.'
        else:
            PaymentProof.objects.create(
                student=student,
                spp_bill=bill,
                image=request.FILES['proof_image'],
                description=request.POST.get('description', ''),
            )
            msg = 'Bukti pembayaran berhasil diupload.'
        messages.success(request, msg)
        return redirect('payments:history')

    existing = PaymentProof.objects.filter(student=student, spp_bill=bill).first()
    return render(request, 'payments/upload_proof.html', {'bill': bill, 'existing': existing})


@login_required
def upload_proof_for_transaction(request, tx_id):
    if request.user.role != 'parent':
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    student = get_student(request.user)
    if not student:
        messages.error(request, 'Data siswa tidak ditemukan.')
        return redirect('accounts:dashboard')
    tx = get_object_or_404(PaymentTransaction, id=tx_id, student=student)

    if request.method == 'POST' and request.FILES.get('proof_image'):
        existing = PaymentProof.objects.filter(student=student, transaction=tx).first()
        if existing:
            existing.image = request.FILES['proof_image']
            existing.description = request.POST.get('description', '')
            existing.is_verified = False
            existing.save()
            msg = 'Bukti pembayaran berhasil diperbarui.'
        else:
            PaymentProof.objects.create(
                student=student,
                transaction=tx,
                image=request.FILES['proof_image'],
                description=request.POST.get('description', ''),
            )
            msg = 'Bukti pembayaran berhasil diupload.'
        messages.success(request, msg)
        return redirect('payments:history')

    existing = PaymentProof.objects.filter(student=student, transaction=tx).first()
    return render(request, 'payments/upload_proof.html', {'tx': tx, 'existing': existing})


@login_required
def proof_list(request):
    messages.info(request, 'Halaman Bukti Pembayaran dipindahkan ke menu Tagihan masing-masing.')
    return redirect('payments:history')


@login_required
def verify_proof(request, proof_id):
    if request.user.role not in ['admin', 'tu']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    proof = get_object_or_404(PaymentProof, id=proof_id)
    proof.is_verified = True
    proof.save()
    from django.utils import timezone
    if proof.spp_bill:
        proof.spp_bill.is_paid = True
        proof.spp_bill.paid_at = timezone.now()
        proof.spp_bill.save()
    elif proof.transaction:
        proof.transaction.status = 'success'
        proof.transaction.paid_at = timezone.now()
        proof.transaction.save()
    messages.success(request, 'Bukti pembayaran berhasil diverifikasi.')
    return redirect('payments:proof_list')


@login_required
def midtrans_config(request):
    if request.user.role not in ['admin']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import MidtransConfig
    config = MidtransConfig.get_config()
    if request.method == 'POST':
        config.merchant_id = request.POST.get('merchant_id', config.merchant_id)
        config.client_key = request.POST.get('client_key', config.client_key)
        config.server_key = request.POST.get('server_key', config.server_key)
        config.is_production = request.POST.get('is_production') == 'on'
        config.save()
        messages.success(request, 'Konfigurasi Midtrans berhasil disimpan.')
        return redirect('payments:midtrans_config')
    from .models import MidtransConfig
    config = MidtransConfig.get_config()
    return render(request, 'payments/midtrans_config.html', {'config': config})


@login_required
def test_midtrans_connection(request):
    if request.user.role not in ['admin']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from .models import MidtransConfig
    config = MidtransConfig.get_config()
    if not config.server_key or not config.client_key:
        messages.error(request, 'Server Key atau Client Key belum diisi.')
        return redirect('payments:midtrans_config')
    import midtransclient
    import uuid
    try:
        snap = midtransclient.Snap(
            is_production=config.is_production,
            server_key=config.server_key,
        )
        test_order = f"TEST-{uuid.uuid4().hex[:8].upper()}"
        param = {
            "transaction_details": {
                "order_id": test_order,
                "gross_amount": 1000,
            },
            "credit_card": {"secure": True},
            "customer_details": {
                "first_name": "Test",
                "email": "test@example.com",
            }
        }
        response = snap.create_transaction(param)
        redirect_url = response.get('redirect_url', '')
        messages.success(request, f'Koneksi Midtrans berhasil! Order ID: {test_order}, Redirect URL tersedia.')
        logger.info(f'Midtrans test OK: order_id={test_order}, env={"production" if config.is_production else "sandbox"}')
    except Exception as e:
        error_msg = str(e)
        logger.error(f'Midtrans test FAILED: {error_msg}')
        if '401' in error_msg:
            hint = 'Server Key tidak valid. Login ke dashboard Midtrans, periksa Server Key sesuai environment (Sandbox/Production).'
        elif '403' in error_msg:
            hint = 'Akses ditolak. Periksa apakah akun Midtrans aktif.'
        else:
            hint = f'Error: {error_msg}'
        messages.error(request, f'Test gagal! {hint}')
    return redirect('payments:midtrans_config')


@login_required
def eca_payment(request, reg_id, slug=None):
    """Create Midtrans payment for single ECA registration"""
    if request.user.role != 'parent':
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from eca.models import ECARegistration
    children = request.user.children.all()
    if not children:
        messages.error(request, 'Data siswa tidak ditemukan.')
        return redirect('accounts:dashboard')
    registration = get_object_or_404(ECARegistration, id=reg_id, student__in=children)
    if registration.is_paid:
        messages.info(request, 'ECA ini sudah lunas.')
        return redirect('eca:registration_list')
    amount = registration.program.price
    order_id = f"ECA-{reg_id}-{uuid.uuid4().hex[:8].upper()}"
    transaction = PaymentTransaction.objects.create(
        student=registration.student,
        transaction_type='eca',
        amount=amount,
        midtrans_order_id=order_id,
        status='pending',
    )
    from .models import MidtransConfig
    config = MidtransConfig.get_config()
    if config.server_key and config.client_key:
        import midtransclient
        snap = midtransclient.Snap(
            is_production=config.is_production,
            server_key=config.server_key,
        )
        param = {
            "transaction_details": {
                "order_id": order_id,
                "gross_amount": int(amount),
            },
            "credit_card": {"secure": True},
            "customer_details": {
                "first_name": registration.student.full_name,
                "email": request.user.email,
                "phone": request.user.phone,
            }
        }
        try:
            response = snap.create_transaction(param)
            transaction.midtrans_redirect_url = response.get('redirect_url', '')
            transaction.snap_token = response.get('token', '')
            transaction.save()
        except Exception as e:
            messages.error(request, f'Gagal membuat transaksi Midtrans: {e}')
            return redirect('eca:registration_list')
    else:
        # If no Midtrans config, just mark as paid (offline)
        registration.is_paid = True
        registration.save()
        messages.success(request, 'Pembayaran ECA dicatat sebagai lunas (offline).')
        return redirect('eca:registration_list')

    if transaction.midtrans_redirect_url:
        return redirect(transaction.midtrans_redirect_url)
    return render(request, 'payments/eca_payment.html', {
        'transaction': transaction,
        'registration': registration,
        'snap_token': transaction.snap_token,
    })


@login_required
def eca_bulk_payment(request):
    """Bulk payment for all unpaid approved ECA registrations"""
    if request.user.role != 'parent':
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from eca.models import ECARegistration
    students = request.user.children.all()
    if not students:
        messages.error(request, 'Data siswa tidak ditemukan.')
        return redirect('accounts:dashboard')

    unpaid = ECARegistration.objects.filter(student__in=students, status='approved', is_paid=False)
    if not unpaid.exists():
        messages.info(request, 'Tidak ada tagihan ECA yang perlu dibayar.')
        return redirect('eca:registration_list')

    total = sum(r.program.price for r in unpaid)
    reg_ids = '_'.join(str(r.id) for r in unpaid)
    student_ids = '_'.join(str(s.nisn) for s in students)
    order_id = f"ECA-BULK-{student_ids}-{uuid.uuid4().hex[:8].upper()}"

    transaction = PaymentTransaction.objects.create(
        student=unpaid.first().student,
        transaction_type='eca',
        amount=total,
        midtrans_order_id=order_id,
        status='pending',
    )

    from .models import MidtransConfig
    config = MidtransConfig.get_config()
    if config.server_key and config.client_key:
        import midtransclient
        snap = midtransclient.Snap(
            is_production=config.is_production,
            server_key=config.server_key,
        )
        first_student = unpaid.first().student
        param = {
            "transaction_details": {
                "order_id": order_id,
                "gross_amount": int(total),
            },
            "credit_card": {"secure": True},
            "customer_details": {
                "first_name": first_student.full_name,
                "email": request.user.email,
                "phone": request.user.phone,
            },
            "item_details": [
                {
                    "id": str(r.id),
                    "price": int(r.program.price),
                    "quantity": 1,
                    "name": r.program.name[:50],
                }
                for r in unpaid
            ],
        }
        try:
            response = snap.create_transaction(param)
            transaction.midtrans_redirect_url = response.get('redirect_url', '')
            transaction.snap_token = response.get('token', '')
            transaction.save()
        except Exception as e:
            messages.error(request, f'Gagal membuat transaksi Midtrans: {e}')
            return redirect('eca:registration_list')
    else:
        messages.error(request, 'Konfigurasi Midtrans belum diatur oleh Admin.')
        return redirect('eca:registration_list')

    return render(request, 'payments/eca_checkout.html', {
        'transaction': transaction,
        'client_key': config.client_key,
    })


@login_required
def payment_history(request):
    from eca.models import ECARegistration
    from finance.models import CambridgeAssessment
    if request.user.role == 'parent':
        children = request.user.children.all()
        if not children:
            messages.error(request, 'Data siswa tidak ditemukan.')
            return redirect('accounts:dashboard')
        transactions = PaymentTransaction.objects.filter(student__in=children)
        spp_bills = SPPBill.objects.filter(student__in=children).order_by('-year', '-month')

        spp_tx = transactions.filter(transaction_type='spp')
        eca_tx = transactions.filter(transaction_type='eca')
        kegiatan_tx = transactions.filter(transaction_type='kegiatan')
        dpp_tx = transactions.filter(transaction_type='dpp')
        cam_tx = transactions.filter(transaction_type='cambridge')

        eca_regs = ECARegistration.objects.filter(student__in=children, is_paid=False, status='approved').select_related('program')
        cam_unpaid = CambridgeAssessment.objects.filter(student__in=children, is_paid=False)

        proofs = PaymentProof.objects.filter(student__in=children)
        proofed_bills = set(proofs.filter(spp_bill__isnull=False).values_list('spp_bill_id', flat=True))
        proofed_txs = set(proofs.filter(transaction__isnull=False).values_list('transaction_id', flat=True))
        verified_bills = set(proofs.filter(spp_bill__isnull=False, is_verified=True).values_list('spp_bill_id', flat=True))
        verified_txs = set(proofs.filter(transaction__isnull=False, is_verified=True).values_list('transaction_id', flat=True))

        # Map bill_id -> transaction id (for Invoice Penerimaan after TU confirms)
        bill_tx_map = {}
        spp_txs = PaymentTransaction.objects.filter(
            student__in=children, transaction_type='spp', status='success'
        ).values('id', 'student_id', 'amount')
        for bill in spp_bills:
            matching = [t for t in spp_txs if t['student_id'] == bill.student_id and t['amount'] == bill.amount]
            if matching:
                bill_tx_map[bill.id] = matching[0]['id']

        # Map bill_id -> proof image URL for preview column
        proof_images = {}
        for p in proofs.filter(spp_bill__isnull=False).exclude(image=''):
            if p.image:
                try:
                    proof_images[p.spp_bill_id] = p.image.url
                except:
                    pass

        # Bank info per level — from database (configurable by admin/TU)
        first_student = children.first()
        level = first_student.class_grade.grade.level if first_student and first_student.class_grade else 'sma'
        from finance.models import BankAccount
        bank = BankAccount.objects.filter(level=level, payment_type='spp', is_active=True).first()
        if bank:
            ctx_extra = {'bank_name': bank.bank_name, 'bank_account': bank.account_number, 'bank_owner': bank.account_holder, 'unpaid_va_bills': spp_bills.filter(is_paid=False, virtual_account__gt='')[:10]}
        else:
            ctx_extra = {'bank_name': '', 'bank_account': '', 'bank_owner': '', 'unpaid_va_bills': spp_bills.filter(is_paid=False, virtual_account__gt='')[:10]}
    else:
        transactions = PaymentTransaction.objects.none()
        spp_bills = SPPBill.objects.none()
        spp_tx = eca_tx = kegiatan_tx = dpp_tx = cam_tx = PaymentTransaction.objects.none()
        eca_regs = ECARegistration.objects.none()
        cam_unpaid = CambridgeAssessment.objects.none()
        proofed_bills = set()
        proofed_txs = set()
        verified_bills = set()
        verified_txs = set()
        ctx_extra = {'bank_name': '', 'bank_account': '', 'bank_owner': '', 'unpaid_va_bills': SPPBill.objects.none()}
        proof_images = {}
        bill_tx_map = {}

    return render(request, 'payments/history.html', {
        'transactions': transactions,
        'spp_bills': spp_bills,
        'spp_tx': spp_tx,
        'eca_tx': eca_tx,
        'kegiatan_tx': kegiatan_tx,
        'dpp_tx': dpp_tx,
        'cam_tx': cam_tx,
        'eca_regs': eca_regs,
        'cam_unpaid': cam_unpaid,
        'proofed_bills': proofed_bills,
        'proofed_txs': proofed_txs,
        'verified_bills': verified_bills,
        'verified_txs': verified_txs,
        **ctx_extra,
        'proof_images': proof_images,
        'bill_tx_map': bill_tx_map,
    })


@login_required
def cambridge_payment(request, ca_id):
    if request.user.role != 'parent':
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from finance.models import CambridgeAssessment
    student = get_student(request.user)
    if not student:
        messages.error(request, 'Data siswa tidak ditemukan.')
        return redirect('accounts:dashboard')
    ca = get_object_or_404(CambridgeAssessment, id=ca_id, student=student)
    if ca.is_paid:
        messages.info(request, 'Cambridge Assessment ini sudah lunas.')
        return redirect('payments:history')
    order_id = f"CAM-{ca.id}-{uuid.uuid4().hex[:8].upper()}"
    transaction = PaymentTransaction.objects.create(
        student=student,
        transaction_type='cambridge',
        amount=ca.amount,
        midtrans_order_id=order_id,
        status='pending',
        virtual_account=f'VA-MID-{student.nisn[-6:]}',
    )
    config = MidtransConfig.get_config()
    if config.server_key and config.client_key:
        import midtransclient
        snap = midtransclient.Snap(
            is_production=config.is_production,
            server_key=config.server_key,
        )
        param = {
            "transaction_details": {
                "order_id": order_id,
                "gross_amount": int(ca.amount),
            },
            "credit_card": {"secure": True},
            "customer_details": {
                "first_name": student.full_name,
                "email": request.user.email,
                "phone": request.user.phone,
            }
        }
        try:
            response = snap.create_transaction(param)
            transaction.midtrans_redirect_url = response.get('redirect_url', '')
            transaction.snap_token = response.get('token', '')
            transaction.invoice_number = f'INV/CAM/{student.nisn}/{uuid.uuid4().hex[:4].upper()}'
            transaction.save()
        except Exception as e:
            messages.error(request, f'Gagal membuat transaksi Midtrans: {e}')
            return redirect('payments:history')
    else:
        transaction.invoice_number = f'INV/CAM/{student.nisn}/{uuid.uuid4().hex[:4].upper()}'
        transaction.save()

    from django.utils import timezone
    ca.is_paid = True
    ca.paid_at = timezone.now()
    ca.payment_method = 'midtrans'
    ca.save()

    messages.success(request, 'Pembayaran Cambridge Assessment berhasil diproses.')
    return redirect('payments:history')


@login_required
def download_invoice(request, pk):
    from django.template.loader import render_to_string
    from xhtml2pdf import pisa
    from public_info.models import SchoolProfile
    import qrcode
    import io, base64
    from django.utils import timezone
    tx = get_object_or_404(PaymentTransaction, id=pk)
    user = request.user
    if user.role == 'parent':
        if tx.student not in user.children.all():
            messages.error(request, 'Akses ditolak.')
            return redirect('accounts:dashboard')
    elif user.role in ['kepsek', 'vp_activity', 'tu'] and user.assigned_level:
        level_map = {
            'sd': ['sd'], 'smp': ['smp'], 'sma': ['sma'],
            'sd_smp': ['sd','smp'], 'smp_sma': ['smp','sma'],
            'sd_smp_sma': ['sd','smp','sma'],
        }
        allowed = level_map.get(user.assigned_level, [])
        if tx.student.class_grade.grade.level not in allowed:
            messages.error(request, 'Akses ditolak.')
            return redirect('accounts:dashboard')
    school = SchoolProfile.get_profile()

    # Generate QR code
    qr_data = f"INV:{tx.invoice_number or tx.id}|CODE:{tx.verification_code}"
    qr = qrcode.make(qr_data)
    buf = io.BytesIO()
    qr.save(buf, format='PNG')
    qr_b64 = base64.b64encode(buf.getvalue()).decode()

    portal = request.build_absolute_uri('/')
    html = render_to_string('payments/invoice_pdf.html', {
        'tx': tx, 'school': school, 'qr_b64': qr_b64, 'portal_url': portal,
        'downloaded_by': request.user.get_full_name() or request.user.username,
        'downloaded_at': timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M'),
    })
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="invoice_{tx.invoice_number or tx.id}.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Gagal membuat PDF', status=500)
    return response


@login_required
def download_bill_invoice(request, bill_id):
    """Generate PDF invoice for an SPP bill with QR validation."""
    from django.template.loader import render_to_string
    from xhtml2pdf import pisa
    from public_info.models import SchoolProfile
    import qrcode, io, base64, uuid
    from finance.models import SPPBill
    from django.utils import timezone

    bill = get_object_or_404(SPPBill, id=bill_id)
    user = request.user
    if user.role == 'parent':
        if bill.student not in user.children.all():
            messages.error(request, 'Akses ditolak.')
            return redirect('accounts:dashboard')

    school = SchoolProfile.get_profile()
    import hashlib
    # Store or reuse verification code on the bill
    if not bill.verification_code:
        bill.verification_code = hashlib.sha256(f"spp-bill-{bill.id}-{bill.student.nis}".encode()).hexdigest()[:12].upper()
        bill.save(update_fields=['verification_code'])
    verify_code = bill.verification_code
    inv_number = f"SPP/{bill.student.class_grade.grade.level.upper()}/{bill.student.nis}/{bill.year}{bill.month:02d}"

    qr_data = f"INV:{inv_number}|CODE:{verify_code}"
    qr = qrcode.make(qr_data)
    buf = io.BytesIO()
    qr.save(buf, format='PNG')
    qr_b64 = base64.b64encode(buf.getvalue()).decode()

    portal = request.build_absolute_uri('/')
    html = render_to_string('payments/bill_invoice_pdf.html', {
        'bill': bill, 'school': school, 'qr_b64': qr_b64,
        'portal_url': portal, 'inv_number': inv_number, 'verify_code': verify_code,
        'downloaded_by': request.user.get_full_name() or request.user.username,
        'downloaded_at': timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M'),
    })
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="tagihan_spp_{bill.student.nis}_{bill.year}_{bill.month}.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Gagal membuat PDF', status=500)
    return response


@login_required
def verify_invoice(request):
    if request.user.role not in ['admin', 'tu', 'kepsek', 'vp_activity']:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    result = None
    code = ''
    if request.method == 'POST':
        code = request.POST.get('code', '').strip().upper()
        tx = PaymentTransaction.objects.filter(verification_code=code).first()
        if tx:
            result = tx
        else:
            messages.error(request, 'Kode verifikasi tidak ditemukan.')
    return render(request, 'payments/verify_invoice.html', {'result': result, 'code': code})


@login_required
def download_va(request, tx_id):
    from django.template.loader import render_to_string
    from xhtml2pdf import pisa
    from django.utils import timezone
    tx = get_object_or_404(PaymentTransaction, id=tx_id)
    user = request.user

    if user.role == 'parent':
        if tx.student not in user.children.all():
            messages.error(request, 'Akses ditolak.')
            return redirect('accounts:dashboard')
    elif user.role == 'admin':
        pass
    elif user.role in ['kepsek', 'vp_activity', 'tu'] and user.assigned_level:
        level_map = {'sd': ['sd'], 'smp': ['smp'], 'sma': ['sma'],
            'sd_smp': ['sd','smp'], 'smp_sma': ['smp','sma'], 'sd_smp_sma': ['sd','smp','sma']}
        allowed = level_map.get(user.assigned_level, [])
        if tx.student.class_grade.grade.level not in allowed:
            messages.error(request, 'Akses ditolak.')
            return redirect('accounts:dashboard')
    else:
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')

    from public_info.models import SchoolProfile
    school = SchoolProfile.get_profile()
    import qrcode, io, base64
    qr_data = f"VA:{tx.invoice_number or tx.id}|CODE:{tx.verification_code}"
    qr = qrcode.make(qr_data)
    buf = io.BytesIO()
    qr.save(buf, format='PNG')
    qr_b64 = base64.b64encode(buf.getvalue()).decode()
    portal = request.build_absolute_uri('/')
    html = render_to_string('payments/va_pdf.html', {
        'tx': tx, 'school': school, 'qr_b64': qr_b64, 'portal_url': portal,
        'downloaded_by': request.user.get_full_name() or request.user.username,
        'downloaded_at': timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M'),
    })
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="VA_{tx.invoice_number or tx.id}.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Gagal membuat PDF', status=500)
    return response


@login_required
def export_history_xlsx(request):
    if request.user.role != 'parent':
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from finance.models import SPPBill
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.utils import timezone
    import hashlib
    children = request.user.children.all()
    spp_bills = SPPBill.objects.filter(student__in=children).order_by('-year', '-month')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Riwayat Pembayaran"
    
    # User identity
    ws.cell(row=1, column=1, value='Nama Orang Tua:').font = Font(bold=True)
    ws.cell(row=1, column=2, value=request.user.get_full_name() or request.user.username)
    ws.cell(row=2, column=1, value='Email:').font = Font(bold=True)
    ws.cell(row=2, column=2, value=request.user.email or '-')
    ws.cell(row=3, column=1, value='Tanggal Export:').font = Font(bold=True)
    ws.cell(row=3, column=2, value=timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M'))
    
    # Generate a stable verification code for the export
    export_code = hashlib.sha256(f'export-{request.user.id}-{timezone.localtime(timezone.now()).strftime("%Y%m%d")}'.encode()).hexdigest()[:12].upper()
    ws.cell(row=4, column=1, value='Kode Verifikasi:').font = Font(bold=True)
    ws.cell(row=4, column=2, value=export_code)
    
    # Header row
    headers = ['No', 'Nama Siswa', 'NIS', 'Kelas', 'Bulan', 'Tahun', 'Jumlah', 'Virtual Account', 'Status', 'Invoice']
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for i, bill in enumerate(spp_bills, 1):
        inv = f"SPP/{bill.student.class_grade.grade.level.upper()}/{bill.student.nis}/{bill.year}{bill.month:02d}"
        row = i + 6
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=bill.student.full_name)
        ws.cell(row=row, column=3, value=bill.student.nis)
        ws.cell(row=row, column=4, value=str(bill.student.class_grade))
        ws.cell(row=row, column=5, value=bill.get_month_display())
        ws.cell(row=row, column=6, value=bill.year)
        ws.cell(row=row, column=7, value=int(bill.amount))
        ws.cell(row=row, column=8, value=bill.virtual_account or '-')
        ws.cell(row=row, column=9, value='Lunas' if bill.is_paid else 'Belum')
        ws.cell(row=row, column=10, value=inv)
    
    # Auto-width
    for col in range(1, 11):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=riwayat_pembayaran_{timezone.localtime(timezone.now()).strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
def export_history_pdf(request):
    if request.user.role != 'parent':
        messages.error(request, 'Akses ditolak.')
        return redirect('accounts:dashboard')
    from finance.models import SPPBill
    from django.template.loader import render_to_string
    from xhtml2pdf import pisa
    from django.utils import timezone
    from public_info.models import SchoolProfile
    children = request.user.children.all()
    spp_bills = SPPBill.objects.filter(student__in=children).order_by('-year', '-month')
    
    school = SchoolProfile.get_profile()
    
    rows = []
    for i, bill in enumerate(spp_bills, 1):
        inv = f"SPP/{bill.student.class_grade.grade.level.upper()}/{bill.student.nis}/{bill.year}{bill.month:02d}"
        rows.append({
            'no': i,
            'student': bill.student.full_name,
            'nis': bill.student.nis,
            'kelas': str(bill.student.class_grade),
            'bulan': bill.get_month_display(),
            'tahun': str(bill.year),
            'jumlah': f'Rp{int(bill.amount):,}',
            'va': bill.virtual_account or '-',
            'status': 'Lunas' if bill.is_paid else 'Belum',
            'invoice': inv,
        })
    
    # Determine jenjang from student data
    first_student = children.first()
    level_name = first_student.class_grade.grade.get_level_display() if first_student and first_student.class_grade else 'Semua Jenjang'
    
    html = render_to_string('payments/pdf_history.html', {
        'rows': rows,
        'title': f'Riwayat Pembayaran SPP - {level_name}',
        'headers': ['No', 'Nama', 'NIS', 'Kelas', 'Bulan', 'Tahun', 'Jumlah', 'VA', 'Status', 'Invoice'],
        'generated_at': timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M'),
        'user_name': request.user.get_full_name() or request.user.username,
        'user_email': request.user.email or '-',
        'school': school,
        'level_name': level_name,
    })
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=riwayat_pembayaran_{timezone.localtime(timezone.now()).strftime("%Y%m%d")}.pdf'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Gagal membuat PDF', status=500)
    return response
